# routes.py - Kinessia Hub v2.0
# Rutas actualizadas con sistema de tarjetas corporativas, autorizaciones y gestión de usuarios

from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from flask_login import login_required, current_user
from .models import (
    db, Usuario, Rol, Papeleta, Desglose, Empresa, Aerolinea, EmpresaBooking, 
    CargoServicio, Descuento, TarifaFija, Sucursal, TarjetaCorporativa, Autorizacion,
    TarjetaUsuario, AuditLog, ReporteVenta, DetalleReporteVenta, EntregaCorte, 
    DetalleArqueo, HistorialEntrega, crear_entrega_desde_reporte, Notificacion
)
from datetime import datetime, timedelta, date
from sqlalchemy import func
from collections import OrderedDict
from functools import wraps


# Importar servicio de notificaciones (con manejo de error si no existe aún)
try:
    from app.services.notificaciones import NotificacionService, obtener_notificaciones_pendientes
except ImportError:
    NotificacionService = None
    obtener_notificaciones_pendientes = None

# Zona horaria de México (Tijuana/Ensenada)
from zoneinfo import ZoneInfo
TIMEZONE_MX = ZoneInfo('America/Tijuana')

def fecha_mexico():
    """Retorna la fecha actual en zona horaria de México (Tijuana)"""
    return datetime.now(TIMEZONE_MX).date()


main = Blueprint('main', __name__)


# Context processor para inyectar notificaciones a todos los templates
@main.app_context_processor
def inject_notificaciones():
    """Inyecta el conteo de autorizaciones pendientes a todos los templates"""
    if current_user.is_authenticated and current_user.rol in ['director', 'administrador', 'admin']:
        try:
            pendientes = Autorizacion.query.filter_by(estatus='pendiente').count()
            return {'autorizaciones_pendientes': pendientes}
        except:
            return {'autorizaciones_pendientes': 0}
    return {'autorizaciones_pendientes': 0}

# =============================================================================
# RUTAS PRINCIPALES
# =============================================================================

@main.route('/')
@login_required
def index():
    return redirect(url_for('main.dashboard'))

# ============================================================
# REEMPLAZA tu función dashboard() en routes.py con esto:
# Esta versión incluye manejo de errores de transacción
# ============================================================

@main.route('/dashboard')
@login_required
def dashboard():
    """Dashboard unificado con manejo de errores"""
    from datetime import date, timedelta
    
    # IMPORTANTE: Limpiar cualquier transacción fallida anterior
    try:
        db.session.rollback()
    except:
        pass
    
    fecha_hoy = fecha_mexico()
    
    # Variables con valores por defecto
    context = {
        'fecha_hoy': fecha_hoy,
        'usuario': current_user,
        'autorizaciones_pendientes': 0,
        'mis_papeletas_hoy': 0,
        'mi_total_hoy': 0,
        'mi_efectivo_hoy': 0,
        'mis_papeletas_pendientes': [],
        'papeletas_pendientes': 0,
        'papeletas_urgentes': 0,
        'mi_efectivo_pendiente': 0,
        'mis_reportes_mes': 0,
        # Para admin
        'total_papeletas_pendientes': 0,
        'total_efectivo_pendiente': 0,
        'entregas_por_recibir': 0,
        'entregas_pendientes': [],
        'reportes_por_revisar': 0,
        'papeletas_hoy_total': 0,
        'total_ventas_hoy': 0,
        'total_efectivo_hoy': 0,
        'resumen_agentes': [],
    }
    
    try:
        # Autorizaciones pendientes (para admin)
        if current_user.rol in ['director', 'administrador', 'admin']:
            context['autorizaciones_pendientes'] = Autorizacion.query.filter_by(estatus='pendiente').count()
    except Exception as e:
        print(f"Error autorizaciones: {e}")
        db.session.rollback()
    
    try:
        # Papeletas del usuario actual - HOY
        context['mis_papeletas_hoy'] = Papeleta.query.filter(
            Papeleta.usuario_id == current_user.id,
            Papeleta.fecha_venta == fecha_hoy
        ).count()
    except Exception as e:
        print(f"Error mis_papeletas_hoy: {e}")
        db.session.rollback()
    
    try:
        mi_total = db.session.query(func.sum(Papeleta.total)).filter(
            Papeleta.usuario_id == current_user.id,
            Papeleta.fecha_venta == fecha_hoy
        ).scalar()
        context['mi_total_hoy'] = float(mi_total or 0)
    except Exception as e:
        print(f"Error mi_total_hoy: {e}")
        db.session.rollback()
    
    try:
        mi_efectivo = db.session.query(func.sum(Papeleta.total)).filter(
            Papeleta.usuario_id == current_user.id,
            Papeleta.fecha_venta == fecha_hoy,
            db.or_(
                Papeleta.forma_pago.ilike('%contado%'),
                Papeleta.forma_pago.ilike('%efectivo%')
            )
        ).scalar()
        context['mi_efectivo_hoy'] = float(mi_efectivo or 0)
    except Exception as e:
        print(f"Error mi_efectivo_hoy: {e}")
        db.session.rollback()
    
    try:
        # Papeletas pendientes = MOSTRADOR sin reporte Y sin factura
        mis_pendientes = Papeleta.query.filter(
            Papeleta.usuario_id == current_user.id,
            Papeleta.fecha_venta >= fecha_hoy - timedelta(days=30),
            Papeleta.reporte_venta_id.is_(None),
            db.or_(
                Papeleta.numero_factura.is_(None),
                Papeleta.numero_factura == ''
            ),
            db.or_(
                Papeleta.empresa_id.is_(None),
                Papeleta.facturar_a.ilike('%MOSTRADOR%')
            )
        ).order_by(Papeleta.fecha_venta.desc()).all()
        
        for p in mis_pendientes:
            p.dias = (fecha_hoy - p.fecha_venta).days if p.fecha_venta else 0
        
        context['mis_papeletas_pendientes'] = mis_pendientes
        context['papeletas_pendientes'] = len(mis_pendientes)
        context['papeletas_urgentes'] = len([p for p in mis_pendientes if p.dias > 3])
        context['mi_efectivo_pendiente'] = sum([
            float(p.total or 0) for p in mis_pendientes 
            if p.forma_pago and ('contado' in p.forma_pago.lower() or 'efectivo' in p.forma_pago.lower())
        ])
    except Exception as e:
        print(f"Error papeletas pendientes: {e}")
        db.session.rollback()
    
    try:
        # Reportes del mes
        primer_dia_mes = fecha_hoy.replace(day=1)
        context['mis_reportes_mes'] = ReporteVenta.query.filter(
            ReporteVenta.usuario_id == current_user.id,
            ReporteVenta.fecha >= primer_dia_mes
        ).count()
    except Exception as e:
        print(f"Error reportes mes: {e}")
        db.session.rollback()
    
    # ============================================================
    # DATOS PARA ADMIN
    # ============================================================
    
    if current_user.es_admin():
        try:
            # Papeletas pendientes = sin reporte Y sin factura
            context['total_papeletas_pendientes'] = Papeleta.query.filter(
                Papeleta.fecha_venta < fecha_hoy,
                Papeleta.fecha_venta >= fecha_hoy - timedelta(days=30),
                Papeleta.reporte_venta_id.is_(None),
                db.or_(
                    Papeleta.numero_factura.is_(None),
                    Papeleta.numero_factura == ''
                )
            ).count()
        except Exception as e:
            print(f"Error total pendientes: {e}")
            db.session.rollback()
        
        try:
            # Efectivo pendiente = papeletas en CONTADO sin reporte
            total_efec = db.session.query(func.sum(Papeleta.total)).filter(
                Papeleta.fecha_venta < fecha_hoy,
                Papeleta.fecha_venta >= fecha_hoy - timedelta(days=30),
                Papeleta.reporte_venta_id.is_(None),
                db.or_(
                    Papeleta.forma_pago.ilike('%contado%'),
                    Papeleta.forma_pago.ilike('%efectivo%'),
                    Papeleta.forma_pago == 'Contado'
                )
            ).scalar()
            context['total_efectivo_pendiente'] = float(total_efec or 0)
        except Exception as e:
            print(f"Error total efectivo: {e}")
            db.session.rollback()
        
        try:
            context['entregas_por_recibir'] = EntregaCorte.query.filter(
                EntregaCorte.estatus.in_(['pendiente', 'entregado'])
            ).count()
            
            context['entregas_pendientes'] = EntregaCorte.query.filter(
                EntregaCorte.estatus.in_(['pendiente', 'entregado', 'en_custodia'])
            ).order_by(EntregaCorte.fecha.desc()).limit(5).all()
        except Exception as e:
            print(f"Error entregas: {e}")
            db.session.rollback()
        
        try:
            context['reportes_por_revisar'] = ReporteVenta.query.filter(
                ReporteVenta.estatus == 'enviado'
            ).count()
        except Exception as e:
            print(f"Error reportes revisar: {e}")
            db.session.rollback()
        
        try:
            context['papeletas_hoy_total'] = Papeleta.query.filter(
                Papeleta.fecha_venta == fecha_hoy
            ).count()
            
            total_hoy = db.session.query(func.sum(Papeleta.total)).filter(
                Papeleta.fecha_venta == fecha_hoy
            ).scalar()
            context['total_ventas_hoy'] = float(total_hoy or 0)
            
            efec_hoy = db.session.query(func.sum(Papeleta.total)).filter(
                Papeleta.fecha_venta == fecha_hoy,
                db.or_(
                    Papeleta.forma_pago.ilike('%contado%'),
                    Papeleta.forma_pago.ilike('%efectivo%'),
                    Papeleta.forma_pago == 'Contado'
                )
            ).scalar()
            context['total_efectivo_hoy'] = float(efec_hoy or 0)
        except Exception as e:
            print(f"Error papeletas hoy: {e}")
            db.session.rollback()
        
        try:
            # Resumen por agente - incluye todos los agentes + usuario actual
            agentes = Usuario.query.filter(
                Usuario.activo == True,
                db.or_(
                    # Agentes (roles operativos)
                    Usuario.rol.in_(['agente', 'mostrador', 'vendedor', 'ejecutivo']),
                    # O el usuario actual (aunque sea admin)
                    Usuario.id == current_user.id
                )
            ).order_by(Usuario.nombre).all()
            
            resumen = []
            for agente in agentes:
                try:
                    paps_hoy = Papeleta.query.filter(
                        Papeleta.usuario_id == agente.id,
                        Papeleta.fecha_venta == fecha_hoy
                    ).count()
                    
                    # Pendientes = sin reporte Y sin factura
                    paps_pend = Papeleta.query.filter(
                        Papeleta.usuario_id == agente.id,
                        Papeleta.fecha_venta < fecha_hoy,
                        Papeleta.fecha_venta >= fecha_hoy - timedelta(days=30),
                        Papeleta.reporte_venta_id.is_(None),
                        db.or_(
                            Papeleta.numero_factura.is_(None),
                            Papeleta.numero_factura == ''
                        )
                    ).count()
                    
                    # Efectivo pendiente del agente (sin reporte) - CONTADO
                    efec = db.session.query(func.sum(Papeleta.total)).filter(
                        Papeleta.usuario_id == agente.id,
                        Papeleta.fecha_venta < fecha_hoy,
                        Papeleta.fecha_venta >= fecha_hoy - timedelta(days=30),
                        Papeleta.reporte_venta_id.is_(None),
                        db.or_(
                            Papeleta.forma_pago.ilike('%contado%'),
                            Papeleta.forma_pago.ilike('%efectivo%'),
                            Papeleta.forma_pago == 'Contado'
                        )
                    ).scalar()
                    
                    resumen.append({
                        'id': agente.id,
                        'agente': agente.nombre,
                        'papeletas_hoy': paps_hoy,
                        'pendientes': paps_pend,
                        'efectivo': float(efec or 0)
                    })
                except:
                    db.session.rollback()
            
            resumen.sort(key=lambda x: (-x['pendientes'], x['agente']))
            context['resumen_agentes'] = resumen
        except Exception as e:
            print(f"Error resumen agentes: {e}")
            db.session.rollback()
        
        # ============================================================
        # DATOS DE FACTURACIÓN PARA DASHBOARD
        # ============================================================
        try:
            # Papeletas pendientes de facturar (Low Cost con empresa)
            context['papeletas_facturacion_pendientes'] = Papeleta.query.filter(
                Papeleta.empresa_id.isnot(None),
                Papeleta.numero_factura.is_(None)
            ).count()
        except Exception as e:
            print(f"Error papeletas facturacion: {e}")
            db.session.rollback()
            context['papeletas_facturacion_pendientes'] = 0
        
        try:
            # Desgloses BSP pendientes de facturar
            desgloses_bsp = Desglose.query.join(Aerolinea).filter(
                Desglose.empresa_id.isnot(None),
                db.or_(
                    Desglose.numero_factura.is_(None),
                    Desglose.estatus_facturacion.in_(['pendiente', None])
                ),
                db.or_(
                    Aerolinea.es_bsp == True,
                    Aerolinea.nombre.ilike('%aeromexico%'),
                    Aerolinea.nombre.ilike('%american%'),
                    Aerolinea.nombre.ilike('%united%'),
                    Aerolinea.nombre.ilike('%delta%')
                )
            ).count()
            context['desgloses_bsp_pendientes'] = desgloses_bsp
        except Exception as e:
            print(f"Error desgloses BSP: {e}")
            db.session.rollback()
            context['desgloses_bsp_pendientes'] = 0
        
        try:
            # Total facturación pendiente
            context['facturacion_pendiente'] = context.get('papeletas_facturacion_pendientes', 0) + context.get('desgloses_bsp_pendientes', 0)
        except:
            context['facturacion_pendiente'] = 0
        
        try:
            # Desgloses de hoy
            context['desgloses_hoy'] = Desglose.query.filter(
                Desglose.fecha_emision == fecha_hoy
            ).count()
        except Exception as e:
            print(f"Error desgloses hoy: {e}")
            context['desgloses_hoy'] = 0
        
        try:
            # Últimas facturas registradas
            context['ultimas_facturas'] = Papeleta.query.filter(
                Papeleta.numero_factura.isnot(None)
            ).order_by(Papeleta.fecha_facturacion.desc()).limit(5).all()
        except Exception as e:
            print(f"Error ultimas facturas: {e}")
            context['ultimas_facturas'] = []

    # Papeletas pendientes del usuario

    papeletas_pendientes = Papeleta.query.filter(
        Papeleta.usuario_id == current_user.id,
        Papeleta.fecha_venta < fecha_mexico(),
        Papeleta.estatus_control.in_(['activa', None]),
        Papeleta.justificacion_pendiente.is_(None)
    ).count()        
    papeletas_pendientes=papeletas_pendientes
    return render_template('dashboard.html', **context)
  
# =============================================================================
# API ENDPOINTS
# =============================================================================

@main.route('/api/siguiente-folio-desglose')
@login_required
def siguiente_folio_desglose():
    ultimo_folio = db.session.query(func.max(Desglose.folio)).scalar()
    siguiente = (ultimo_folio or 0) + 1
    return jsonify({'folio': siguiente})


@main.route('/api/siguiente-folio-papeleta')
@login_required
def siguiente_folio_papeleta():
    tarjeta = request.args.get('tarjeta', '')
    tarjeta_id = request.args.get('tarjeta_id', '')

    if tarjeta_id:
        tarjeta_obj = TarjetaCorporativa.query.get(tarjeta_id)
        if tarjeta_obj:
            tarjeta = tarjeta_obj.numero_tarjeta

    if not tarjeta or len(tarjeta) < 2:
        return jsonify({'error': 'Tarjeta inválida'}), 400

    ultima_papeleta = Papeleta.query.filter(
        Papeleta.folio.like(f"{tarjeta}-%")
    ).order_by(Papeleta.id.desc()).first()

    if ultima_papeleta:
        try:
            ultimo_numero = int(ultima_papeleta.folio.split('-')[1])
            siguiente = ultimo_numero + 1
        except:
            siguiente = 1
    else:
        siguiente = 1

    folio = f"{tarjeta}-{siguiente:03d}"
    return jsonify({'folio': folio, 'numero': siguiente})


@main.route('/api/cargos-empresa/<int:empresa_id>')
@login_required
def cargos_empresa(empresa_id):
    """
    Obtiene los cargos por servicio de una empresa.
    Parámetro opcional: tipo_servicio (nacional, internacional, hotel, auto, otro)
    """
    empresa = Empresa.query.get_or_404(empresa_id)
    
    # Obtener tipo de servicio del query string (default: nacional)
    tipo_servicio = request.args.get('tipo_servicio', 'nacional')
    
    # Mapear tipos de cargo del formulario a tipos de servicio de la BD
    tipo_map = {
        'aerolinea': 'nacional',  # Por defecto nacional, puede ser internacional
        'hotel': 'hotel',
        'auto': 'auto',
        'otro': 'otro',
        'nacional': 'nacional',
        'internacional': 'internacional'
    }
    
    tipo_servicio_bd = tipo_map.get(tipo_servicio, 'nacional')
    
    # Buscar cargos para este tipo de servicio
    cargo_visible = CargoServicio.query.filter_by(
        empresa_id=empresa_id,
        tipo='visible',
        tipo_servicio=tipo_servicio_bd,
        activo=True
    ).first()
    
    cargo_oculto = CargoServicio.query.filter_by(
        empresa_id=empresa_id,
        tipo='oculto',
        tipo_servicio=tipo_servicio_bd,
        activo=True
    ).first()
    
    # Si no hay cargos para el tipo específico, buscar nacional como fallback
    if not cargo_visible and not cargo_oculto and tipo_servicio_bd != 'nacional':
        cargo_visible = CargoServicio.query.filter_by(
            empresa_id=empresa_id,
            tipo='visible',
            tipo_servicio='nacional',
            activo=True
        ).first()
        
        cargo_oculto = CargoServicio.query.filter_by(
            empresa_id=empresa_id,
            tipo='oculto',
            tipo_servicio='nacional',
            activo=True
        ).first()
    
    return jsonify({
        'empresa_id': empresa_id,
        'empresa_nombre': empresa.nombre_empresa,
        'tipo_servicio': tipo_servicio_bd,
        'cargo_visible': float(cargo_visible.monto) if cargo_visible else 0,
        'cargo_oculto': float(cargo_oculto.monto) if cargo_oculto else 0
    })


@main.route('/api/papeleta/<int:id>')
@login_required
def api_papeleta_detalle(id):
    """API para obtener el detalle de una papeleta."""
    papeleta = Papeleta.query.get(id)
    if not papeleta:
        return jsonify({'error': 'Papeleta no encontrada'}), 404
    
    tarjeta_nombre = papeleta.tarjeta_rel.nombre_tarjeta if papeleta.tarjeta_rel else ''
    aerolinea_nombre = papeleta.aerolinea.nombre if papeleta.aerolinea else ''
    sucursal_nombre = papeleta.sucursal.nombre if papeleta.sucursal else ''
    
    papeleta_relacionada_folio = ''
    if papeleta.papeleta_relacionada_id:
        pap_rel = Papeleta.query.get(papeleta.papeleta_relacionada_id)
        if pap_rel:
            papeleta_relacionada_folio = pap_rel.folio
    
    # Obtener datos del reporte de ventas
    reporte_folio = ''
    reporte_fecha = ''
    reporte_estatus = ''
    if papeleta.reporte_venta_id:
        reporte = ReporteVenta.query.get(papeleta.reporte_venta_id)
        if reporte:
            reporte_folio = reporte.folio
            reporte_fecha = reporte.fecha.strftime('%d/%m/%Y') if reporte.fecha else ''
            reporte_estatus = reporte.estatus or ''
    
    return jsonify({
        'id': papeleta.id,
        'folio': papeleta.folio,
        'tarjeta': papeleta.tarjeta,
        'tarjeta_nombre': tarjeta_nombre,
        'fecha_venta': papeleta.fecha_venta.strftime('%d/%m/%Y') if papeleta.fecha_venta else '',
        'total_ticket': float(papeleta.total_ticket or 0),
        'diez_porciento': float(papeleta.diez_porciento or 0),
        'cargo': float(papeleta.cargo or 0),
        'total': float(papeleta.total or 0),
        'facturar_a': papeleta.facturar_a or '',
        'solicito': papeleta.solicito or '',
        'clave_sabre': papeleta.clave_sabre or '',
        'forma_pago': papeleta.forma_pago or '',
        'aerolinea': aerolinea_nombre,
        'proveedor': papeleta.proveedor or '',
        'tipo_cargo': papeleta.tipo_cargo or '',
        'sucursal': sucursal_nombre,
        'extemporanea': papeleta.extemporanea or False,
        'fecha_cargo_real': papeleta.fecha_cargo_real.strftime('%d/%m/%Y') if papeleta.fecha_cargo_real else '',
        'motivo_extemporanea': papeleta.motivo_extemporanea or '',
        'tiene_reembolso': papeleta.tiene_reembolso or False,
        'estatus_reembolso': papeleta.estatus_reembolso or '',
        'motivo_reembolso': papeleta.motivo_reembolso or '',
        'monto_reembolso': float(papeleta.monto_reembolso) if papeleta.monto_reembolso else None,
        'fecha_solicitud_reembolso': papeleta.fecha_solicitud_reembolso.strftime('%d/%m/%Y') if papeleta.fecha_solicitud_reembolso else '',
        'referencia_reembolso': papeleta.referencia_reembolso or '',
        'papeleta_relacionada': papeleta_relacionada_folio,
        'usuario': papeleta.usuario.nombre if papeleta.usuario else '',
        'created_at': papeleta.created_at.strftime('%d/%m/%Y %H:%M') if papeleta.created_at else '',
        # Campos del reporte de ventas
        'reporte_venta_id': papeleta.reporte_venta_id,
        'reporte_folio': reporte_folio,
        'reporte_fecha': reporte_fecha,
        'reporte_estatus': reporte_estatus,
        'esta_reportada': papeleta.reporte_venta_id is not None,
        # Factura
        'numero_factura': papeleta.numero_factura,
        'esta_facturada': papeleta.numero_factura is not None,
        # Archivo del boleto
        'archivo_boleto': papeleta.archivo_boleto,
        'tiene_archivo': papeleta.archivo_boleto is not None
    })


@main.route('/api/verificar-tarjeta/<int:tarjeta_id>')
@login_required
def verificar_tarjeta(tarjeta_id):
    tarjeta = TarjetaCorporativa.query.get_or_404(tarjeta_id)
    requiere_autorizacion = tarjeta.requiere_autorizacion(current_user)
    tiene_autorizacion = False
    autorizacion_id = None
    tiempo_restante = None
    fecha_expiracion = None
    
    if requiere_autorizacion:
        autorizacion = Autorizacion.query.filter_by(
            tarjeta_id=tarjeta_id,
            solicitante_id=current_user.id,
            estatus='aprobada'
        ).order_by(Autorizacion.fecha_respuesta.desc()).first()
        
        if autorizacion and autorizacion.esta_vigente(horas=24):
            tiene_autorizacion = True
            autorizacion_id = autorizacion.id
            fecha_resp = autorizacion.fecha_respuesta
            if fecha_resp.tzinfo is not None:
                fecha_resp = fecha_resp.replace(tzinfo=None)
            expira = fecha_resp + timedelta(hours=24)
            fecha_expiracion = expira.strftime('%d/%m/%Y %H:%M')
            diferencia = expira - datetime.utcnow()
            horas_restantes = diferencia.total_seconds() / 3600
            if horas_restantes > 1:
                tiempo_restante = f"{int(horas_restantes)} hora{'s' if int(horas_restantes) > 1 else ''}"
            else:
                tiempo_restante = f"{int(horas_restantes * 60)} minuto{'s' if int(horas_restantes * 60) > 1 else ''}"
    
    return jsonify({
        'tarjeta_id': tarjeta_id,
        'numero_tarjeta': tarjeta.numero_tarjeta,
        'nombre_tarjeta': tarjeta.nombre_tarjeta,
        'sucursal': tarjeta.sucursal.nombre if tarjeta.sucursal else 'Sin asignar',
        'requiere_autorizacion': requiere_autorizacion,
        'tiene_autorizacion': tiene_autorizacion,
        'autorizacion_id': autorizacion_id,
        'puede_usar': not requiere_autorizacion or tiene_autorizacion,
        'tiempo_restante': tiempo_restante,
        'fecha_expiracion': fecha_expiracion
    })


@main.route('/api/tarjetas-disponibles')
@login_required
def tarjetas_disponibles():
    tarjetas = TarjetaCorporativa.query.filter_by(activa=True).all()
    resultado = []
    for t in tarjetas:
        requiere_auth = t.requiere_autorizacion(current_user)
        tiene_auth = False
        if requiere_auth:
            auth = Autorizacion.query.filter_by(
                tarjeta_id=t.id, solicitante_id=current_user.id, estatus='aprobada'
            ).order_by(Autorizacion.fecha_respuesta.desc()).first()
            tiene_auth = auth and auth.esta_vigente(horas=24)
        resultado.append({
            'id': t.id, 'numero_tarjeta': t.numero_tarjeta, 'nombre_tarjeta': t.nombre_tarjeta,
            'banco': t.banco, 'sucursal': t.sucursal.nombre if t.sucursal else None,
            'requiere_autorizacion': requiere_auth, 'tiene_autorizacion': tiene_auth,
            'puede_usar': not requiere_auth or tiene_auth
        })
    resultado.sort(key=lambda x: (not x['puede_usar'], x['nombre_tarjeta']))
    return jsonify(resultado)


@main.route('/api/usuario/<int:id>')
@login_required
def api_usuario_detalle(id):
    """API para obtener el detalle de un usuario."""
    usuario = Usuario.query.get(id)
    if not usuario:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    
    tarjetas = []
    for asignacion in usuario.tarjetas_asignadas:
        if asignacion.activo and asignacion.tarjeta:
            tarjetas.append({
                'id': asignacion.tarjeta.id,
                'nombre': asignacion.tarjeta.nombre_tarjeta,
                'numero': asignacion.tarjeta.numero_tarjeta
            })
    
    return jsonify({
        'id': usuario.id, 'nombre': usuario.nombre, 'correo': usuario.correo,
        'telefono': usuario.telefono, 'rol': usuario.rol, 'rol_id': usuario.rol_id,
        'sucursal': usuario.sucursal.nombre if usuario.sucursal else None,
        'sucursal_id': usuario.sucursal_id, 'tipo_agente': usuario.tipo_agente,
        'activo': usuario.activo, 'tarjetas': tarjetas, 'tarjetas_count': len(tarjetas),
        'created_at': usuario.created_at.strftime('%d/%m/%Y %H:%M') if usuario.created_at else '',
        'updated_at': usuario.updated_at.strftime('%d/%m/%Y %H:%M') if usuario.updated_at else ''
    })


@main.route('/api/usuario/<int:id>/tarjetas')
@login_required
def api_usuario_tarjetas(id):
    """API para obtener las tarjetas asignadas a un usuario."""
    usuario = Usuario.query.get(id)
    if not usuario:
        return jsonify({'error': 'Usuario no encontrado'}), 404
    tarjetas_ids = [asig.tarjeta_id for asig in usuario.tarjetas_asignadas if asig.activo]
    return jsonify({'tarjetas': tarjetas_ids})


# =============================================================================
# RUTAS DE USUARIOS
# =============================================================================

@main.route('/usuarios')
@login_required
def usuarios():
    """Lista de usuarios del sistema."""
    if not current_user.es_admin():
        flash('Acceso no autorizado. Solo administradores pueden gestionar usuarios.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    usuarios_list = Usuario.query.order_by(Usuario.nombre).all()
    roles_list = Rol.query.order_by(Rol.nombre).all()
    sucursales_list = Sucursal.query.filter_by(activa=True).order_by(Sucursal.nombre).all()
    tarjetas_list = TarjetaCorporativa.query.filter_by(activa=True).order_by(TarjetaCorporativa.nombre_tarjeta).all()
    
    return render_template('usuarios.html', 
                           usuarios=usuarios_list, roles=roles_list,
                           sucursales=sucursales_list, tarjetas=tarjetas_list)


@main.route('/usuarios/nuevo', methods=['POST'])
@login_required
def nuevo_usuario():
    """Crea un nuevo usuario."""
    if not current_user.es_admin():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        nombre = request.form.get('nombre', '').strip()
        correo = request.form.get('correo', '').strip().lower()
        contrasena = request.form.get('contrasena', '')
        telefono = request.form.get('telefono', '').strip() or None
        rol_id = request.form.get('rol_id')
        sucursal_id = request.form.get('sucursal_id')
        tipo_agente = request.form.get('tipo_agente', '').strip() or None
        
        if not nombre or not correo or not contrasena:
            flash('Nombre, correo y contraseña son obligatorios.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        if len(contrasena) < 6:
            flash('La contraseña debe tener al menos 6 caracteres.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        if Usuario.query.filter_by(correo=correo).first():
            flash('Ya existe un usuario con ese correo electrónico.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        rol = Rol.query.get(rol_id)
        if not rol:
            flash('Rol no válido.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        nuevo = Usuario(
            nombre=nombre, correo=correo, telefono=telefono, rol=rol.nombre,
            rol_id=int(rol_id), sucursal_id=int(sucursal_id) if sucursal_id else None,
            tipo_agente=tipo_agente, activo=True
        )
        nuevo.set_password(contrasena)
        db.session.add(nuevo)
        db.session.commit()
        
        audit = AuditLog(
            tabla_nombre='usuarios', registro_id=str(nuevo.id), accion='INSERT',
            datos_nuevos={'nombre': nombre, 'correo': correo, 'rol': rol.nombre, 'sucursal_id': sucursal_id},
            usuario_id=current_user.id, usuario_email=current_user.correo, sucursal_id=current_user.sucursal_id
        )
        db.session.add(audit)
        db.session.commit()
        flash(f'Usuario "{nombre}" creado exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear usuario: {str(e)}', 'danger')
    
    return redirect(url_for('main.usuarios'))


@main.route('/usuarios/editar/<int:id>', methods=['POST'])
@login_required
def editar_usuario(id):
    """Edita un usuario existente."""
    if not current_user.es_admin():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    usuario = Usuario.query.get_or_404(id)
    
    try:
        datos_anteriores = {
            'nombre': usuario.nombre, 'correo': usuario.correo,
            'rol': usuario.rol, 'sucursal_id': usuario.sucursal_id
        }
        
        nombre = request.form.get('nombre', '').strip()
        correo = request.form.get('correo', '').strip().lower()
        telefono = request.form.get('telefono', '').strip() or None
        rol_id = request.form.get('rol_id')
        sucursal_id = request.form.get('sucursal_id')
        tipo_agente = request.form.get('tipo_agente', '').strip() or None
        contrasena = request.form.get('contrasena', '').strip()
        
        if not nombre or not correo:
            flash('Nombre y correo son obligatorios.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        usuario_existente = Usuario.query.filter(Usuario.correo == correo, Usuario.id != id).first()
        if usuario_existente:
            flash('Ya existe otro usuario con ese correo electrónico.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        rol = Rol.query.get(rol_id)
        if not rol:
            flash('Rol no válido.', 'warning')
            return redirect(url_for('main.usuarios'))
        
        usuario.nombre = nombre
        usuario.correo = correo
        usuario.telefono = telefono
        usuario.rol = rol.nombre
        usuario.rol_id = int(rol_id)
        usuario.sucursal_id = int(sucursal_id) if sucursal_id else None
        usuario.tipo_agente = tipo_agente
        
        if contrasena:
            if len(contrasena) < 6:
                flash('La contraseña debe tener al menos 6 caracteres.', 'warning')
                return redirect(url_for('main.usuarios'))
            usuario.set_password(contrasena)
        
        audit = AuditLog(
            tabla_nombre='usuarios', registro_id=str(usuario.id), accion='UPDATE',
            datos_anteriores=datos_anteriores,
            datos_nuevos={'nombre': nombre, 'correo': correo, 'rol': rol.nombre, 
                         'sucursal_id': sucursal_id, 'contrasena_cambiada': bool(contrasena)},
            usuario_id=current_user.id, usuario_email=current_user.correo, sucursal_id=current_user.sucursal_id
        )
        db.session.add(audit)
        db.session.commit()
        flash(f'Usuario "{nombre}" actualizado exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al actualizar usuario: {str(e)}', 'danger')
    
    return redirect(url_for('main.usuarios'))


@main.route('/usuarios/<int:id>/toggle-estatus', methods=['POST'])
@login_required
def toggle_estatus_usuario(id):
    """Activa o desactiva un usuario."""
    if not current_user.es_admin():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    if id == current_user.id:
        flash('No puedes desactivar tu propia cuenta.', 'warning')
        return redirect(url_for('main.usuarios'))
    
    usuario = Usuario.query.get_or_404(id)
    
    try:
        accion = request.form.get('accion')
        motivo = request.form.get('motivo', '').strip()
        
        if not motivo or len(motivo) < 10:
            flash('Debe proporcionar un motivo (mínimo 10 caracteres).', 'warning')
            return redirect(url_for('main.usuarios'))
        
        estado_anterior = usuario.activo
        usuario.activo = not usuario.activo
        
        audit = AuditLog(
            tabla_nombre='usuarios', registro_id=str(usuario.id), accion='UPDATE',
            datos_anteriores={'activo': estado_anterior},
            datos_nuevos={'activo': usuario.activo, 'motivo': motivo, 'accion': accion},
            usuario_id=current_user.id, usuario_email=current_user.correo, sucursal_id=current_user.sucursal_id
        )
        db.session.add(audit)
        db.session.commit()
        
        estado = "activado" if usuario.activo else "desactivado"
        flash(f'Usuario "{usuario.nombre}" ha sido {estado}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('main.usuarios'))


@main.route('/usuarios/<int:id>/tarjetas', methods=['POST'])
@login_required
def asignar_tarjetas_usuario(id):
    """Asigna tarjetas a un usuario."""
    if not current_user.es_admin():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    usuario = Usuario.query.get_or_404(id)
    
    try:
        tarjetas_ids = request.form.getlist('tarjetas[]')
        TarjetaUsuario.query.filter_by(usuario_id=id).update({'activo': False})
        
        for tarjeta_id in tarjetas_ids:
            asignacion = TarjetaUsuario.query.filter_by(tarjeta_id=int(tarjeta_id), usuario_id=id).first()
            if asignacion:
                asignacion.activo = True
            else:
                nueva_asignacion = TarjetaUsuario(
                    tarjeta_id=int(tarjeta_id), usuario_id=id,
                    asignado_por=current_user.id, activo=True
                )
                db.session.add(nueva_asignacion)
        
        audit = AuditLog(
            tabla_nombre='tarjetas_usuarios', registro_id=str(id), accion='UPDATE',
            datos_nuevos={'usuario_id': id, 'tarjetas_asignadas': tarjetas_ids, 'asignado_por': current_user.nombre},
            usuario_id=current_user.id, usuario_email=current_user.correo, sucursal_id=current_user.sucursal_id
        )
        db.session.add(audit)
        db.session.commit()
        flash(f'Tarjetas asignadas a "{usuario.nombre}" exitosamente.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al asignar tarjetas: {str(e)}', 'danger')
    
    return redirect(url_for('main.usuarios'))


# =============================================================================
# RUTAS DE TARJETAS CORPORATIVAS
# =============================================================================

@main.route('/tarjetas')
@login_required
def tarjetas():
    if not current_user.es_gerente_o_superior():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    tarjetas_list = TarjetaCorporativa.query.order_by(TarjetaCorporativa.nombre_tarjeta).all()
    sucursales = Sucursal.query.filter_by(activa=True).all()
    return render_template('tarjetas.html', tarjetas=tarjetas_list, sucursales=sucursales)


@main.route('/tarjetas/nueva', methods=['POST'])
@login_required
def nueva_tarjeta():
    if not current_user.es_gerente_o_superior():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    try:
        numero = request.form.get('numero_tarjeta', '').strip()
        nombre = request.form.get('nombre_tarjeta', '').strip()
        
        if not numero or not nombre:
            flash('El número y nombre de la tarjeta son obligatorios.', 'warning')
            return redirect(url_for('main.tarjetas'))
        
        if TarjetaCorporativa.query.filter_by(numero_tarjeta=numero).first():
            flash(f'Ya existe una tarjeta con el número {numero}.', 'warning')
            return redirect(url_for('main.tarjetas'))
        
        sucursal_id = request.form.get('sucursal_id')
        nueva = TarjetaCorporativa(
            numero_tarjeta=numero, nombre_tarjeta=nombre,
            banco=request.form.get('banco', '').strip() or None,
            titular=request.form.get('titular', '').strip() or None,
            sucursal_id=int(sucursal_id) if sucursal_id else None,
            activa=True
        )
        db.session.add(nueva)
        db.session.commit()
        flash(f'Tarjeta "{nombre}" registrada con éxito.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar la tarjeta: {str(e)}', 'danger')
    
    return redirect(url_for('main.tarjetas'))


@main.route('/tarjetas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_tarjeta(id):
    if not current_user.es_gerente_o_superior():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    tarjeta = TarjetaCorporativa.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            tarjeta.numero_tarjeta = request.form.get('numero_tarjeta', '').strip()
            tarjeta.nombre_tarjeta = request.form.get('nombre_tarjeta', '').strip()
            tarjeta.banco = request.form.get('banco', '').strip() or None
            tarjeta.titular = request.form.get('titular', '').strip() or None
            sucursal_id = request.form.get('sucursal_id')
            tarjeta.sucursal_id = int(sucursal_id) if sucursal_id else None
            tarjeta.activa = request.form.get('activa') == 'on'
            
            agentes_ids = [int(aid) for aid in request.form.getlist('agentes_ids') if aid]
            TarjetaUsuario.query.filter_by(tarjeta_id=tarjeta.id).delete()
            
            for usuario_id in agentes_ids:
                nueva_asignacion = TarjetaUsuario(
                    tarjeta_id=tarjeta.id, usuario_id=usuario_id,
                    asignado_por=current_user.id, activo=True
                )
                db.session.add(nueva_asignacion)
            
            db.session.commit()
            flash(f'Tarjeta "{tarjeta.nombre_tarjeta}" actualizada.', 'success')
            return redirect(url_for('main.tarjetas'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
    
    sucursales = Sucursal.query.filter_by(activa=True).all()
    usuarios = Usuario.query.order_by(Usuario.nombre).all()
    usuarios_asignados_ids = [asig.usuario_id for asig in TarjetaUsuario.query.filter_by(tarjeta_id=tarjeta.id, activo=True).all()]
    
    return render_template('tarjeta_edit.html', tarjeta=tarjeta, sucursales=sucursales,
                           usuarios=usuarios, usuarios_asignados_ids=usuarios_asignados_ids)


@main.route('/tarjetas/eliminar/<int:id>')
@login_required
def eliminar_tarjeta(id):
    if not current_user.es_admin():
        flash('Solo los administradores pueden eliminar tarjetas.', 'danger')
        return redirect(url_for('main.tarjetas'))
    
    tarjeta = TarjetaCorporativa.query.get_or_404(id)
    if tarjeta.papeletas.count() > 0:
        flash('No se puede eliminar: la tarjeta tiene papeletas asociadas.', 'warning')
        return redirect(url_for('main.tarjetas'))
    
    nombre = tarjeta.nombre_tarjeta
    db.session.delete(tarjeta)
    db.session.commit()
    flash(f'Tarjeta "{nombre}" eliminada.', 'info')
    return redirect(url_for('main.tarjetas'))


# =============================================================================
# RUTAS DE AUTORIZACIONES
# =============================================================================

@main.route('/autorizaciones')
@login_required
def autorizaciones():
    if current_user.rol in ['director', 'administrador', 'admin']:
        lista = Autorizacion.query.filter_by(estatus='pendiente').order_by(Autorizacion.fecha_solicitud.asc()).all()
        es_director = True
    else:
        lista = Autorizacion.query.filter_by(solicitante_id=current_user.id).order_by(Autorizacion.fecha_solicitud.desc()).all()
        es_director = False
    return render_template('autorizaciones.html', autorizaciones=lista, es_director=es_director)


@main.route('/autorizaciones/solicitar', methods=['POST'])
@login_required
def solicitar_autorizacion():
    try:
        tarjeta_id = request.form.get('tarjeta_id')
        motivo = request.form.get('motivo', '').strip()
        
        if not tarjeta_id or not motivo:
            flash('Debe seleccionar una tarjeta y proporcionar un motivo.', 'warning')
            return redirect(url_for('main.nueva_papeleta_form'))
        
        tarjeta = TarjetaCorporativa.query.get_or_404(tarjeta_id)
        existente = Autorizacion.query.filter_by(
            tarjeta_id=tarjeta_id, 
            solicitante_id=current_user.id, 
            estatus='pendiente'
        ).first()
        
        if existente:
            flash('Ya tienes una solicitud pendiente para esta tarjeta.', 'warning')
            return redirect(url_for('main.autorizaciones'))
        
        # Crear la autorización
        nueva_auth = Autorizacion(
            tipo='uso_tarjeta',
            solicitante_id=current_user.id,
            tarjeta_id=tarjeta_id,
            motivo=motivo,
            sucursal_id=current_user.sucursal_id or 1
        )
        db.session.add(nueva_auth)
        db.session.commit()
        
        # Enviar notificaciones (sistema + email)
        try:
            resultado = NotificacionService.notificar_autorizacion_solicitada(
                nueva_auth, db, Notificacion
            )
            if resultado['email']:
                flash(f'Solicitud enviada. Se notificó a Dirección por correo.', 'success')
            else:
                flash(f'Solicitud enviada para tarjeta {tarjeta.nombre_tarjeta}.', 'success')
                if resultado['errores']:
                    print(f"Errores de notificación: {resultado['errores']}")
        except Exception as notif_error:
            # Si falla la notificación, no afecta la solicitud
            print(f"Error enviando notificación: {notif_error}")
            flash(f'Solicitud enviada para tarjeta {tarjeta.nombre_tarjeta}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al solicitar autorización: {str(e)}', 'danger')
    
    return redirect(url_for('main.autorizaciones'))


@main.route('/autorizaciones/responder/<int:id>', methods=['POST'])
@login_required
def responder_autorizacion(id):
    if current_user.rol not in ['director', 'administrador', 'admin']:
        flash('Solo Dirección puede aprobar autorizaciones.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    autorizacion = Autorizacion.query.get_or_404(id)
    if autorizacion.estatus != 'pendiente':
        flash('Esta autorización ya fue respondida.', 'warning')
        return redirect(url_for('main.autorizaciones'))
    
    accion = request.form.get('accion')
    comentario = request.form.get('comentario', '').strip()
    
    try:
        if accion == 'aprobar':
            autorizacion.aprobar(current_user, comentario)
            flash(f'Autorización APROBADA para {autorizacion.solicitante.nombre}.', 'success')
        elif accion == 'rechazar':
            autorizacion.rechazar(current_user, comentario)
            flash(f'Autorización RECHAZADA para {autorizacion.solicitante.nombre}.', 'info')
        else:
            flash('Acción no válida.', 'warning')
            return redirect(url_for('main.autorizaciones'))
        
        db.session.commit()
        
        # Notificar al agente
        try:
            NotificacionService.notificar_autorizacion_respondida(
                autorizacion, db, Notificacion
            )
        except Exception as notif_error:
            print(f"Error enviando notificación de respuesta: {notif_error}")
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('main.autorizaciones'))


@main.route('/autorizacion/aprobar/<token>')
def aprobar_por_email(token):
    """Aprueba una autorización via link de email (sin login requerido)"""
    autorizacion = Autorizacion.query.filter_by(token=token).first()
    
    if not autorizacion:
        flash('Enlace de autorización inválido o expirado.', 'danger')
        return render_template('autorizacion_resultado.html', 
                               exito=False, 
                               mensaje='Enlace inválido o expirado')
    
    if autorizacion.estatus != 'pendiente':
        flash(f'Esta autorización ya fue {autorizacion.estatus}.', 'warning')
        return render_template('autorizacion_resultado.html', 
                               exito=False, 
                               mensaje=f'Esta autorización ya fue {autorizacion.estatus}')
    
    try:
        director = Usuario.query.filter_by(rol='director').first()
        autorizacion.aprobar_por_token(director.id if director else None)
        autorizacion.token = None
        db.session.commit()
        
        try:
            if NotificacionService:
                NotificacionService.notificar_autorizacion_respondida(autorizacion, db, Notificacion)
        except:
            pass
        
        return render_template('autorizacion_resultado.html', 
                               exito=True, 
                               mensaje='Autorización APROBADA',
                               autorizacion=autorizacion)
    except Exception as e:
        db.session.rollback()
        return render_template('autorizacion_resultado.html', 
                               exito=False, 
                               mensaje=f'Error: {str(e)}')


@main.route('/autorizacion/rechazar/<token>')
def rechazar_por_email(token):
    """Rechaza una autorización via link de email (sin login requerido)"""
    autorizacion = Autorizacion.query.filter_by(token=token).first()
    
    if not autorizacion:
        flash('Enlace de autorización inválido o expirado.', 'danger')
        return render_template('autorizacion_resultado.html', 
                               exito=False, 
                               mensaje='Enlace inválido o expirado')
    
    if autorizacion.estatus != 'pendiente':
        flash(f'Esta autorización ya fue {autorizacion.estatus}.', 'warning')
        return render_template('autorizacion_resultado.html', 
                               exito=False, 
                               mensaje=f'Esta autorización ya fue {autorizacion.estatus}')
    
    try:
        director = Usuario.query.filter_by(rol='director').first()
        autorizacion.rechazar_por_token(director.id if director else None)
        autorizacion.token = None
        db.session.commit()
        
        try:
            if NotificacionService:
                NotificacionService.notificar_autorizacion_respondida(autorizacion, db, Notificacion)
        except:
            pass
        
        return render_template('autorizacion_resultado.html', 
                               exito=True, 
                               mensaje='Autorización RECHAZADA',
                               autorizacion=autorizacion)
    except Exception as e:
        db.session.rollback()
        return render_template('autorizacion_resultado.html', 
                               exito=False, 
                               mensaje=f'Error: {str(e)}')

# =============================================================================
# RUTAS DE FACTURACIÓN
# =============================================================================

@main.route('/facturacion')
@login_required
def facturacion():
    """Vista principal de facturación - muestra papeletas y desgloses BSP pendientes de facturar"""
    # Solo facturación, admin y gerente pueden acceder
    if current_user.rol not in ['facturacion', 'administrador', 'admin', 'gerente', 'director']:
        flash('No tienes permiso para acceder a este módulo.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Diccionarios para traducir fechas al español
    DIAS_ES = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }
    MESES_ES = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
        'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
        'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    
    def fecha_espanol(fecha):
        if not fecha:
            return 'Sin fecha'
        fecha_str = fecha.strftime('%A, %d de %B %Y')
        for en, es in DIAS_ES.items():
            fecha_str = fecha_str.replace(en, es)
        for en, es in MESES_ES.items():
            fecha_str = fecha_str.replace(en, es)
        return fecha_str
    
    # 1. Papeletas pendientes de facturar (Low Cost con empresa = crédito)
    papeletas_pendientes = Papeleta.query.filter(
        Papeleta.estatus_facturacion == 'pendiente',
        Papeleta.empresa_id.isnot(None)  # Solo las que tienen empresa (crédito)
    ).order_by(Papeleta.fecha_venta.desc()).all()
    
    # 2. Desgloses BSP pendientes de facturar
    # Buscar desgloses que:
    # - Sean de aerolíneas BSP (es_bsp=True) O sean Aeromexico por nombre
    # - Tengan empresa_id (cliente de facturación/crédito)
    # - No tengan factura registrada aún (numero_factura es NULL o estatus_facturacion = 'pendiente')
    desgloses_bsp_pendientes = Desglose.query.join(Aerolinea).filter(
        db.or_(
            Aerolinea.es_bsp == True,
            Aerolinea.nombre.ilike('%aeromexico%'),
            Aerolinea.nombre.ilike('%aerom%xico%'),
            Aerolinea.nombre.ilike('%american%'),
            Aerolinea.nombre.ilike('%united%'),
            Aerolinea.nombre.ilike('%delta%')
        ),
        Desglose.empresa_id.isnot(None),  # Tiene cliente de facturación
        db.or_(
            Desglose.estatus_facturacion == 'pendiente',
            Desglose.estatus_facturacion.is_(None),
            Desglose.numero_factura.is_(None)
        )
    ).order_by(Desglose.fecha_emision.desc()).all()
    
    # Crear lista unificada de expedientes pendientes
    expedientes_pendientes = []
    
    # Agregar papeletas
    for p in papeletas_pendientes:
        expedientes_pendientes.append({
            'tipo': 'papeleta',
            'id': p.id,
            'folio': p.folio,
            'fecha': p.fecha_venta,
            'fecha_display': fecha_espanol(p.fecha_venta),
            'empresa': p.empresa.nombre_empresa if p.empresa else p.facturar_a,
            'pasajero': p.solicito,
            'aerolinea': p.aerolinea.nombre if p.aerolinea else (p.proveedor or 'N/A'),
            'total': float(p.total or 0),
            'clave': p.clave_sabre,
            'tiene_desglose': bool(p.clave_sabre),
            'tiene_boleto': bool(p.archivo_boleto),
            'objeto': p
        })
    
    # Agregar desgloses BSP
    for d in desgloses_bsp_pendientes:
        expedientes_pendientes.append({
            'tipo': 'desglose_bsp',
            'id': d.folio,
            'folio': f'BSP-{d.folio}',
            'fecha': d.fecha_emision,
            'fecha_display': fecha_espanol(d.fecha_emision),
            'empresa': d.empresa.nombre_empresa if d.empresa else 'N/A',
            'pasajero': d.pasajero_nombre or 'N/A',
            'aerolinea': d.aerolinea.nombre if d.aerolinea else 'N/A',
            'total': float(d.total or 0),
            'clave': d.clave_sabre or d.clave_reserva,
            'tiene_desglose': True,  # BSP siempre tiene desglose
            'tiene_boleto': bool(d.numero_boleto),
            'objeto': d
        })
    
    # Ordenar por fecha descendente
    expedientes_pendientes.sort(key=lambda x: x['fecha'] or date.min, reverse=True)
    
    # Agrupar por fecha
    def agrupar_por_fecha(expedientes):
        grupos = OrderedDict()
        for exp in expedientes:
            if exp['fecha']:
                fecha_key = exp['fecha'].strftime('%Y-%m-%d')
                fecha_display = exp['fecha_display']
            else:
                fecha_key = ''
                fecha_display = 'Sin fecha'
            
            if fecha_key not in grupos:
                grupos[fecha_key] = {'fecha_display': fecha_display, 'expedientes': []}
            grupos[fecha_key]['expedientes'].append(exp)
        
        return list(grupos.items())
    
    expedientes_agrupados = agrupar_por_fecha(expedientes_pendientes)
    
    # Papeletas ya facturadas (para historial)
    papeletas_facturadas = Papeleta.query.filter(
        Papeleta.estatus_facturacion.in_(['facturada', 'aprobada', 'rechazada'])
    ).order_by(Papeleta.fecha_facturacion.desc()).limit(50).all()
    
    # Desgloses BSP ya facturados
    desgloses_bsp_facturados = Desglose.query.join(Aerolinea).filter(
        Aerolinea.es_bsp == True,
        Desglose.estatus_facturacion.in_(['facturada', 'aprobada', 'rechazada'])
    ).order_by(Desglose.fecha_facturacion.desc()).limit(50).all()
    
    # Obtener desgloses para vincular con papeletas
    desgloses = Desglose.query.all()
    desgloses_dict = {}
    for d in desgloses:
        if d.clave_sabre:
            desgloses_dict[d.clave_sabre] = d
        if d.clave_reserva:
            desgloses_dict[d.clave_reserva] = d
    
    return render_template('facturacion.html',
                           papeletas_pendientes=papeletas_pendientes,
                           desgloses_bsp_pendientes=desgloses_bsp_pendientes,
                           expedientes_pendientes=expedientes_pendientes,
                           expedientes_agrupados=expedientes_agrupados,
                           papeletas_facturadas=papeletas_facturadas,
                           desgloses_bsp_facturados=desgloses_bsp_facturados,
                           desgloses_dict=desgloses_dict)


@main.route('/facturacion/registrar/<int:id>', methods=['POST'])
@login_required
def registrar_factura(id):
    """Registra la factura para una papeleta"""
    if current_user.rol not in ['facturacion', 'administrador', 'admin', 'gerente', 'director']:
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No tienes permiso para esta acción.'})
        flash('No tienes permiso para esta acción.', 'danger')
        return redirect(url_for('main.facturacion'))
    
    papeleta = Papeleta.query.get_or_404(id)
    
    numero_factura = request.form.get('numero_factura', '').strip()
    monto_factura = request.form.get('monto_factura', '').strip()
    
    if not numero_factura or not monto_factura:
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Debe ingresar el número y monto de la factura.'})
        flash('Debe ingresar el número y monto de la factura.', 'warning')
        return redirect(url_for('main.facturacion'))
    
    try:
        papeleta.numero_factura = numero_factura
        papeleta.monto_factura = float(monto_factura)
        papeleta.estatus_facturacion = 'facturada'
        papeleta.facturada_por_id = current_user.id
        papeleta.fecha_facturacion = datetime.utcnow()
        
        db.session.commit()
        
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': f'Factura {numero_factura} registrada correctamente.',
                'folio': papeleta.folio
            })
        
        flash(f'Factura {numero_factura} registrada para papeleta {papeleta.folio}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
        flash(f'Error al registrar factura: {str(e)}', 'danger')
    
    return redirect(url_for('main.facturacion'))


@main.route('/facturacion/registrar-bsp/<int:folio>', methods=['POST'])
@login_required
def registrar_factura_bsp(folio):
    """Registra la factura para un desglose BSP"""
    if current_user.rol not in ['facturacion', 'administrador', 'admin', 'gerente', 'director']:
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No tienes permiso para esta acción.'})
        flash('No tienes permiso para esta acción.', 'danger')
        return redirect(url_for('main.facturacion'))
    
    desglose = Desglose.query.get_or_404(folio)
    
    numero_factura = request.form.get('numero_factura', '').strip()
    monto_factura = request.form.get('monto_factura', '').strip()
    
    if not numero_factura or not monto_factura:
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Debe ingresar el número y monto de la factura.'})
        flash('Debe ingresar el número y monto de la factura.', 'warning')
        return redirect(url_for('main.facturacion'))
    
    try:
        desglose.numero_factura = numero_factura
        desglose.monto_factura = float(monto_factura)
        desglose.estatus_facturacion = 'facturada'
        desglose.facturada_por_id = current_user.id
        desglose.fecha_facturacion = datetime.utcnow()
        
        db.session.commit()
        
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': f'Factura {numero_factura} registrada correctamente.',
                'folio': f'BSP-{desglose.folio}'
            })
        
        flash(f'Factura {numero_factura} registrada para desglose BSP-{desglose.folio}.', 'success')
        
    except Exception as e:
        db.session.rollback()
        if request.args.get('ajax') == '1' or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
        flash(f'Error al registrar factura: {str(e)}', 'danger')
    
    return redirect(url_for('main.facturacion'))


@main.route('/desglose/subir-boleto/<int:folio>', methods=['POST'])
@login_required
def subir_boleto_desglose(folio):
    """Sube el PDF del boleto para un desglose BSP"""
    import os
    from werkzeug.utils import secure_filename
    
    desglose = Desglose.query.get_or_404(folio)
    
    if 'archivo_boleto' not in request.files:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        flash('No se seleccionó ningún archivo', 'warning')
        return redirect(request.referrer or url_for('main.facturacion'))
    
    archivo = request.files['archivo_boleto']
    
    if archivo.filename == '':
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'No se seleccionó ningún archivo'})
        flash('No se seleccionó ningún archivo', 'warning')
        return redirect(request.referrer or url_for('main.facturacion'))
    
    # Validar extensión
    ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
    extension = archivo.filename.rsplit('.', 1)[1].lower() if '.' in archivo.filename else ''
    
    if extension not in ALLOWED_EXTENSIONS:
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': 'Solo se permiten archivos PDF, PNG, JPG'})
        flash('Solo se permiten archivos PDF, PNG, JPG', 'warning')
        return redirect(request.referrer or url_for('main.facturacion'))
    
    try:
        # Crear nombre único para el archivo
        nombre_archivo = f"bsp_{desglose.folio}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}"
        nombre_archivo = secure_filename(nombre_archivo)
        
        # Ruta de guardado
        upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads', 'boletos')
        os.makedirs(upload_folder, exist_ok=True)
        
        ruta_completa = os.path.join(upload_folder, nombre_archivo)
        archivo.save(ruta_completa)
        
        # Actualizar registro
        desglose.archivo_boleto = nombre_archivo
        db.session.commit()
        
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True, 
                'message': 'Boleto subido correctamente',
                'archivo': nombre_archivo
            })
        
        flash('Boleto subido correctamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'message': f'Error: {str(e)}'})
        flash(f'Error al subir boleto: {str(e)}', 'danger')
    
    return redirect(request.referrer or url_for('main.facturacion'))


@main.route('/uploads/boletos/<filename>')
@login_required
def ver_boleto(filename):
    """Servir archivos de boletos PDF"""
    import os
    from flask import send_from_directory
    upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads', 'boletos')
    return send_from_directory(upload_folder, filename)


@main.route('/revision-facturas')
@login_required
def revision_facturas():
    """Vista para que el gerente revise facturas registradas"""
    if current_user.rol not in ['gerente', 'director', 'administrador', 'admin']:
        flash('Solo gerentes pueden revisar facturas.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Diccionarios para traducir fechas al español
    DIAS_ES = {
        'Monday': 'Lunes', 'Tuesday': 'Martes', 'Wednesday': 'Miércoles',
        'Thursday': 'Jueves', 'Friday': 'Viernes', 'Saturday': 'Sábado', 'Sunday': 'Domingo'
    }
    MESES_ES = {
        'January': 'Enero', 'February': 'Febrero', 'March': 'Marzo', 'April': 'Abril',
        'May': 'Mayo', 'June': 'Junio', 'July': 'Julio', 'August': 'Agosto',
        'September': 'Septiembre', 'October': 'Octubre', 'November': 'Noviembre', 'December': 'Diciembre'
    }
    
    def fecha_espanol(fecha):
        """Convierte una fecha a formato español: Lunes, 30 de Enero 2026"""
        if not fecha:
            return 'Sin fecha'
        fecha_str = fecha.strftime('%A, %d de %B %Y')
        for en, es in DIAS_ES.items():
            fecha_str = fecha_str.replace(en, es)
        for en, es in MESES_ES.items():
            fecha_str = fecha_str.replace(en, es)
        return fecha_str
    
    # Facturas pendientes de revisión (estatus = 'facturada')
    pendientes = Papeleta.query.filter(
        Papeleta.estatus_facturacion == 'facturada'
    ).order_by(Papeleta.fecha_facturacion.desc()).all()
    
    # Agrupar pendientes por fecha (solo día), ordenado de más reciente a más antiguo
    def agrupar_por_fecha(papeletas):
        grupos = OrderedDict()
        for p in papeletas:
            if p.fecha_facturacion:
                fecha_key = p.fecha_facturacion.strftime('%Y-%m-%d')
                fecha_display = fecha_espanol(p.fecha_facturacion)
            else:
                fecha_key = ''
                fecha_display = 'Sin fecha'
            
            if fecha_key not in grupos:
                grupos[fecha_key] = {'fecha_display': fecha_display, 'facturas': []}
            grupos[fecha_key]['facturas'].append(p)
        
        # Ordenar facturas dentro de cada día por número de factura (consecutivo ascendente)
        for fecha_key in grupos:
            grupos[fecha_key]['facturas'] = sorted(
                grupos[fecha_key]['facturas'], 
                key=lambda p: p.numero_factura or ''
            )
        
        return list(grupos.items())
    
    pendientes_agrupados = agrupar_por_fecha(pendientes)
    
    # Calcular monto total pendiente
    monto_pendiente = sum(float(p.monto_factura or 0) for p in pendientes)
    
    # Historial de revisiones (aprobadas y rechazadas)
    historial = Papeleta.query.filter(
        Papeleta.estatus_facturacion.in_(['aprobada', 'rechazada'])
    ).order_by(Papeleta.fecha_aprobacion.desc()).limit(50).all()
    
    # Estadísticas del día
    hoy_inicio = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    
    aprobadas_hoy = Papeleta.query.filter(
        Papeleta.estatus_facturacion == 'aprobada',
        Papeleta.fecha_aprobacion >= hoy_inicio
    ).count()
    
    rechazadas_hoy = Papeleta.query.filter(
        Papeleta.estatus_facturacion == 'rechazada',
        Papeleta.fecha_aprobacion >= hoy_inicio
    ).count()
    
    # Obtener desgloses para el modal
    desgloses = Desglose.query.all()
    desgloses_dict = {}
    for d in desgloses:
        if d.clave_sabre:
            desgloses_dict[d.clave_sabre] = d
        if d.clave_reserva:
            desgloses_dict[d.clave_reserva] = d
    
    return render_template('revision_facturas.html',
                           pendientes=pendientes,
                           pendientes_agrupados=pendientes_agrupados,
                           historial=historial,
                           monto_pendiente=monto_pendiente,
                           aprobadas_hoy=aprobadas_hoy,
                           rechazadas_hoy=rechazadas_hoy,
                           desgloses_dict=desgloses_dict)


@main.route('/revision-facturas/aprobar/<int:id>', methods=['POST'])
@login_required
def aprobar_factura(id):
    """Aprueba una factura - montos coinciden"""
    if current_user.rol not in ['gerente', 'director', 'administrador', 'admin']:
        flash('Solo gerentes pueden aprobar facturas.', 'danger')
        return redirect(url_for('main.revision_facturas'))
    
    papeleta = Papeleta.query.get_or_404(id)
    
    if papeleta.estatus_facturacion != 'facturada':
        flash('Esta papeleta no está lista para aprobación.', 'warning')
        return redirect(url_for('main.revision_facturas'))
    
    try:
        papeleta.estatus_facturacion = 'aprobada'
        papeleta.aprobada_por_id = current_user.id
        papeleta.fecha_aprobacion = datetime.utcnow()
        
        db.session.commit()
        flash(f'Factura {papeleta.numero_factura} APROBADA.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('main.revision_facturas'))


@main.route('/revision-facturas/rechazar/<int:id>', methods=['POST'])
@login_required
def rechazar_factura(id):
    """Rechaza una factura - montos no coinciden o hay error"""
    if current_user.rol not in ['gerente', 'director', 'administrador', 'admin']:
        flash('Solo gerentes pueden rechazar facturas.', 'danger')
        return redirect(url_for('main.revision_facturas'))
    
    papeleta = Papeleta.query.get_or_404(id)
    comentario = request.form.get('comentario', '').strip()
    
    if papeleta.estatus_facturacion != 'facturada':
        flash('Esta papeleta no está lista para revisión.', 'warning')
        return redirect(url_for('main.revision_facturas'))
    
    if not comentario:
        flash('Debe proporcionar un motivo para rechazar la factura.', 'warning')
        return redirect(url_for('main.revision_facturas'))
    
    try:
        papeleta.estatus_facturacion = 'rechazada'
        papeleta.aprobada_por_id = current_user.id
        papeleta.fecha_aprobacion = datetime.utcnow()
        # Guardar motivo en justificacion_pendiente
        papeleta.justificacion_pendiente = f'Factura rechazada: {comentario}'
        
        db.session.commit()
        flash(f'Factura {papeleta.numero_factura} RECHAZADA. Motivo: {comentario}', 'warning')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    
    return redirect(url_for('main.revision_facturas'))


# =============================================================================
# RUTAS DE PAPELETAS
# =============================================================================

@main.route('/papeletas', methods=['GET'])
@login_required
def consulta_papeletas():
    """Muestra la lista de papeletas agrupadas por tarjeta."""
    if current_user.es_admin():
        papeletas_list = Papeleta.query.order_by(Papeleta.tarjeta, Papeleta.fecha_venta.desc()).all()
    elif current_user.es_gerente_o_superior():
        papeletas_list = Papeleta.query.filter_by(sucursal_id=current_user.sucursal_id).order_by(Papeleta.tarjeta, Papeleta.fecha_venta.desc()).all()
    else:
        papeletas_list = Papeleta.query.filter_by(usuario_id=current_user.id).order_by(Papeleta.tarjeta, Papeleta.fecha_venta.desc()).all()
    
    papeletas_por_tarjeta = OrderedDict()
    
    for papeleta in papeletas_list:
        tarjeta_numero = papeleta.tarjeta
        tarjeta_info = None
        if papeleta.tarjeta_id:
            tarjeta_obj = TarjetaCorporativa.query.get(papeleta.tarjeta_id)
            if tarjeta_obj:
                tarjeta_info = {'numero': tarjeta_obj.numero_tarjeta, 'nombre': tarjeta_obj.nombre_tarjeta, 'banco': tarjeta_obj.banco}
        
        if not tarjeta_info:
            tarjeta_info = {'numero': tarjeta_numero, 'nombre': None, 'banco': None}
        
        class TarjetaKey:
            def __init__(self, numero, nombre, banco):
                self.numero = numero
                self.nombre = nombre
                self.banco = banco
            def __hash__(self):
                return hash(self.numero)
            def __eq__(self, other):
                return self.numero == other.numero
        
        tarjeta_key = TarjetaKey(tarjeta_info['numero'], tarjeta_info['nombre'], tarjeta_info['banco'])
        if tarjeta_key not in papeletas_por_tarjeta:
            papeletas_por_tarjeta[tarjeta_key] = []
        papeletas_por_tarjeta[tarjeta_key].append(papeleta)
    
    # === ESTADÍSTICAS PARA EL TEMPLATE ===
    fecha_hoy = fecha_mexico()
    
    # Papeletas de hoy
    papeletas_hoy_list = [p for p in papeletas_list if p.fecha_venta == fecha_hoy]
    papeletas_hoy = len(papeletas_hoy_list)
    total_hoy = sum([float(p.total or 0) for p in papeletas_hoy_list])
    
    # Sin reportar (no tienen reporte_venta_id y no tienen factura)
    sin_reportar_list = [p for p in papeletas_list if not p.reporte_venta_id and not p.numero_factura]
    sin_reportar = len(sin_reportar_list)
    total_sin_reportar = sum([float(p.total or 0) for p in sin_reportar_list])
    
    # Sin facturar (tienen reporte pero no factura)
    sin_facturar_list = [p for p in papeletas_list if p.reporte_venta_id and not p.numero_factura]
    sin_facturar = len(sin_facturar_list)
    total_sin_facturar = sum([float(p.total or 0) for p in sin_facturar_list])
    
    # Facturadas
    facturadas_list = [p for p in papeletas_list if p.numero_factura]
    facturadas = len(facturadas_list)
    total_facturadas = sum([float(p.total or 0) for p in facturadas_list])
    
    # Urgentes (más de 3 días sin reportar)
    urgentes = len([p for p in sin_reportar_list 
                    if p.fecha_venta and (fecha_hoy - p.fecha_venta).days > 3])
    
    # Efectivo pendiente
    efectivo_pendiente = sum([float(p.total or 0) for p in sin_reportar_list 
                              if p.forma_pago and 'efectivo' in p.forma_pago.lower()])
    
    return render_template('consulta_papeletas.html', 
        papeletas_por_tarjeta=papeletas_por_tarjeta,
        papeletas_hoy=papeletas_hoy,
        total_hoy=total_hoy,
        sin_reportar=sin_reportar,
        total_sin_reportar=total_sin_reportar,
        sin_facturar=sin_facturar,
        total_sin_facturar=total_sin_facturar,
        facturadas=facturadas,
        total_facturadas=total_facturadas,
        urgentes=urgentes,
        efectivo_pendiente=efectivo_pendiente,
        fecha_actual=fecha_hoy)


@main.route('/papeletas/nueva', methods=['GET'])
@login_required
def nueva_papeleta_form():
    """Muestra el formulario para crear una nueva papeleta."""
    empresas_list = Empresa.query.order_by(Empresa.nombre_empresa).all()
    aerolineas_list = Aerolinea.query.order_by(Aerolinea.nombre).all()
    tarjetas_list = TarjetaCorporativa.query.filter_by(activa=True).order_by(TarjetaCorporativa.nombre_tarjeta).all()
    
    tarjetas_info = []
    for t in tarjetas_list:
        requiere_auth = t.requiere_autorizacion(current_user)
        tiene_auth = False
        tiempo_restante = None
        fecha_expiracion = None
        
        if requiere_auth:
            auth = Autorizacion.query.filter_by(tarjeta_id=t.id, solicitante_id=current_user.id, estatus='aprobada').order_by(Autorizacion.fecha_respuesta.desc()).first()
            if auth and auth.esta_vigente(horas=24):
                tiene_auth = True
                fecha_resp = auth.fecha_respuesta
                if fecha_resp.tzinfo is not None:
                    fecha_resp = fecha_resp.replace(tzinfo=None)
                expira = fecha_resp + timedelta(hours=24)
                fecha_expiracion = expira.strftime('%d/%m/%Y %H:%M')
                diferencia = expira - datetime.utcnow()
                horas_restantes = diferencia.total_seconds() / 3600
                tiempo_restante = f"{int(horas_restantes)}h" if horas_restantes > 1 else f"{int(horas_restantes * 60)}min"
        
        usuarios_asignados = [asig.usuario.nombre for asig in TarjetaUsuario.query.filter_by(tarjeta_id=t.id, activo=True).all() if asig.usuario]
        
        tarjetas_info.append({
            'tarjeta': t, 'requiere_autorizacion': requiere_auth, 'tiene_autorizacion': tiene_auth,
            'puede_usar': not requiere_auth or tiene_auth, 'usuarios_asignados': usuarios_asignados,
            'tiempo_restante': tiempo_restante, 'fecha_expiracion': fecha_expiracion
        })
    
    return render_template('papeletas.html', 
                           empresas=empresas_list, 
                           aerolineas=aerolineas_list, 
                           tarjetas_info=tarjetas_info,
                           fecha_hoy=fecha_mexico().strftime('%Y-%m-%d'))


@main.route('/papeletas/nueva', methods=['POST'])
@login_required
def nueva_papeleta_post():
    """Recibe los datos del formulario y crea una nueva papeleta."""
    import os
    from werkzeug.utils import secure_filename
    from flask import current_app
    
    # Configuración de archivos - usar la carpeta static de la app
    UPLOAD_FOLDER = os.path.join(current_app.static_folder, 'uploads', 'boletos')
    ALLOWED_EXTENSIONS = {'pdf'}
    
    def allowed_file(filename):
        return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS
    
    # Crear carpeta si no existe
    if not os.path.exists(UPLOAD_FOLDER):
        os.makedirs(UPLOAD_FOLDER)
    
    try:
        # Validar archivo PDF
        if 'archivo_boleto' not in request.files:
            flash('Debe adjuntar el PDF del boleto.', 'warning')
            return redirect(url_for('main.nueva_papeleta_form'))
        
        archivo = request.files['archivo_boleto']
        if archivo.filename == '':
            flash('Debe seleccionar un archivo PDF del boleto.', 'warning')
            return redirect(url_for('main.nueva_papeleta_form'))
        
        if not allowed_file(archivo.filename):
            flash('Solo se permiten archivos PDF.', 'warning')
            return redirect(url_for('main.nueva_papeleta_form'))
        
        tarjeta_id = request.form.get('tarjeta_id')
        tarjeta_manual = request.form.get('tarjeta_manual', '').strip()
        es_extemporanea = request.form.get('extemporanea') == '1'
        
        tarjeta_numero = None
        tarjeta_obj = None
        autorizacion_id = None
        
        if tarjeta_id:
            tarjeta_obj = TarjetaCorporativa.query.get(tarjeta_id)
            if not tarjeta_obj:
                flash('Tarjeta no encontrada.', 'danger')
                return redirect(url_for('main.nueva_papeleta_form'))
            tarjeta_numero = tarjeta_obj.numero_tarjeta
            
            if not es_extemporanea and tarjeta_obj.requiere_autorizacion(current_user):
                auth = Autorizacion.query.filter_by(tarjeta_id=tarjeta_id, solicitante_id=current_user.id, estatus='aprobada').order_by(Autorizacion.fecha_respuesta.desc()).first()
                if not auth or not auth.esta_vigente(horas=24):
                    flash('Necesitas autorización vigente para usar esta tarjeta.', 'danger')
                    return redirect(url_for('main.nueva_papeleta_form'))
                autorizacion_id = auth.id
        elif tarjeta_manual:
            if len(tarjeta_manual) != 4 or not tarjeta_manual.isdigit():
                flash('La terminación de tarjeta debe ser de 4 dígitos.', 'warning')
                return redirect(url_for('main.nueva_papeleta_form'))
            tarjeta_numero = tarjeta_manual
        else:
            flash('Debe seleccionar o ingresar una tarjeta.', 'warning')
            return redirect(url_for('main.nueva_papeleta_form'))
        
        fecha_cargo_real = None
        motivo_extemporanea = None
        if es_extemporanea:
            fecha_cargo_real_str = request.form.get('fecha_cargo_real')
            if not fecha_cargo_real_str:
                flash('Debe ingresar la fecha real del cargo para papeletas extemporáneas.', 'warning')
                return redirect(url_for('main.nueva_papeleta_form'))
            fecha_cargo_real = datetime.strptime(fecha_cargo_real_str, '%Y-%m-%d').date()
            motivo_extemporanea = request.form.get('motivo_extemporanea', '').strip()
            if not motivo_extemporanea:
                flash('Debe ingresar el motivo del registro tardío.', 'warning')
                return redirect(url_for('main.nueva_papeleta_form'))
        
        fecha_venta_str = request.form.get('fecha_venta')
        fecha_venta = datetime.strptime(fecha_venta_str, '%Y-%m-%d').date() if fecha_venta_str else None
        empresa_id_str = request.form.get('facturar_a')
        empresa_id = int(empresa_id_str) if empresa_id_str else None
        facturar_a_nombre = ''
        if empresa_id:
            empresa = Empresa.query.get(empresa_id)
            facturar_a_nombre = empresa.nombre_empresa if empresa else ''
        aerolinea_id_str = request.form.get('aerolinea_id')
        aerolinea_id = int(aerolinea_id_str) if aerolinea_id_str else None

        folio = request.form.get('folio')
        if not folio:
            ultima = Papeleta.query.filter(Papeleta.folio.like(f"{tarjeta_numero}-%")).order_by(Papeleta.id.desc()).first()
            num = int(ultima.folio.split('-')[1]) + 1 if ultima else 1
            folio = f"{tarjeta_numero}-{num:03d}"

        tiene_reembolso = request.form.get('tiene_reembolso') == '1'
        motivo_reembolso = None
        monto_reembolso = None
        estatus_reembolso = None
        fecha_solicitud_reembolso = None
        referencia_reembolso = None
        papeleta_relacionada_id = None
        
        if tiene_reembolso:
            motivo_reembolso = request.form.get('motivo_reembolso', '')
            if motivo_reembolso == 'otro':
                motivo_reembolso = request.form.get('motivo_reembolso_otro', '').strip()
            monto_str = request.form.get('monto_reembolso', '').strip()
            monto_reembolso = float(monto_str) if monto_str else None
            estatus_reembolso = request.form.get('estatus_reembolso', 'pendiente')
            fecha_sol_str = request.form.get('fecha_solicitud_reembolso', '').strip()
            if fecha_sol_str:
                fecha_solicitud_reembolso = datetime.strptime(fecha_sol_str, '%Y-%m-%d').date()
            referencia_reembolso = request.form.get('referencia_reembolso', '').strip() or None
            folio_relacionado = request.form.get('papeleta_relacionada_folio', '').strip()
            if folio_relacionado:
                pap_rel = Papeleta.query.filter_by(folio=folio_relacionado).first()
                if pap_rel:
                    papeleta_relacionada_id = pap_rel.id
        
        nueva = Papeleta(
            folio=folio, tarjeta=tarjeta_numero, tarjeta_id=int(tarjeta_id) if tarjeta_id else None,
            fecha_venta=fecha_venta, total_ticket=float(request.form.get('total_ticket', 0)),
            diez_porciento=float(request.form.get('diez_porciento', 0)), cargo=float(request.form.get('cargo', 0)),
            total=float(request.form.get('total', 0)), facturar_a=facturar_a_nombre,
            solicito=request.form.get('solicito', ''), clave_sabre=request.form.get('clave_sabre', ''),
            clave_reserva=request.form.get('clave_reserva', '').strip().upper() or None,
            pasajero_nombre=request.form.get('pasajero_nombre', '').strip().upper() or None,
            forma_pago=request.form.get('forma_pago', ''), empresa_id=empresa_id, aerolinea_id=aerolinea_id,
            usuario_id=current_user.id, autorizacion_id=autorizacion_id, sucursal_id=current_user.sucursal_id,
            tipo_cargo=request.form.get('tipo_cargo', ''), proveedor=request.form.get('proveedor', ''),
            extemporanea=es_extemporanea, fecha_cargo_real=fecha_cargo_real, motivo_extemporanea=motivo_extemporanea,
            tiene_reembolso=tiene_reembolso, motivo_reembolso=motivo_reembolso, monto_reembolso=monto_reembolso,
            estatus_reembolso=estatus_reembolso, fecha_solicitud_reembolso=fecha_solicitud_reembolso,
            referencia_reembolso=referencia_reembolso, papeleta_relacionada_id=papeleta_relacionada_id
        )
        db.session.add(nueva)
        db.session.flush()  # Para obtener el ID antes del commit
        
        # Guardar archivo PDF
        nombre_archivo = f"{folio}_boleto.pdf"
        nombre_archivo = secure_filename(nombre_archivo)
        ruta_archivo = os.path.join(UPLOAD_FOLDER, nombre_archivo)
        archivo.save(ruta_archivo)
        nueva.archivo_boleto = nombre_archivo
        
        db.session.commit()

        if tiene_reembolso:
            flash(f'Papeleta {nueva.folio} registrada con reembolso {estatus_reembolso}.', 'success')
        elif es_extemporanea:
            flash(f'Papeleta extemporánea {nueva.folio} registrada.', 'success')
        else:
            flash(f'Papeleta {nueva.folio} creada con éxito.', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error al crear la papeleta: {str(e)}', 'danger')

    return redirect(url_for('main.consulta_papeletas'))


@main.route('/papeletas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_papeleta(id):
    """Edita una papeleta existente. Solo administración puede editar."""
    if not current_user.es_admin():
        flash('Solo administración puede editar papeletas.', 'danger')
        return redirect(url_for('main.consulta_papeletas'))
    
    papeleta = Papeleta.query.get_or_404(id)
    
    if request.method == 'POST':
        try:
            motivo_edicion = request.form.get('motivo_edicion', '').strip()
            if not motivo_edicion or len(motivo_edicion) < 10:
                flash('Debe proporcionar un motivo de edición (mínimo 10 caracteres).', 'warning')
                return redirect(url_for('main.consulta_papeletas'))
            
            datos_anteriores = {
                'fecha_venta': str(papeleta.fecha_venta), 'total_ticket': float(papeleta.total_ticket),
                'total': float(papeleta.total), 'facturar_a': papeleta.facturar_a,
                'solicito': papeleta.solicito, 'clave_sabre': papeleta.clave_sabre
            }
            
            papeleta.fecha_venta = datetime.strptime(request.form.get('fecha_venta'), '%Y-%m-%d').date()
            papeleta.total_ticket = float(request.form.get('total_ticket', 0))
            papeleta.diez_porciento = float(request.form.get('diez_porciento', 0))
            papeleta.cargo = float(request.form.get('cargo', 0))
            papeleta.total = float(request.form.get('total', 0))
            papeleta.solicito = request.form.get('solicito', '')
            papeleta.pasajero_nombre = request.form.get('pasajero_nombre', '').strip().upper() or None
            papeleta.clave_sabre = request.form.get('clave_sabre', '')
            papeleta.clave_reserva = request.form.get('clave_reserva', '').strip().upper() or None
            papeleta.forma_pago = request.form.get('forma_pago', '')
            
            empresa_id = request.form.get('facturar_a')
            if empresa_id:
                papeleta.empresa_id = int(empresa_id)
                empresa = Empresa.query.get(empresa_id)
                papeleta.facturar_a = empresa.nombre_empresa if empresa else ''
            
            aerolinea_id = request.form.get('aerolinea_id')
            papeleta.aerolinea_id = int(aerolinea_id) if aerolinea_id else None
            
            audit = AuditLog(
                tabla_nombre='papeletas', registro_id=str(papeleta.id), accion='UPDATE',
                datos_anteriores=datos_anteriores,
                datos_nuevos={'fecha_venta': str(papeleta.fecha_venta), 'total_ticket': float(papeleta.total_ticket),
                             'total': float(papeleta.total), 'facturar_a': papeleta.facturar_a, 'motivo_edicion': motivo_edicion},
                usuario_id=current_user.id, usuario_email=current_user.correo, sucursal_id=current_user.sucursal_id
            )
            db.session.add(audit)
            db.session.commit()
            flash(f'Papeleta {papeleta.folio} actualizada con éxito.', 'success')
            return redirect(url_for('main.consulta_papeletas'))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
    
    empresas_list = Empresa.query.order_by(Empresa.nombre_empresa).all()
    aerolineas_list = Aerolinea.query.order_by(Aerolinea.nombre).all()
    return render_template('papeletas_edit.html', papeleta=papeleta, empresas=empresas_list, aerolineas=aerolineas_list)


@main.route('/papeletas/eliminar/<int:id>', methods=['POST'])
@login_required
def eliminar_papeleta(id):
    """Elimina una papeleta. Solo administración puede eliminar con motivo."""
    if not current_user.es_admin():
        flash('Solo administración puede eliminar papeletas.', 'danger')
        return redirect(url_for('main.consulta_papeletas'))
    
    papeleta = Papeleta.query.get_or_404(id)
    motivo_tipo = request.form.get('motivo_eliminacion_tipo', '').strip()
    motivo_detalle = request.form.get('motivo_eliminacion_detalle', '').strip()
    
    if not motivo_tipo:
        flash('Debe seleccionar un tipo de motivo para eliminar.', 'warning')
        return redirect(url_for('main.consulta_papeletas'))
    
    if not motivo_detalle or len(motivo_detalle) < 10:
        flash('Debe proporcionar detalle del motivo (mínimo 10 caracteres).', 'warning')
        return redirect(url_for('main.consulta_papeletas'))
    
    try:
        folio = papeleta.folio
        audit = AuditLog(
            tabla_nombre='papeletas', registro_id=str(papeleta.id), accion='DELETE',
            datos_anteriores={
                'folio': papeleta.folio, 'tarjeta': papeleta.tarjeta, 'fecha_venta': str(papeleta.fecha_venta),
                'total_ticket': float(papeleta.total_ticket), 'total': float(papeleta.total),
                'facturar_a': papeleta.facturar_a, 'solicito': papeleta.solicito,
                'clave_sabre': papeleta.clave_sabre, 'usuario_id': papeleta.usuario_id
            },
            datos_nuevos={'motivo_tipo': motivo_tipo, 'motivo_detalle': motivo_detalle, 'eliminado_por': current_user.nombre},
            usuario_id=current_user.id, usuario_email=current_user.correo, sucursal_id=current_user.sucursal_id
        )
        db.session.add(audit)
        db.session.delete(papeleta)
        db.session.commit()
        flash(f'Papeleta {folio} eliminada. Motivo: {motivo_tipo}', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')
    
    return redirect(url_for('main.consulta_papeletas'))


# =============================================================================
# IMPRESIÓN DE TICKETS TÉRMICOS
# =============================================================================

@main.route('/papeleta/<int:id>/ticket')
@login_required
def imprimir_ticket_papeleta(id):
    """
    Genera ticket térmico 80mm para uso interno.
    Parámetros URL opcionales:
    - autoprint=1 : Imprime automáticamente al cargar
    """
    papeleta = Papeleta.query.get_or_404(id)
    
    # Verificar permisos: solo el dueño o admin pueden imprimir
    if papeleta.usuario_id != current_user.id and not current_user.es_admin():
        flash('No tienes permiso para imprimir esta papeleta.', 'danger')
        return redirect(url_for('main.dashboard'))
    
    # Obtener nombre de aerolínea
    aerolinea_nombre = None
    if papeleta.aerolinea_id:
        aerolinea = Aerolinea.query.get(papeleta.aerolinea_id)
        if aerolinea:
            aerolinea_nombre = aerolinea.nombre
    
    # Obtener nombre de tarjeta (usando relación o búsqueda)
    tarjeta_nombre = None
    if hasattr(papeleta, 'tarjeta_rel') and papeleta.tarjeta_rel:
        tarjeta_nombre = papeleta.tarjeta_rel.nombre_tarjeta
    elif hasattr(papeleta, 'tarjeta_id') and papeleta.tarjeta_id:
        tarjeta_obj = TarjetaCorporativa.query.get(papeleta.tarjeta_id)
        if tarjeta_obj:
            tarjeta_nombre = tarjeta_obj.nombre_tarjeta
    elif papeleta.tarjeta:
        tarjeta_obj = TarjetaCorporativa.query.filter_by(numero_tarjeta=papeleta.tarjeta).first()
        if tarjeta_obj:
            tarjeta_nombre = tarjeta_obj.nombre_tarjeta
    
    # Obtener nombre de sucursal
    sucursal_nombre = None
    if papeleta.sucursal_id:
        sucursal = Sucursal.query.get(papeleta.sucursal_id)
        if sucursal:
            sucursal_nombre = sucursal.nombre
    elif current_user.sucursal:
        sucursal_nombre = current_user.sucursal.nombre
    
    return render_template('tickets/ticket_papeleta.html',
        papeleta=papeleta,
        aerolinea_nombre=aerolinea_nombre,
        tarjeta_nombre=tarjeta_nombre,
        sucursal_nombre=sucursal_nombre,
        now=datetime.now(TIMEZONE_MX)
    )

# =============================================================================
# RUTAS DE DESGLOSES
# =============================================================================

@main.route('/desgloses', methods=['GET'])
@login_required
def desgloses():
    return redirect(url_for('main.consulta_desgloses'))


@main.route('/desgloses/consulta', methods=['GET'])
@login_required
def consulta_desgloses():
    """Vista de consulta de desgloses con filtros"""
    from datetime import date
    
    # Base query - cada agente ve sus desgloses, admin ve todos
    if current_user.es_admin():
        query = Desglose.query
    else:
        query = Desglose.query.filter_by(usuario_id=current_user.id)
    
    # Aplicar filtros
    folio = request.args.get('folio', '').strip()
    empresa_id = request.args.get('empresa_id', type=int)
    clave = request.args.get('clave', '').strip()
    estatus = request.args.get('estatus', '').strip()
    
    if folio:
        query = query.filter(Desglose.folio == int(folio))
    if empresa_id:
        query = query.filter(Desglose.empresa_id == empresa_id)
    if clave:
        query = query.filter(Desglose.clave_reserva.ilike(f'%{clave}%'))
    if estatus:
        query = query.filter(Desglose.estatus == estatus)
    
    # Ordenar por folio descendente (más recientes primero)
    desgloses_list = query.order_by(Desglose.folio.desc()).limit(100).all()
    
    # Estadísticas
    total_desgloses = query.count()
    suma_total = sum(float(d.total or 0) for d in desgloses_list)
    
    # Desgloses de hoy
    hoy = fecha_mexico()
    if current_user.es_admin():
        desgloses_hoy = Desglose.query.filter(Desglose.fecha_emision == hoy).count()
    else:
        desgloses_hoy = Desglose.query.filter(
            Desglose.usuario_id == current_user.id,
            Desglose.fecha_emision == hoy
        ).count()
    
    # Lista de empresas para el filtro
    empresas_list = Empresa.query.filter_by(activa=True).order_by(Empresa.nombre_empresa).all()
    
    return render_template('consulta_desgloses.html',
                           desgloses=desgloses_list,
                           empresas=empresas_list,
                           total_desgloses=total_desgloses,
                           suma_total=suma_total,
                           desgloses_hoy=desgloses_hoy)


@main.route('/desgloses/nuevo', methods=['GET'])
@login_required
def nuevo_desglose_form():
    """Redirige a la calculadora de desglose"""
    return redirect(url_for('main.calculadora_desglose'))


# Ruta POST obsoleta - mantenida por compatibilidad pero redirige
@main.route('/desgloses/nuevo', methods=['POST'])
@login_required
def nuevo_desglose_post():
    """Redirige a la calculadora - formulario antiguo obsoleto"""
    return redirect(url_for('main.calculadora_desglose'))


@main.route('/desgloses/editar/<int:folio>', methods=['GET', 'POST'])
@login_required
def editar_desglose(folio):
    desglose = Desglose.query.get_or_404(folio)
    if request.method == 'POST':
        try:
            desglose.empresa_id = int(request.form.get('empresa_id'))
            desglose.aerolinea_id = int(request.form.get('aerolinea_id'))
            desglose.tarifa_base = float(request.form.get('tarifa_base') or 0)
            desglose.iva = float(request.form.get('iva') or 0)
            desglose.tua = float(request.form.get('tua') or 0)
            desglose.yr = float(request.form.get('yr') or 0)
            desglose.otros_cargos = float(request.form.get('otros_cargos') or 0)
            desglose.cargo_por_servicio = float(request.form.get('cargo_por_servicio') or 0)
            desglose.total = float(request.form.get('total') or 0)
            desglose.clave_reserva = request.form.get('clave_reserva')
            desglose.pasajero_nombre = request.form.get('pasajero_nombre') or None
            desglose.ruta = request.form.get('ruta') or None
            desglose.numero_boleto = request.form.get('numero_boleto', '').strip() or None
            db.session.commit()
            flash(f'Desglose {desglose.folio} actualizado con éxito.', 'success')
            return redirect(url_for('main.consulta_desgloses'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
    
    empresas_list = Empresa.query.filter_by(activa=True).order_by(Empresa.nombre_empresa).all()
    aerolineas_list = Aerolinea.query.order_by(Aerolinea.nombre).all()
    return render_template('desglose_edit.html', desglose=desglose, empresas=empresas_list, aerolineas=aerolineas_list)


@main.route('/desgloses/eliminar/<int:folio>')
@login_required
def eliminar_desglose(folio):
    desglose_a_eliminar = Desglose.query.get_or_404(folio)
    try:
        db.session.delete(desglose_a_eliminar)
        db.session.commit()
        flash(f'Desglose {folio} eliminado.', 'info')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al eliminar: {str(e)}', 'danger')
    return redirect(url_for('main.desgloses'))


# =============================================================================
# RUTAS DE EXPEDIENTES
# =============================================================================

@main.route('/expedientes')
@login_required
def expedientes():
    """
    Vista de expedientes - Agrupa documentos por clave de reservación.
    Un expediente es una vista virtual que muestra todos los documentos
    relacionados con una misma clave de reservación (desglose, papeleta, factura).
    
    Permisos:
    - Agente: Solo expedientes que él creó (sus propios desgloses y papeletas)
    - Gerente: Todos los de su sucursal
    - Director/Admin/Facturación: Todos
    """
    from datetime import date
    
    # Filtros
    q = request.args.get('q', '').strip().upper()
    estatus = request.args.get('estatus', '')
    empresa_id = request.args.get('empresa_id', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    
    # Determinar nivel de permisos
    rol_usuario = current_user.rol_relacion.nombre if current_user.rol_relacion else current_user.rol
    roles_ver_todo = ['administrador', 'admin', 'director', 'facturacion', 'gerente']
    
    puede_ver_todo = rol_usuario in roles_ver_todo
    
    # Query base para desgloses
    query_desgloses = db.session.query(
        Desglose.clave_reserva,
        Desglose.clave_sabre,
        Desglose.folio,
        Desglose.pasajero_nombre,
        Desglose.ruta,
        Desglose.total,
        Desglose.fecha_emision,
        Desglose.empresa_id,
        Desglose.aerolinea_id,
        Desglose.numero_boleto,
        Desglose.numero_factura,
        Desglose.sucursal_id,
        Desglose.usuario_id
    )
    
    # Query base para papeletas
    query_papeletas = db.session.query(
        Papeleta.clave_sabre,
        Papeleta.id,
        Papeleta.folio,
        Papeleta.facturar_a,
        Papeleta.total,
        Papeleta.fecha_venta,
        Papeleta.empresa_id,
        Papeleta.aerolinea_id,
        Papeleta.numero_factura,
        Papeleta.sucursal_id,
        Papeleta.usuario_id
    )
    
    # Aplicar filtros según rol
    if puede_ver_todo:
        # Admin/Director/Facturación/Gerente: Ve todos
        pass
    else:
        # Agente: Solo ve los que él creó
        query_desgloses = query_desgloses.filter(Desglose.usuario_id == current_user.id)
        query_papeletas = query_papeletas.filter(Papeleta.usuario_id == current_user.id)
    
    claves_desgloses = query_desgloses.all()
    claves_papeletas = query_papeletas.all()
    
    # Crear diccionario de expedientes
    expedientes_dict = {}
    
    # Procesar desgloses — agrupar por clave_sabre (vínculo con papeleta)
    for d in claves_desgloses:
        clave = (d.clave_sabre or d.clave_reserva or '').upper() or 'SIN-CLAVE'
        if clave not in expedientes_dict:
            expedientes_dict[clave] = {
                'clave': clave,
                'pasajero': d.pasajero_nombre or 'Sin nombre',
                'ruta': d.ruta,
                'total': float(d.total or 0),
                'fecha': d.fecha_emision,
                'empresa_id': d.empresa_id,
                'empresa': None,
                'aerolinea_id': d.aerolinea_id,
                'aerolinea': None,
                'desglose': d,
                'papeleta': None,
                'factura': d.numero_factura,  # Número de factura real (no boleto)
                'es_lowcost': False,
                'es_facturacion': True
            }
        else:
            # Actualizar con datos del desglose si ya existe
            expedientes_dict[clave]['desglose'] = d
            if d.pasajero_nombre:
                expedientes_dict[clave]['pasajero'] = d.pasajero_nombre
            if d.ruta:
                expedientes_dict[clave]['ruta'] = d.ruta
            if d.total:
                expedientes_dict[clave]['total'] = max(expedientes_dict[clave]['total'], float(d.total))
    
    # Procesar papeletas
    for p in claves_papeletas:
        clave = p.clave_sabre.upper() if p.clave_sabre else 'SIN-CLAVE'
        if clave not in expedientes_dict:
            expedientes_dict[clave] = {
                'clave': clave,
                'pasajero': p.facturar_a or 'Sin nombre',
                'ruta': None,
                'total': float(p.total or 0),
                'fecha': p.fecha_venta,
                'empresa_id': p.empresa_id,
                'empresa': None,
                'aerolinea_id': p.aerolinea_id,
                'aerolinea': None,
                'desglose': None,
                'papeleta': p,
                'factura': p.numero_factura,
                'es_lowcost': True,
                'es_facturacion': True
            }
        else:
            # Vincular papeleta al expediente existente
            expedientes_dict[clave]['papeleta'] = p
            expedientes_dict[clave]['es_lowcost'] = True
            if p.numero_factura:
                expedientes_dict[clave]['factura'] = p.numero_factura
            if not expedientes_dict[clave]['fecha'] and p.fecha_venta:
                expedientes_dict[clave]['fecha'] = p.fecha_venta
    
    # Obtener nombres de empresas y aerolíneas
    empresas_dict = {e.id: e.nombre_empresa for e in Empresa.query.all()}
    aerolineas_dict = {a.id: a.nombre for a in Aerolinea.query.all()}
    
    # Crear lista de expedientes con datos completos
    expedientes_list = []
    for clave, exp in expedientes_dict.items():
        # Obtener nombres
        exp['empresa'] = empresas_dict.get(exp['empresa_id'], '')
        exp['aerolinea'] = aerolineas_dict.get(exp['aerolinea_id'], 'Sin aerolínea')
        
        # Determinar requisitos
        exp['requiere_desglose'] = exp['es_facturacion']
        exp['requiere_papeleta'] = exp['es_lowcost']
        exp['requiere_factura'] = exp['es_facturacion']
        
        # Calcular progreso
        docs_requeridos = 0
        docs_completos = 0
        
        if exp['requiere_desglose']:
            docs_requeridos += 1
            if exp['desglose']:
                docs_completos += 1
        
        if exp['requiere_papeleta']:
            docs_requeridos += 1
            if exp['papeleta']:
                docs_completos += 1
        
        if exp['requiere_factura']:
            docs_requeridos += 1
            if exp['factura']:
                docs_completos += 1
        
        exp['docs_requeridos'] = docs_requeridos if docs_requeridos > 0 else 1
        exp['docs_completos'] = docs_completos
        exp['progreso'] = int((docs_completos / exp['docs_requeridos']) * 100) if exp['docs_requeridos'] > 0 else 0
        exp['completo'] = docs_completos >= docs_requeridos and docs_requeridos > 0
        
        expedientes_list.append(exp)
    
    # Aplicar filtros
    if q:
        expedientes_list = [e for e in expedientes_list if q in e['clave'] or q in (e['pasajero'] or '').upper()]
    
    if estatus == 'completo':
        expedientes_list = [e for e in expedientes_list if e['completo']]
    elif estatus == 'pendiente':
        expedientes_list = [e for e in expedientes_list if not e['completo']]
    
    if empresa_id:
        expedientes_list = [e for e in expedientes_list if e['empresa_id'] == int(empresa_id)]
    
    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
        expedientes_list = [e for e in expedientes_list if e['fecha'] and e['fecha'] >= fecha_desde_dt]
    
    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
        expedientes_list = [e for e in expedientes_list if e['fecha'] and e['fecha'] <= fecha_hasta_dt]
    
    # Ordenar por fecha más reciente
    expedientes_list.sort(key=lambda x: x['fecha'] or date.min, reverse=True)
    
    # Estadísticas
    hoy = fecha_mexico()
    stats = {
        'completos': len([e for e in expedientes_list if e['completo']]),
        'pendientes': len([e for e in expedientes_list if not e['completo']]),
        'hoy': len([e for e in expedientes_list if e['fecha'] == hoy])
    }
    
    # Empresas para filtro
    empresas_filtro = Empresa.query.filter_by(activa=True).order_by(Empresa.nombre_empresa).all()
    
    return render_template('expedientes.html',
                           expedientes=expedientes_list,
                           stats=stats,
                           empresas=empresas_filtro)


# =============================================================================
# RUTAS DE EMPRESAS
# =============================================================================

@main.route('/empresas', methods=['GET'])
@login_required
def empresas():
    lista_empresas = Empresa.query.order_by(Empresa.nombre_empresa).all()
    return render_template('empresas.html', empresas_registradas=lista_empresas)


@main.route('/empresas/nueva', methods=['GET', 'POST'])
@login_required
def nueva_empresa():
    if current_user.rol_relacion.nombre not in ['administrador', 'admin', 'director']:
        return redirect(url_for('main.dashboard'))
    
    # Si es GET, mostrar formulario vacío
    if request.method == 'GET':
        return render_template('empresa_edit.html', empresa=None)
    
    # Si es POST, procesar el formulario
    try:
        nombre = request.form.get('nombre_empresa')
        if not nombre:
            flash('El nombre de la empresa no puede estar vacío.', 'warning')
            return redirect(url_for('main.empresas'))
        
        # Crear empresa con campos básicos y de esquema
        nueva = Empresa(
            nombre_empresa=nombre,
            sucursal_id=current_user.sucursal_id,
            tipo_cliente=request.form.get('tipo_cliente', 'corporativo'),
            rfc=request.form.get('rfc') or None,
            razon_social=request.form.get('razon_social') or None,
            contacto_nombre=request.form.get('contacto_nombre') or None,
            contacto_email=request.form.get('contacto_email') or None,
            contacto_telefono=request.form.get('contacto_telefono') or None,
            # Campos de esquema de facturación
            esquema_facturacion=request.form.get('esquema_facturacion', 'cargo_servicio'),
            bonificacion_porcentaje=float(request.form.get('bonificacion_porcentaje', 0)) / 100 if request.form.get('bonificacion_porcentaje') else None,
            bonificacion_aplica_sobre=request.form.get('bonificacion_aplica_sobre', 'tarifa_base'),
            requiere_desglose=request.form.get('requiere_desglose') == 'on',
            tipo_desglose=request.form.get('tipo_desglose', 'standard'),
            instrucciones_cotizacion=request.form.get('instrucciones_cotizacion') or None,
            notas_emision=request.form.get('notas_emision') or None,
            personas_autorizadas=request.form.get('personas_autorizadas') or None,
            encargado_cuenta=request.form.get('encargado_cuenta') or None,
        )
        db.session.add(nueva)
        db.session.flush()

        # Agregar cargos por servicio
        cargo_facturado = request.form.get('cargoServicioFacturado')
        if cargo_facturado:
            db.session.add(CargoServicio(
                empresa_id=nueva.id, 
                tipo='visible', 
                monto=float(cargo_facturado),
                tipo_servicio=request.form.get('cargo_tipo_servicio', 'todos')
            ))
        cargo_oculto = request.form.get('cargoServicioOculto')
        if cargo_oculto:
            db.session.add(CargoServicio(
                empresa_id=nueva.id, 
                tipo='oculto', 
                monto=float(cargo_oculto),
                tipo_servicio=request.form.get('cargo_tipo_servicio', 'todos')
            ))
        
        # Agregar descuento
        monto_descuento = request.form.get('montoDescuento')
        if monto_descuento:
            db.session.add(Descuento(
                empresa_id=nueva.id, 
                tipo=request.form.get('descuento_tipo', 'fijo'),
                valor=float(monto_descuento),
                aplica_sobre=request.form.get('descuento_aplica_sobre', 'tarifa_base'),
                incluye_iva=request.form.get('descuento_incluye_iva') == 'on'
            ))
        
        # Agregar tarifas fijas
        tarifa_origenes = request.form.getlist('tarifa_origen[]')
        tarifa_destinos = request.form.getlist('tarifa_destino[]')
        tarifa_montos = request.form.getlist('tarifa_monto[]')
        tarifa_tipos = request.form.getlist('tarifa_tipo_viaje[]')
        
        for i, monto in enumerate(tarifa_montos):
            if monto:
                es_default = request.form.get(f'tarifa_default_{i}') == '1'
                db.session.add(TarifaFija(
                    empresa_id=nueva.id,
                    ruta_origen=tarifa_origenes[i].upper() if i < len(tarifa_origenes) and tarifa_origenes[i] else None,
                    ruta_destino=tarifa_destinos[i].upper() if i < len(tarifa_destinos) and tarifa_destinos[i] else None,
                    monto=float(monto),
                    tipo_viaje=tarifa_tipos[i] if i < len(tarifa_tipos) else 'sencillo',
                    es_ruta_default=es_default
                ))

        db.session.commit()
        flash(f'Empresa "{nombre}" registrada con éxito.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error al registrar empresa: {str(e)}', 'danger')
    return redirect(url_for('main.empresas'))


@main.route('/empresas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_empresa(id):
    if current_user.rol_relacion.nombre not in ['administrador', 'admin', 'director']:
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    empresa = Empresa.query.get_or_404(id)
    if request.method == 'POST':
        try:
            # Actualizar campos básicos
            empresa.nombre_empresa = request.form.get('nombre_empresa')
            empresa.tipo_cliente = request.form.get('tipo_cliente', 'corporativo')
            empresa.rfc = request.form.get('rfc') or None
            empresa.razon_social = request.form.get('razon_social') or None
            empresa.contacto_nombre = request.form.get('contacto_nombre') or None
            empresa.contacto_email = request.form.get('contacto_email') or None
            empresa.contacto_telefono = request.form.get('contacto_telefono') or None
            
            # Actualizar campos de esquema de facturación
            empresa.esquema_facturacion = request.form.get('esquema_facturacion', 'cargo_servicio')
            bonif_pct = request.form.get('bonificacion_porcentaje')
            empresa.bonificacion_porcentaje = float(bonif_pct) / 100 if bonif_pct else None
            empresa.bonificacion_aplica_sobre = request.form.get('bonificacion_aplica_sobre', 'tarifa_base')
            empresa.requiere_desglose = request.form.get('requiere_desglose') == 'on'
            empresa.tipo_desglose = request.form.get('tipo_desglose', 'standard')
            empresa.instrucciones_cotizacion = request.form.get('instrucciones_cotizacion') or None
            empresa.notas_emision = request.form.get('notas_emision') or None
            empresa.personas_autorizadas = request.form.get('personas_autorizadas') or None
            empresa.encargado_cuenta = request.form.get('encargado_cuenta') or None
            
            # Eliminar cargos, descuentos y tarifas existentes
            CargoServicio.query.filter_by(empresa_id=id).delete()
            Descuento.query.filter_by(empresa_id=id).delete()
            TarifaFija.query.filter_by(empresa_id=id).delete()

            # Agregar nuevos cargos por servicio
            cargo_facturado = request.form.get('cargoServicioFacturado')
            if cargo_facturado:
                db.session.add(CargoServicio(
                    empresa_id=id, 
                    tipo='visible', 
                    monto=float(cargo_facturado),
                    tipo_servicio=request.form.get('cargo_tipo_servicio', 'todos')
                ))
            cargo_oculto = request.form.get('cargoServicioOculto')
            if cargo_oculto:
                db.session.add(CargoServicio(
                    empresa_id=id, 
                    tipo='oculto', 
                    monto=float(cargo_oculto),
                    tipo_servicio=request.form.get('cargo_tipo_servicio', 'todos')
                ))
            
            # Agregar nuevo descuento
            monto_descuento = request.form.get('montoDescuento')
            if monto_descuento:
                db.session.add(Descuento(
                    empresa_id=id, 
                    tipo=request.form.get('descuento_tipo', 'fijo'),
                    valor=float(monto_descuento),
                    aplica_sobre=request.form.get('descuento_aplica_sobre', 'tarifa_base'),
                    incluye_iva=request.form.get('descuento_incluye_iva') == 'on'
                ))
            
            # Agregar nuevas tarifas fijas
            tarifa_origenes = request.form.getlist('tarifa_origen[]')
            tarifa_destinos = request.form.getlist('tarifa_destino[]')
            tarifa_montos = request.form.getlist('tarifa_monto[]')
            tarifa_tipos = request.form.getlist('tarifa_tipo_viaje[]')
            
            for i, monto in enumerate(tarifa_montos):
                if monto:
                    es_default = request.form.get(f'tarifa_default_{i}') == '1'
                    db.session.add(TarifaFija(
                        empresa_id=id,
                        ruta_origen=tarifa_origenes[i].upper() if i < len(tarifa_origenes) and tarifa_origenes[i] else None,
                        ruta_destino=tarifa_destinos[i].upper() if i < len(tarifa_destinos) and tarifa_destinos[i] else None,
                        monto=float(monto),
                        tipo_viaje=tarifa_tipos[i] if i < len(tarifa_tipos) else 'sencillo',
                        es_ruta_default=es_default
                    ))

            db.session.commit()
            flash(f'Empresa "{empresa.nombre_empresa}" actualizada.', 'success')
            return redirect(url_for('main.empresas'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
    return render_template('empresa_edit.html', empresa=empresa)


@main.route('/empresas/eliminar/<int:id>')
@login_required
def eliminar_empresa(id):
    if current_user.rol_relacion.nombre not in ['administrador', 'admin', 'director']:
        return redirect(url_for('main.dashboard'))
    empresa_a_eliminar = Empresa.query.get_or_404(id)
    db.session.delete(empresa_a_eliminar)
    db.session.commit()
    flash(f'Empresa "{empresa_a_eliminar.nombre_empresa}" eliminada.', 'info')
    return redirect(url_for('main.empresas'))


# =============================================================================
# RUTAS DE SUCURSALES
# =============================================================================

@main.route('/sucursales')
@login_required
def sucursales():
    if not current_user.es_admin():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    lista = Sucursal.query.order_by(Sucursal.nombre).all()
    return render_template('sucursales.html', sucursales=lista)


@main.route('/sucursales/nueva', methods=['POST'])
@login_required
def nueva_sucursal():
    if not current_user.es_admin():
        return redirect(url_for('main.dashboard'))
    try:
        nueva = Sucursal(
            nombre=request.form.get('nombre', '').strip(),
            ciudad=request.form.get('ciudad', '').strip(),
            direccion=request.form.get('direccion', '').strip() or None,
            telefono=request.form.get('telefono', '').strip() or None
        )
        db.session.add(nueva)
        db.session.commit()
        flash(f'Sucursal "{nueva.nombre}" creada.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'danger')
    return redirect(url_for('main.sucursales'))


@main.route('/sucursales/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_sucursal(id):
    if not current_user.es_admin():
        flash('Acceso no autorizado.', 'danger')
        return redirect(url_for('main.dashboard'))
    sucursal = Sucursal.query.get_or_404(id)
    if request.method == 'POST':
        try:
            sucursal.nombre = request.form.get('nombre', '').strip()
            sucursal.ciudad = request.form.get('ciudad', '').strip()
            sucursal.direccion = request.form.get('direccion', '').strip() or None
            sucursal.telefono = request.form.get('telefono', '').strip() or None
            sucursal.activa = request.form.get('activa') == 'on'
            db.session.commit()
            flash(f'Sucursal "{sucursal.nombre}" actualizada.', 'success')
            return redirect(url_for('main.sucursales'))
        except Exception as e:
            db.session.rollback()
            flash(f'Error: {str(e)}', 'danger')
    return render_template('sucursal_edit.html', sucursal=sucursal)


@main.route('/sucursales/eliminar/<int:id>')
@login_required
def eliminar_sucursal(id):
    if not current_user.es_admin():
        return redirect(url_for('main.dashboard'))
    sucursal = Sucursal.query.get_or_404(id)
    if Usuario.query.filter_by(sucursal_id=id).count() > 0:
        flash('No se puede eliminar: la sucursal tiene usuarios asignados.', 'warning')
        return redirect(url_for('main.sucursales'))
    nombre = sucursal.nombre
    db.session.delete(sucursal)
    db.session.commit()
    flash(f'Sucursal "{nombre}" eliminada.', 'info')
    return redirect(url_for('main.sucursales'))

# Agregar esta ruta al archivo routes.py

# =============================================================================
# API - ACTUALIZAR NÚMERO DE FACTURA
# =============================================================================

@main.route('/api/papeleta/<int:id>/factura', methods=['POST'])
@login_required
def actualizar_factura_papeleta(id):
    """Actualiza el número de factura de una papeleta (solo admin/contabilidad)"""
    if not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No tienes permisos para esta acción'}), 403
    
    papeleta = Papeleta.query.get_or_404(id)
    data = request.get_json()
    
    numero_factura = data.get('numero_factura', '').strip()
    
    # Guardar valor anterior para auditoría
    valor_anterior = papeleta.numero_factura
    
    # Actualizar
    papeleta.numero_factura = numero_factura if numero_factura else None
    
    try:
        db.session.commit()
        
        # Log de auditoría (opcional)
        # audit_log = AuditLog(
        #     tabla_nombre='papeletas',
        #     registro_id=str(id),
        #     accion='UPDATE',
        #     datos_anteriores={'numero_factura': valor_anterior},
        #     datos_nuevos={'numero_factura': papeleta.numero_factura},
        #     campos_modificados=['numero_factura'],
        #     usuario_id=current_user.id,
        #     usuario_email=current_user.correo
        # )
        # db.session.add(audit_log)
        # db.session.commit()
        
        return jsonify({
            'success': True,
            'numero_factura': papeleta.numero_factura,
            'message': 'Factura actualizada correctamente'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# =============================================================================
# API - ACTUALIZAR PDF DEL BOLETO
# =============================================================================

@main.route('/api/papeleta/<int:id>/archivo', methods=['POST'])
@login_required
def actualizar_archivo_papeleta(id):
    """Actualiza el archivo PDF del boleto de una papeleta"""
    import os
    from werkzeug.utils import secure_filename
    from flask import current_app
    
    papeleta = Papeleta.query.get_or_404(id)
    
    # Verificar que el usuario sea el dueño o admin
    if papeleta.usuario_id != current_user.id and not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No tienes permisos para modificar esta papeleta'}), 403
    
    if 'archivo_boleto' not in request.files:
        return jsonify({'success': False, 'error': 'No se envió ningún archivo'}), 400
    
    archivo = request.files['archivo_boleto']
    if archivo.filename == '':
        return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400
    
    # Validar extensión
    if not archivo.filename.lower().endswith('.pdf'):
        return jsonify({'success': False, 'error': 'Solo se permiten archivos PDF'}), 400
    
    try:
        UPLOAD_FOLDER = os.path.join(current_app.static_folder, 'uploads', 'boletos')
        
        # Crear carpeta si no existe
        if not os.path.exists(UPLOAD_FOLDER):
            os.makedirs(UPLOAD_FOLDER)
        
        # Eliminar archivo anterior si existe
        if papeleta.archivo_boleto:
            archivo_anterior = os.path.join(UPLOAD_FOLDER, papeleta.archivo_boleto)
            if os.path.exists(archivo_anterior):
                os.remove(archivo_anterior)
        
        # Guardar nuevo archivo
        nombre_archivo = f"{papeleta.folio}_boleto.pdf"
        nombre_archivo = secure_filename(nombre_archivo)
        ruta_archivo = os.path.join(UPLOAD_FOLDER, nombre_archivo)
        archivo.save(ruta_archivo)
        
        papeleta.archivo_boleto = nombre_archivo
        db.session.commit()
        
        return jsonify({
            'success': True,
            'archivo': nombre_archivo,
            'message': 'Archivo actualizado correctamente'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/papeleta/<int:id>/archivo', methods=['DELETE'])
@login_required
def eliminar_archivo_papeleta(id):
    """Elimina el archivo PDF del boleto de una papeleta (solo admin)"""
    import os
    from flask import current_app
    
    if not current_user.es_admin():
        return jsonify({'success': False, 'error': 'Solo administración puede eliminar archivos'}), 403
    
    papeleta = Papeleta.query.get_or_404(id)
    
    if not papeleta.archivo_boleto:
        return jsonify({'success': False, 'error': 'La papeleta no tiene archivo adjunto'}), 400
    
    try:
        UPLOAD_FOLDER = os.path.join(current_app.static_folder, 'uploads', 'boletos')
        archivo_path = os.path.join(UPLOAD_FOLDER, papeleta.archivo_boleto)
        
        if os.path.exists(archivo_path):
            os.remove(archivo_path)
        
        papeleta.archivo_boleto = None
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Archivo eliminado correctamente'
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

# =============================================================================
# RUTAS DE REPORTES DE VENTAS
# Agregar a routes.py
# =============================================================================

# Agregar estos imports al inicio de routes.py:
# 

# =============================================================================
# CONSULTA DE REPORTES DE VENTAS
# =============================================================================

def _generar_folio_reporte():
    """Genera folio automático RV-YYYY-NNNN"""
    from datetime import date
    anio = date.today().year
    ultimo = ReporteVenta.query.filter(
        ReporteVenta.folio.like(f'RV-{anio}-%')
    ).order_by(ReporteVenta.folio.desc()).first()
    
    if ultimo and ultimo.folio:
        try:
            num = int(ultimo.folio.split('-')[-1]) + 1
        except (ValueError, IndexError):
            num = 1
    else:
        num = 1
    return f'RV-{anio}-{num:04d}'


def _recalcular_totales_reporte(reporte):
    """Recalcula todos los totales del reporte sumando sus detalles"""
    detalles = reporte.detalles.all()
    
    reporte.total_bsp = sum(float(d.monto_bsp or 0) for d in detalles)
    reporte.total_volaris = sum(float(d.monto_volaris or 0) for d in detalles)
    reporte.total_vivaerobus = sum(float(d.monto_vivaerobus or 0) for d in detalles)
    reporte.total_compra_tc = sum(float(d.monto_compra_tc or 0) for d in detalles)
    reporte.total_cargo_expedicion = sum(float(d.cargo_expedicion or 0) for d in detalles)
    reporte.total_cargo_315 = sum(float(d.cargo_315 or 0) for d in detalles)
    reporte.total_seguros = sum(float(d.monto_seguros or 0) for d in detalles)
    reporte.total_hoteles_paquetes = sum(float(d.monto_hoteles_paquetes or 0) for d in detalles)
    reporte.total_transporte_terrestre = sum(float(d.monto_transporte_terrestre or 0) for d in detalles)
    reporte.total_pago_directo_tc = sum(float(d.pago_directo_tc or 0) for d in detalles)
    reporte.total_voucher_tc = sum(float(d.voucher_tc or 0) for d in detalles)
    reporte.total_efectivo = sum(float(d.efectivo or 0) for d in detalles)
    reporte.total_general = sum(float(d.total_linea or 0) for d in detalles)
    reporte.total_boletos = sum(int(d.num_boletos or 0) for d in detalles)
    reporte.total_recibos = len(detalles)


@main.route('/reportes-ventas')
@login_required
def reportes_ventas():
    """Lista de reportes de ventas"""
    # Filtros
    fecha_desde = request.args.get('desde')
    fecha_hasta = request.args.get('hasta')
    estatus = request.args.get('estatus')
    
    query = ReporteVenta.query
    
    # Si no es admin, solo ver sus reportes
    if not current_user.es_admin():
        query = query.filter_by(usuario_id=current_user.id)
    
    if fecha_desde:
        query = query.filter(ReporteVenta.fecha >= fecha_desde)
    if fecha_hasta:
        query = query.filter(ReporteVenta.fecha <= fecha_hasta)
    if estatus:
        query = query.filter_by(estatus=estatus)
    
    reportes = query.order_by(ReporteVenta.fecha.desc(), ReporteVenta.id.desc()).all()
    
    return render_template('reportes_ventas/lista.html', reportes=reportes)


@main.route('/reportes-ventas/nuevo', methods=['GET', 'POST'])
@login_required
def nuevo_reporte_venta():
    """Crear nuevo reporte de ventas"""
    from datetime import date
    
    if request.method == 'POST':
        try:
            fecha = request.form.get('fecha')
            if not fecha:
                fecha = fecha_mexico()
            else:
                fecha = datetime.strptime(fecha, '%Y-%m-%d').date()
            
            # Verificar si ya existe reporte para esta fecha y usuario
            existente = ReporteVenta.query.filter_by(
                fecha=fecha,
                usuario_id=current_user.id
            ).first()
            
            # Si existe pero YA TIENE VALE, permitir crear uno nuevo
            # Si existe y NO tiene vale, redirigir al existente para editarlo
            if existente:
                if existente.entrega_corte:
                    # Ya tiene vale, permitir crear otro reporte
                    pass  # Continúa con la creación
                else:
                    # No tiene vale, redirigir a editar el existente
                    flash(f'Ya existe un reporte para el {fecha.strftime("%d/%m/%Y")}. Folio: {existente.folio}', 'warning')
                    return redirect(url_for('main.editar_reporte_venta', id=existente.id))
            
            reporte = ReporteVenta(
                folio=_generar_folio_reporte(),
                fecha=fecha,
                usuario_id=current_user.id,
                sucursal_id=current_user.sucursal_id,
                notas=request.form.get('notas', '')
            )
            db.session.add(reporte)
            db.session.commit()
            
            flash(f'Reporte {reporte.folio} creado exitosamente.', 'success')
            return redirect(url_for('main.editar_reporte_venta', id=reporte.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear reporte: {str(e)}', 'danger')
    
    return render_template('reportes_ventas/nuevo.html', fecha_hoy=fecha_mexico())


@main.route('/reportes-ventas/<int:id>')
@login_required
def ver_reporte_venta(id):
    """Ver detalle de reporte de ventas"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    # Verificar acceso
    if not current_user.es_admin() and reporte.usuario_id != current_user.id:
        flash('No tienes acceso a este reporte.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    detalles = reporte.detalles.order_by(DetalleReporteVenta.orden).all()
    
    # Contar boletos por aerolínea
    boletos_aerolinea = db.session.query(
        DetalleReporteVenta.clave_aerolinea,
        func.sum(DetalleReporteVenta.num_boletos)
    ).filter_by(reporte_id=id).group_by(DetalleReporteVenta.clave_aerolinea).all()
    
    return render_template('reportes_ventas/ver.html', 
                          reporte=reporte, 
                          detalles=detalles,
                          boletos_aerolinea=dict(boletos_aerolinea))


@main.route('/reportes-ventas/<int:id>/editar', methods=['GET', 'POST'])
@login_required
def editar_reporte_venta(id):
    """Editar reporte de ventas"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    # Verificar acceso y estado
    if not current_user.es_admin() and reporte.usuario_id != current_user.id:
        flash('No tienes acceso a este reporte.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    if not reporte.puede_editar and not current_user.es_admin():
        flash('Este reporte ya no puede ser editado.', 'warning')
        return redirect(url_for('main.ver_reporte_venta', id=id))
    
    if request.method == 'POST':
        try:
            # Actualizar datos de depósitos
            reporte.deposito_pesos_efectivo = float(request.form.get('deposito_pesos_efectivo') or 0)
            reporte.deposito_dolares_efectivo = float(request.form.get('deposito_dolares_efectivo') or 0)
            reporte.deposito_pesos_cheques = float(request.form.get('deposito_pesos_cheques') or 0)
            reporte.tipo_cambio = float(request.form.get('tipo_cambio') or 0)
            reporte.cuenta_deposito = request.form.get('cuenta_deposito', '')
            reporte.notas = request.form.get('notas', '')
            
            db.session.commit()
            flash('Reporte actualizado.', 'success')
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al actualizar: {str(e)}', 'danger')
    
    # Obtener papeletas de MOSTRADOR del usuario que no están en ningún reporte
    # Solo mostrador = sin empresa asignada o facturar_a = 'MOSTRADOR'
    fecha_limite = reporte.fecha - timedelta(days=30)
    
    papeletas_disponibles = Papeleta.query.filter(
        Papeleta.usuario_id == reporte.usuario_id,
        Papeleta.reporte_venta_id.is_(None),
        Papeleta.fecha_venta >= fecha_limite,
        Papeleta.fecha_venta <= reporte.fecha,
        db.or_(
            Papeleta.empresa_id.is_(None),
            Papeleta.facturar_a.ilike('%MOSTRADOR%')
        )
    ).order_by(Papeleta.fecha_venta.desc(), Papeleta.id).all()
    
    # Papeletas ya agregadas
    detalles = reporte.detalles.order_by(DetalleReporteVenta.orden).all()
    
    aerolineas = Aerolinea.query.filter_by(activa=True).order_by(Aerolinea.nombre).all()
    
    return render_template('reportes_ventas/editar.html',
                          reporte=reporte,
                          detalles=detalles,
                          papeletas_disponibles=papeletas_disponibles,
                          aerolineas=aerolineas)


@main.route('/reportes-ventas/<int:id>/enviar', methods=['POST'])
@login_required
def enviar_reporte_venta(id):
    """Enviar reporte para aprobación"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    if reporte.usuario_id != current_user.id and not current_user.es_admin():
        flash('No tienes permiso para enviar este reporte.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    if reporte.estatus != 'borrador':
        flash('Este reporte ya fue enviado.', 'warning')
        return redirect(url_for('main.ver_reporte_venta', id=id))
    
    if reporte.total_recibos == 0:
        flash('No puedes enviar un reporte vacío.', 'warning')
        return redirect(url_for('main.editar_reporte_venta', id=id))
    
    reporte.estatus = 'enviado'
    reporte.fecha_envio = datetime.utcnow()
    db.session.commit()
    
    flash(f'Reporte {reporte.folio} enviado para revisión.', 'success')
    return redirect(url_for('main.ver_reporte_venta', id=id))


@main.route('/reportes-ventas/<int:id>/aprobar', methods=['POST'])
@login_required
def aprobar_reporte_venta(id):
    """Aprobar reporte de ventas (solo admin)"""
    if not current_user.es_admin():
        flash('No tienes permiso para aprobar reportes.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    reporte = ReporteVenta.query.get_or_404(id)
    
    if reporte.estatus != 'enviado':
        flash('Solo se pueden aprobar reportes enviados.', 'warning')
        return redirect(url_for('main.ver_reporte_venta', id=id))
    
    reporte.estatus = 'aprobado'
    reporte.fecha_aprobacion = datetime.utcnow()
    reporte.aprobado_por = current_user.id
    
    # === AUTO-APROBAR VALE DE ENTREGA ASOCIADO ===
    # Si el reporte tiene un vale de entrega vinculado y está en estatus 'depositado',
    # se aprueba automáticamente
    if reporte.entrega_corte:
        entrega = reporte.entrega_corte
        if hasattr(entrega, '__iter__'):
            # Si es una lista (múltiples entregas)
            for e in entrega:
                if e.estatus == 'depositado' and not e.fecha_revision:
                    e.revisar_y_aprobar(
                        director_id=current_user.id,
                        aprobado=True,
                        notas=f'Aprobado automáticamente al aprobar reporte {reporte.folio}'
                    )
        else:
            # Si es un solo objeto
            if entrega.estatus == 'depositado' and not entrega.fecha_revision:
                entrega.revisar_y_aprobar(
                    director_id=current_user.id,
                    aprobado=True,
                    notas=f'Aprobado automáticamente al aprobar reporte {reporte.folio}'
                )
    
    db.session.commit()
    
    flash(f'Reporte {reporte.folio} aprobado.', 'success')
    return redirect(url_for('main.ver_reporte_venta', id=id))


@main.route('/reportes-ventas/<int:id>/rechazar', methods=['POST'])
@login_required
def rechazar_reporte_venta(id):
    """Rechazar reporte de ventas (solo admin)"""
    if not current_user.es_admin():
        flash('No tienes permiso para rechazar reportes.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    reporte = ReporteVenta.query.get_or_404(id)
    motivo = request.form.get('motivo', '')
    
    reporte.estatus = 'rechazado'
    reporte.notas = f"RECHAZADO: {motivo}\n\n{reporte.notas or ''}"
    db.session.commit()
    
    flash(f'Reporte {reporte.folio} rechazado.', 'info')
    return redirect(url_for('main.ver_reporte_venta', id=id))


@main.route('/reportes-ventas/<int:id>/reabrir', methods=['POST'])
@login_required
def reabrir_reporte_venta(id):
    """Reabrir reporte rechazado"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    if reporte.usuario_id != current_user.id and not current_user.es_admin():
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    if reporte.estatus not in ['rechazado', 'enviado']:
        flash('No se puede reabrir este reporte.', 'warning')
        return redirect(url_for('main.ver_reporte_venta', id=id))
    
    reporte.estatus = 'borrador'
    db.session.commit()
    
    flash(f'Reporte {reporte.folio} reabierto para edición.', 'success')
    return redirect(url_for('main.editar_reporte_venta', id=id))


@main.route('/reportes-ventas/<int:id>/eliminar', methods=['POST'])
@login_required
def eliminar_reporte_venta(id):
    """Eliminar reporte de ventas"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    if reporte.usuario_id != current_user.id and not current_user.es_admin():
        flash('No tienes permiso.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    if reporte.estatus == 'aprobado':
        flash('No se puede eliminar un reporte aprobado.', 'danger')
        return redirect(url_for('main.reportes_ventas'))
    
    # Liberar papeletas asociadas
    Papeleta.query.filter_by(reporte_venta_id=id).update({'reporte_venta_id': None})
    
    folio = reporte.folio
    db.session.delete(reporte)
    db.session.commit()
    
    flash(f'Reporte {folio} eliminado.', 'info')
    return redirect(url_for('main.reportes_ventas'))


# =============================================================================
# API - DETALLE DE REPORTE
# =============================================================================

@main.route('/api/reporte-venta/<int:id>/agregar-linea', methods=['POST'])
@login_required
def agregar_linea_reporte(id):
    """Agregar línea al reporte de ventas"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    if not reporte.puede_editar:
        return jsonify({'success': False, 'error': 'Reporte no editable'}), 400
    
    data = request.get_json()
    
    try:
        # Obtener siguiente orden
        max_orden = db.session.query(func.max(DetalleReporteVenta.orden)).filter_by(reporte_id=id).scalar() or 0
        
        detalle = DetalleReporteVenta(
            reporte_id=id,
            papeleta_id=data.get('papeleta_id'),
            clave_aerolinea=data.get('clave_aerolinea', ''),
            num_boletos=int(data.get('num_boletos', 1)),
            reserva=data.get('reserva', ''),
            num_recibo=data.get('num_recibo', ''),
            num_papeleta=data.get('num_papeleta', ''),
            monto_bsp=float(data.get('monto_bsp', 0)),
            monto_volaris=float(data.get('monto_volaris', 0)),
            monto_vivaerobus=float(data.get('monto_vivaerobus', 0)),
            monto_compra_tc=float(data.get('monto_compra_tc', 0)),
            cargo_expedicion=float(data.get('cargo_expedicion', 0)),
            cargo_315=float(data.get('cargo_315', 0)),
            monto_seguros=float(data.get('monto_seguros', 0)),
            monto_hoteles_paquetes=float(data.get('monto_hoteles_paquetes', 0)),
            monto_transporte_terrestre=float(data.get('monto_transporte_terrestre', 0)),
            pago_directo_tc=float(data.get('pago_directo_tc', 0)),
            voucher_tc=float(data.get('voucher_tc', 0)),
            efectivo=float(data.get('efectivo', 0)),
            total_linea=float(data.get('total_linea', 0)),
            orden=max_orden + 1
        )
        
        db.session.add(detalle)
        
        # Si viene de una papeleta, asociarla
        if data.get('papeleta_id'):
            papeleta = Papeleta.query.get(data['papeleta_id'])
            if papeleta:
                papeleta.reporte_venta_id = id
        
        # Recalcular totales del reporte
        _recalcular_totales_reporte(reporte)
        db.session.commit()
        
        # Recargar reporte para obtener totales actualizados
        db.session.refresh(reporte)
        
        return jsonify({
            'success': True,
            'detalle_id': detalle.id,
            'totales': {
                'total_general': float(reporte.total_general or 0),
                'total_efectivo': float(reporte.total_efectivo or 0),
                'total_voucher_tc': float(reporte.total_voucher_tc or 0),
                'total_boletos': reporte.total_boletos,
                'total_recibos': reporte.total_recibos
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/reporte-venta/<int:id>/agregar-papeleta', methods=['POST'])
@login_required
def agregar_papeleta_reporte(id):
    """Agregar papeleta existente al reporte"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    if not reporte.puede_editar:
        return jsonify({'success': False, 'error': 'Reporte no editable'}), 400
    
    data = request.get_json()
    papeleta_id = data.get('papeleta_id')
    
    if not papeleta_id:
        return jsonify({'success': False, 'error': 'Papeleta no especificada'}), 400
    
    papeleta = Papeleta.query.get_or_404(papeleta_id)
    
    # Verificar que no esté ya en otro reporte
    if papeleta.reporte_venta_id and papeleta.reporte_venta_id != id:
        return jsonify({'success': False, 'error': 'Papeleta ya está en otro reporte'}), 400
    
    try:
        # Determinar clave de aerolínea
        clave_aerolinea = ''
        nombre_aerolinea = ''
        if papeleta.aerolinea:
            clave_aerolinea = papeleta.aerolinea.codigo_iata or papeleta.aerolinea.nombre[:2].upper()
            nombre_aerolinea = (papeleta.aerolinea.nombre or '').upper()
        
        # Obtener nombre de la tarjeta corporativa usada
        nombre_tarjeta = (papeleta.tarjeta or '').upper()
        if papeleta.tarjeta_rel:
            nombre_tarjeta = (papeleta.tarjeta_rel.nombre_tarjeta or '').upper()
        
        # Forma de pago del cliente
        forma_pago = (papeleta.forma_pago or '').lower()
        total_ticket = float(papeleta.total_ticket or 0)
        total = float(papeleta.total or 0)
        
        # Inicializar montos
        monto_volaris = 0
        monto_vivaerobus = 0
        monto_compra_tc = 0
        voucher_tc = 0
        efectivo = 0
        pago_directo_tc = 0
        
        # ============================================================
        # CLASIFICACIÓN POR TARJETA CORPORATIVA USADA
        # ============================================================
        # - CREDITO VOLARIS (CVOL) → Columna Volaris
        # - CREDITO VIVAAEROBUS (CVIV) → Columna VivaAerobus  
        # - Otras tarjetas (VISA, MASTER, INVEX, etc.) → Columna Compra TC
        # ============================================================
        
        es_tarjeta_volaris = (
            'VOLARIS' in nombre_tarjeta or 
            'CVOL' in nombre_tarjeta or
            nombre_tarjeta.startswith('CREDITO VOL')
        )
        
        es_tarjeta_vivaerobus = (
            'VIVA' in nombre_tarjeta or 
            'CVIV' in nombre_tarjeta or
            nombre_tarjeta.startswith('CREDITO VIV')
        )
        
        # Determinar si el cliente pagó en efectivo o con voucher
        es_pago_efectivo = (
            'efectivo' in forma_pago or 
            'depósito' in forma_pago or 
            'deposito' in forma_pago or
            'transferencia' in forma_pago or
            'contado' in forma_pago
        )
        
        # Clasificar según la tarjeta corporativa
        if es_tarjeta_volaris:
            monto_volaris = total_ticket
        elif es_tarjeta_vivaerobus:
            monto_vivaerobus = total_ticket
        else:
            # Cualquier otra tarjeta va a Compra TC
            monto_compra_tc = total_ticket
        
        # El total va a efectivo o voucher según cómo pagó el cliente
        if es_pago_efectivo:
            efectivo = total
        else:
            voucher_tc = total
        
        max_orden = db.session.query(func.max(DetalleReporteVenta.orden)).filter_by(reporte_id=id).scalar() or 0
        
        detalle = DetalleReporteVenta(
            reporte_id=id,
            papeleta_id=papeleta_id,
            clave_aerolinea=clave_aerolinea,
            num_boletos=1,
            reserva=papeleta.clave_reserva or papeleta.clave_sabre or '',
            num_recibo='',
            num_papeleta=papeleta.folio,
            monto_volaris=monto_volaris,
            monto_vivaerobus=monto_vivaerobus,
            monto_compra_tc=monto_compra_tc,
            cargo_expedicion=float(papeleta.cargo or 0),
            cargo_315=float(papeleta.diez_porciento or 0),
            voucher_tc=voucher_tc,
            efectivo=efectivo,
            pago_directo_tc=pago_directo_tc,
            total_linea=total,
            orden=max_orden + 1
        )
        
        db.session.add(detalle)
        papeleta.reporte_venta_id = id
        
        # Recalcular totales del reporte
        _recalcular_totales_reporte(reporte)
        db.session.commit()
        
        db.session.refresh(reporte)
        
        return jsonify({
            'success': True,
            'detalle_id': detalle.id,
            'totales': {
                'total_general': float(reporte.total_general or 0),
                'total_efectivo': float(reporte.total_efectivo or 0),
                'total_voucher_tc': float(reporte.total_voucher_tc or 0),
                'total_boletos': reporte.total_boletos,
                'total_recibos': reporte.total_recibos
            }
        })
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500

@main.route('/api/reporte-venta/detalle/<int:detalle_id>', methods=['PUT', 'DELETE'])
@login_required
def modificar_detalle_reporte(detalle_id):
    """Modificar o eliminar línea del reporte"""
    detalle = DetalleReporteVenta.query.get_or_404(detalle_id)
    reporte = detalle.reporte
    
    if not reporte.puede_editar:
        return jsonify({'success': False, 'error': 'Reporte no editable'}), 400
    
    if request.method == 'DELETE':
        try:
            # Liberar papeleta si estaba asociada
            if detalle.papeleta_id:
                papeleta = Papeleta.query.get(detalle.papeleta_id)
                if papeleta:
                    papeleta.reporte_venta_id = None
            
            db.session.delete(detalle)
            
            # Recalcular totales del reporte
            _recalcular_totales_reporte(reporte)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'totales': {
                    'total_general': float(reporte.total_general or 0),
                    'total_efectivo': float(reporte.total_efectivo or 0),
                    'total_voucher_tc': float(reporte.total_voucher_tc or 0),
                    'total_boletos': reporte.total_boletos,
                    'total_recibos': reporte.total_recibos
                }
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500
    
    elif request.method == 'PUT':
        data = request.get_json()
        try:
            detalle.clave_aerolinea = data.get('clave_aerolinea', detalle.clave_aerolinea)
            detalle.num_boletos = int(data.get('num_boletos', detalle.num_boletos))
            detalle.reserva = data.get('reserva', detalle.reserva)
            detalle.num_recibo = data.get('num_recibo', detalle.num_recibo)
            detalle.monto_bsp = float(data.get('monto_bsp', detalle.monto_bsp))
            detalle.monto_volaris = float(data.get('monto_volaris', detalle.monto_volaris))
            detalle.monto_vivaerobus = float(data.get('monto_vivaerobus', detalle.monto_vivaerobus))
            detalle.monto_compra_tc = float(data.get('monto_compra_tc', detalle.monto_compra_tc))
            detalle.cargo_expedicion = float(data.get('cargo_expedicion', detalle.cargo_expedicion))
            detalle.cargo_315 = float(data.get('cargo_315', detalle.cargo_315))
            detalle.monto_seguros = float(data.get('monto_seguros', detalle.monto_seguros))
            detalle.monto_hoteles_paquetes = float(data.get('monto_hoteles_paquetes', detalle.monto_hoteles_paquetes))
            detalle.monto_transporte_terrestre = float(data.get('monto_transporte_terrestre', detalle.monto_transporte_terrestre))
            detalle.pago_directo_tc = float(data.get('pago_directo_tc', detalle.pago_directo_tc))
            detalle.voucher_tc = float(data.get('voucher_tc', detalle.voucher_tc))
            detalle.efectivo = float(data.get('efectivo', detalle.efectivo))
            detalle.total_linea = float(data.get('total_linea', detalle.total_linea))
            
            # Recalcular totales del reporte
            _recalcular_totales_reporte(reporte)
            db.session.commit()
            
            return jsonify({
                'success': True,
                'totales': {
                    'total_general': float(reporte.total_general or 0),
                    'total_efectivo': float(reporte.total_efectivo or 0),
                    'total_voucher_tc': float(reporte.total_voucher_tc or 0),
                    'total_boletos': reporte.total_boletos,
                    'total_recibos': reporte.total_recibos
                }
            })
        except Exception as e:
            db.session.rollback()
            return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/reporte-venta/<int:id>/totales')
@login_required
def obtener_totales_reporte(id):
    """Obtener totales actualizados del reporte"""
    reporte = ReporteVenta.query.get_or_404(id)
    
    # Contar boletos por aerolínea
    boletos_aerolinea = db.session.query(
        DetalleReporteVenta.clave_aerolinea,
        func.sum(DetalleReporteVenta.num_boletos)
    ).filter_by(reporte_id=id).group_by(DetalleReporteVenta.clave_aerolinea).all()
    
    return jsonify({
        'success': True,
        'totales': {
            'total_bsp': float(reporte.total_bsp or 0),
            'total_volaris': float(reporte.total_volaris or 0),
            'total_vivaerobus': float(reporte.total_vivaerobus or 0),
            'total_compra_tc': float(reporte.total_compra_tc or 0),
            'total_cargo_expedicion': float(reporte.total_cargo_expedicion or 0),
            'total_cargo_315': float(reporte.total_cargo_315 or 0),
            'total_seguros': float(reporte.total_seguros or 0),
            'total_hoteles_paquetes': float(reporte.total_hoteles_paquetes or 0),
            'total_transporte_terrestre': float(reporte.total_transporte_terrestre or 0),
            'total_pago_directo_tc': float(reporte.total_pago_directo_tc or 0),
            'total_voucher_tc': float(reporte.total_voucher_tc or 0),
            'total_efectivo': float(reporte.total_efectivo or 0),
            'total_general': float(reporte.total_general or 0),
            'total_boletos': reporte.total_boletos,
            'total_recibos': reporte.total_recibos
        },
        'boletos_aerolinea': {k: int(v) for k, v in boletos_aerolinea if k}
    })




# ============================================================
# Decoradores de permisos
# ============================================================

def admin_required(f):
    """Solo administradores pueden acceder"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.es_admin():
            flash('No tienes permisos para esta acción', 'error')
            return redirect(url_for('main.lista_entregas'))
        return f(*args, **kwargs)
    return decorated_function


def director_required(f):
    """Solo directores pueden acceder"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not current_user.is_authenticated:
            return redirect(url_for('main.login'))
        # Verificar si es director (ajustar según tu modelo)
        if not (current_user.es_admin() or getattr(current_user, 'es_director', lambda: False)()):
            flash('Solo la dirección puede realizar esta acción', 'error')
            return redirect(url_for('main.lista_entregas'))
        return f(*args, **kwargs)
    return decorated_function


# ============================================================
# VISTAS PRINCIPALES
# ============================================================

@main.route('/entregas')
@login_required
def lista_entregas():
    """Lista de entregas según rol del usuario"""
    filtro_estatus = request.args.get('estatus', '')
    filtro_fecha = request.args.get('fecha', '')
    
    # Base query
    query = EntregaCorte.query
    
    # Filtrar por rol
    if current_user.es_admin():
        # Admin ve todo
        pass
    else:
        # Agente solo ve sus entregas
        query = query.filter(EntregaCorte.agente_id == current_user.id)
    
    # Filtros adicionales
    if filtro_estatus:
        query = query.filter(EntregaCorte.estatus == filtro_estatus)
    
    if filtro_fecha:
        try:
            fecha = datetime.strptime(filtro_fecha, '%Y-%m-%d').date()
            query = query.filter(EntregaCorte.fecha == fecha)
        except:
            pass
    
    entregas = query.order_by(EntregaCorte.fecha.desc(), EntregaCorte.id.desc()).all()
    
    # Contadores para dashboard
    pendientes = EntregaCorte.query.filter_by(estatus='pendiente').count()
    en_custodia = EntregaCorte.query.filter_by(estatus='en_custodia').count()
    por_depositar = EntregaCorte.query.filter(EntregaCorte.estatus.in_(['entregado', 'retirado'])).count()
    por_revisar = EntregaCorte.query.filter_by(estatus='depositado').count()
    
    return render_template('entregas/lista.html',
        entregas=entregas,
        pendientes=pendientes,
        en_custodia=en_custodia,
        por_depositar=por_depositar,
        por_revisar=por_revisar,
        filtro_estatus=filtro_estatus,
        filtro_fecha=filtro_fecha
    )


@main.route('/entregas/nueva', methods=['GET', 'POST'])
@login_required
def nueva_entrega():
    """Crear nueva entrega de corte (Vale de Entrega)"""
    if request.method == 'POST':
        try:
            # Obtener datos del formulario
            efectivo_pesos = float(request.form.get('efectivo_pesos', 0) or 0)
            efectivo_dolares = float(request.form.get('efectivo_dolares', 0) or 0)
            tipo_cambio = float(request.form.get('tipo_cambio', 0) or 0)
            cheques = float(request.form.get('cheques', 0) or 0)
            vouchers_tc = float(request.form.get('vouchers_tc', 0) or 0)
            reporte_id = request.form.get('reporte_venta_id')
            notas = request.form.get('notas', '')
            
            # Crear entrega
            entrega = EntregaCorte(
                folio=EntregaCorte.generar_folio(),
                fecha=fecha_mexico(),
                agente_id=current_user.id,
                efectivo_pesos=efectivo_pesos,
                efectivo_dolares=efectivo_dolares,
                tipo_cambio=tipo_cambio,
                cheques=cheques,
                vouchers_tc=vouchers_tc,
                reporte_venta_id=int(reporte_id) if reporte_id else None,
                estatus='pendiente'
            )
            
            entrega.calcular_totales()
            
            db.session.add(entrega)
            db.session.flush()
            
            # Registrar historial
            entrega.registrar_historial('creado', current_user.id, None, notas)
            
            # Procesar arqueo si viene
            denominaciones = request.form.getlist('denominacion[]')
            cantidades = request.form.getlist('cantidad[]')
            tipos = request.form.getlist('tipo_denominacion[]')
            
            for i, denom in enumerate(denominaciones):
                if denom and cantidades[i]:
                    detalle = DetalleArqueo(
                        entrega_id=entrega.id,
                        tipo=tipos[i] if i < len(tipos) else 'billete',
                        denominacion=float(denom),
                        cantidad=int(cantidades[i])
                    )
                    detalle.calcular_subtotal()
                    db.session.add(detalle)
            
            db.session.commit()
            
            flash(f'Vale de entrega {entrega.folio} creado exitosamente', 'success')
            return redirect(url_for('main.ver_entrega', id=entrega.id))
            
        except Exception as e:
            db.session.rollback()
            flash(f'Error al crear entrega: {str(e)}', 'error')
    
    # GET: Mostrar formulario
    # Obtener reportes del día sin entrega asociada
    reportes_disponibles = ReporteVenta.query.filter(
        ReporteVenta.usuario_id == current_user.id,
        ReporteVenta.fecha == fecha_mexico(),
        ~ReporteVenta.id.in_(
            db.session.query(EntregaCorte.reporte_venta_id).filter(
                EntregaCorte.reporte_venta_id.isnot(None)
            )
        )
    ).all()
    
    return render_template('entregas/nueva.html',
        reportes_disponibles=reportes_disponibles,
        fecha_hoy=fecha_mexico()
    )


@main.route('/entregas/<int:id>')
@login_required
def ver_entrega(id):
    """Ver detalle de una entrega"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    # Verificar permisos
    if not current_user.es_admin() and entrega.agente_id != current_user.id:
        flash('No tienes permisos para ver esta entrega', 'error')
        return redirect(url_for('main.lista_entregas'))
    
    historial = entrega.historial.order_by(HistorialEntrega.fecha_hora.desc()).all()
    arqueo = entrega.detalles_arqueo.all()
    
    # Obtener admins para asignar receptor
    admins = Usuario.query.filter_by(activo=True).filter(
        Usuario.rol.in_(['admin', 'supervisor'])
    ).all() if current_user.es_admin() else []
    
    return render_template('entregas/ver.html',
        entrega=entrega,
        historial=historial,
        arqueo=arqueo,
        admins=admins
    )


# ============================================================
# ACCIONES DEL FLUJO
# ============================================================

@main.route('/entregas/<int:id>/entregar', methods=['POST'])
@login_required
def entregar_a_admin(id):
    """Paso 1: Agente entrega a administrativo"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    if not entrega.puede_entregar:
        flash('Esta entrega no puede ser procesada', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    receptor_id = request.form.get('receptor_id')
    notas = request.form.get('notas', '')
    
    if not receptor_id:
        flash('Debes seleccionar quién recibe', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    try:
        entrega.entregar_a_admin(int(receptor_id), notas)
        db.session.commit()
        flash('Entrega registrada exitosamente', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.ver_entrega', id=id))


@main.route('/entregas/<int:id>/confirmar-custodia', methods=['POST'])
@login_required
@admin_required
def confirmar_custodia(id):
    """Administrativo confirma custodia"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    if entrega.estatus != 'entregado':
        flash('Esta entrega no está en el estatus correcto', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    notas = request.form.get('notas', '')
    
    try:
        entrega.confirmar_custodia(current_user.id, notas)
        db.session.commit()
        flash('Custodia confirmada', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.ver_entrega', id=id))


@main.route('/entregas/<int:id>/retirar', methods=['POST'])
@login_required
@admin_required
def retirar_para_deposito(id):
    """Paso 2: Encargado retira para depositar"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    if not entrega.puede_retirar:
        flash('Esta entrega no puede ser retirada', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    notas = request.form.get('notas', '')
    
    try:
        entrega.retirar_para_deposito(current_user.id, notas)
        db.session.commit()
        flash('Retiro registrado. Procede a depositar.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.ver_entrega', id=id))


@main.route('/entregas/<int:id>/depositar', methods=['POST'])
@login_required
@admin_required
def registrar_deposito(id):
    """Paso 3: Registrar depósito bancario"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    if not entrega.puede_depositar:
        flash('Esta entrega no puede ser depositada aún', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    cuenta = request.form.get('cuenta_deposito', '')
    referencia = request.form.get('referencia_deposito', '')
    notas = request.form.get('notas', '')
    
    if not cuenta or not referencia:
        flash('Debes ingresar cuenta y referencia del depósito', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    try:
        entrega.registrar_deposito(current_user.id, cuenta, referencia, notas)
        db.session.commit()
        flash('Depósito registrado. Pendiente de revisión por dirección.', 'success')
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.ver_entrega', id=id))


@main.route('/entregas/<int:id>/revisar', methods=['POST'])
@login_required
@director_required
def revisar_entrega(id):
    """Paso 4: Director revisa y aprueba/rechaza"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    if not entrega.puede_revisar:
        flash('Esta entrega no está lista para revisión', 'error')
        return redirect(url_for('main.ver_entrega', id=id))
    
    accion = request.form.get('accion', 'aprobar')
    notas = request.form.get('notas', '')
    
    try:
        entrega.revisar_y_aprobar(
            current_user.id, 
            aprobado=(accion == 'aprobar'),
            notas=notas
        )
        db.session.commit()
        
        if accion == 'aprobar':
            flash('Entrega aprobada y cerrada', 'success')
        else:
            flash('Entrega rechazada', 'warning')
            
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.ver_entrega', id=id))


# ============================================================
# API ENDPOINTS
# ============================================================

@main.route('/api/entregas/resumen')
@login_required
def api_resumen_entregas():
    """Resumen de entregas para dashboard"""
    try:
        # Totales por estatus
        pendientes = EntregaCorte.query.filter_by(estatus='pendiente').count()
        en_custodia = EntregaCorte.query.filter(
            EntregaCorte.estatus.in_(['entregado', 'en_custodia'])
        ).count()
        por_depositar = EntregaCorte.query.filter_by(estatus='retirado').count()
        por_revisar = EntregaCorte.query.filter_by(estatus='depositado').count()
        
        # Totales monetarios pendientes
        from sqlalchemy import func
        
        total_efectivo_pendiente = db.session.query(
            func.sum(EntregaCorte.efectivo_pesos)
        ).filter(
            EntregaCorte.estatus.in_(['pendiente', 'entregado', 'en_custodia'])
        ).scalar() or 0
        
        total_dolares_pendiente = db.session.query(
            func.sum(EntregaCorte.efectivo_dolares)
        ).filter(
            EntregaCorte.estatus.in_(['pendiente', 'entregado', 'en_custodia'])
        ).scalar() or 0
        
        return jsonify({
            'success': True,
            'contadores': {
                'pendientes': pendientes,
                'en_custodia': en_custodia,
                'por_depositar': por_depositar,
                'por_revisar': por_revisar
            },
            'totales': {
                'efectivo_pendiente': float(total_efectivo_pendiente),
                'dolares_pendiente': float(total_dolares_pendiente)
            }
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/entregas/<int:id>')
@login_required
def api_detalle_entrega(id):
    """Detalle de entrega en JSON"""
    entrega = EntregaCorte.query.get_or_404(id)
    return jsonify({
        'success': True,
        'entrega': entrega.to_dict()
    })


@main.route('/api/entregas/desde-reporte/<int:reporte_id>', methods=['POST'])
@login_required
def api_crear_desde_reporte(reporte_id):
    """Crear entrega desde un reporte de ventas"""
    reporte = ReporteVenta.query.get_or_404(reporte_id)
    
    # Verificar que el reporte sea del usuario actual o sea admin
    if reporte.usuario_id != current_user.id and not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    # Verificar que no tenga ya una entrega
    existente = EntregaCorte.query.filter_by(reporte_venta_id=reporte_id).first()
    if existente:
        return jsonify({
            'success': False, 
            'error': f'Este reporte ya tiene una entrega asociada: {existente.folio}'
        }), 400
    
    try:
        entrega = crear_entrega_desde_reporte(reporte, current_user.id)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'entrega': entrega.to_dict(),
            'redirect': url_for('main.ver_entrega', id=entrega.id)
        })
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============================================================
# IMPRESIÓN
# ============================================================

@main.route('/entregas/<int:id>/imprimir')
@login_required
def imprimir_vale(id):
    """Vista de impresión del vale de entrega"""
    entrega = EntregaCorte.query.get_or_404(id)
    arqueo = entrega.detalles_arqueo.all()
    
    return render_template('entregas/vale_impresion.html',
        entrega=entrega,
        arqueo=arqueo
    )

# ============================================================
# RUTAS DE CONTROL DE PAPELETAS
# Agregar a routes.py
# ============================================================

# ------------------------------------------------------------
# VISTA PRINCIPAL DE CONTROL
# ------------------------------------------------------------

@main.route('/control-papeletas')
@login_required
def control_papeletas():
    """Dashboard de control de papeletas para administración"""
    if not current_user.es_admin():
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('main.dashboard'))
    
    from datetime import date, timedelta
    from sqlalchemy import func, case
    
    fecha_hoy = fecha_mexico()
    
    # Query base - últimos 30 días o sin reportar
    papeletas = Papeleta.query.filter(
        db.or_(
            Papeleta.fecha_venta >= fecha_hoy - timedelta(days=30),
            Papeleta.reporte_venta_id.is_(None)
        )
    ).order_by(
        Papeleta.fecha_venta.desc(),
        Papeleta.id.desc()
    ).all()
    
    # === ESTADÍSTICAS POR ESTATUS ===
    
    # Papeletas de hoy
    papeletas_hoy = len([p for p in papeletas if p.fecha_venta == fecha_hoy])
    total_hoy = sum([float(p.total or 0) for p in papeletas if p.fecha_venta == fecha_hoy])
    
    # Sin reportar Y sin facturar (realmente pendientes de acción)
    # Una papeleta facturada ya está "resuelta" aunque no tenga reporte
    sin_reportar_list = [p for p in papeletas if not p.reporte_venta_id and not p.numero_factura]
    sin_reportar = len(sin_reportar_list)
    total_sin_reportar = sum([float(p.total or 0) for p in sin_reportar_list])
    
    # Reportadas pero sin facturar (tienen reporte_venta_id pero no numero_factura)
    sin_facturar_list = [p for p in papeletas if p.reporte_venta_id and not p.numero_factura]
    sin_facturar = len(sin_facturar_list)
    total_sin_facturar = sum([float(p.total or 0) for p in sin_facturar_list])
    
    # Facturadas (tienen numero_factura)
    facturadas_list = [p for p in papeletas if p.numero_factura]
    facturadas = len(facturadas_list)
    total_facturadas = sum([float(p.total or 0) for p in facturadas_list])
    
    # Urgentes (más de 3 días sin reportar)
    urgentes = len([p for p in sin_reportar_list 
                    if p.fecha_venta and (fecha_hoy - p.fecha_venta).days > 3])
    
    # Efectivo pendiente (sin reportar y forma de pago efectivo)
    efectivo_pendiente = sum([float(p.total or 0) for p in sin_reportar_list 
                              if p.forma_pago and 'efectivo' in p.forma_pago.lower()])
    
    return render_template('control_papeletas.html',
        papeletas=papeletas,
        papeletas_hoy=papeletas_hoy,
        total_hoy=total_hoy,
        sin_reportar=sin_reportar,
        total_sin_reportar=total_sin_reportar,
        sin_facturar=sin_facturar,
        total_sin_facturar=total_sin_facturar,
        facturadas=facturadas,
        total_facturadas=total_facturadas,
        urgentes=urgentes,
        efectivo_pendiente=efectivo_pendiente,
        fecha_actual=fecha_hoy
    )


# ------------------------------------------------------------
# APIs DE CONTROL
# ------------------------------------------------------------

@main.route('/api/papeleta/<int:id>/detalle')
@login_required
def api_papeleta_detalle_control(id):
    """Obtiene detalle completo de una papeleta con historial"""
    try:
        papeleta = Papeleta.query.get_or_404(id)
        
        # Obtener historial
        historial = []
        try:
            from sqlalchemy import text
            result = db.session.execute(text("""
                SELECT accion, campo_modificado, valor_anterior, valor_nuevo,
                       u.nombre as usuario, fecha_hora
                FROM papeletas_historial h
                LEFT JOIN usuarios u ON h.usuario_id = u.id
                WHERE h.papeleta_id = :pid
                ORDER BY h.fecha_hora DESC
                LIMIT 20
            """), {'pid': id})
            
            for row in result:
                historial.append({
                    'accion': row.accion,
                    'campo': row.campo_modificado,
                    'anterior': row.valor_anterior,
                    'nuevo': row.valor_nuevo,
                    'usuario': row.usuario or 'Sistema',
                    'fecha': row.fecha_hora.strftime('%d/%m/%Y %H:%M') if row.fecha_hora else ''
                })
        except Exception as e:
            # Si no existe la tabla de historial aún
            print(f"Error historial: {e}")
            historial = [{'accion': 'creada', 'usuario': papeleta.usuario.nombre if papeleta.usuario else '-', 
                          'fecha': papeleta.created_at.strftime('%d/%m/%Y %H:%M') if papeleta.created_at else ''}]
        
        return jsonify({
            'success': True,
            'papeleta': {
                'id': papeleta.id,
                'folio': papeleta.folio,
                'fecha': papeleta.fecha_venta.strftime('%d/%m/%Y') if papeleta.fecha_venta else '',
                'agente': papeleta.usuario.nombre if papeleta.usuario else '-',
                'cliente': papeleta.facturar_a or (papeleta.empresa.nombre_empresa if papeleta.empresa else '-'),
                'concepto': papeleta.facturar_a or '-',
                'tarifa': '{:,.2f}'.format(float(papeleta.total_ticket or 0)),
                'cargo': '{:,.2f}'.format(float(papeleta.cargo or 0)),
                'total': '{:,.2f}'.format(float(papeleta.total or 0)),
                'forma_pago': papeleta.forma_pago or '-',
                'estatus': papeleta.estatus_control or 'activa',
                'clave_reserva': papeleta.clave_sabre or '-',
                'pasajero': papeleta.solicito or '-',
                'aerolinea': papeleta.aerolinea.nombre if papeleta.aerolinea else '-'
            },
            'historial': historial
        })
    except Exception as e:
        db.session.rollback()
        print(f"Error api_papeleta_detalle_control: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/papeleta/<int:id>/validar', methods=['POST'])
@login_required
def api_validar_papeleta(id):
    """Marca una papeleta como validada/revisada"""
    # Limpiar cualquier transacción pendiente
    try:
        db.session.rollback()
    except:
        pass
    
    if not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    papeleta = Papeleta.query.get_or_404(id)
    
    try:
        papeleta.estatus_control = 'revisada'
        papeleta.revisada_por_id = current_user.id
        papeleta.fecha_revision = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/papeleta/<int:id>/justificar', methods=['POST'])
@login_required
def api_justificar_papeleta(id):
    """Registra justificación de por qué una papeleta queda pendiente"""
    papeleta = Papeleta.query.get_or_404(id)
    
    # Verificar que sea del usuario o admin
    if papeleta.usuario_id != current_user.id and not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    data = request.get_json()
    justificacion = data.get('justificacion', '').strip()
    
    if not justificacion:
        return jsonify({'success': False, 'error': 'La justificación es requerida'}), 400
    
    try:
        papeleta.justificacion_pendiente = justificacion
        papeleta.fecha_justificacion = datetime.utcnow()
        
        # Registrar en historial
        try:
            from sqlalchemy import text
            db.session.execute(text("""
                INSERT INTO papeletas_historial (papeleta_id, accion, campo_modificado, valor_nuevo, usuario_id, motivo)
                VALUES (:pid, 'justificada', 'justificacion_pendiente', :just, :uid, :motivo)
            """), {'pid': id, 'just': justificacion, 'uid': current_user.id, 'motivo': justificacion})
        except:
            pass
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/papeletas/generar-alertas', methods=['POST'])
@login_required
def api_generar_alertas_papeletas():
    """Genera alertas para papeletas pendientes"""
    if not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    try:
        from sqlalchemy import text
        result = db.session.execute(text("SELECT generar_alertas_papeletas_pendientes()"))
        count = result.scalar()
        db.session.commit()
        
        return jsonify({'success': True, 'alertas': count})
    except Exception as e:
        return jsonify({'success': True, 'alertas': 0, 'nota': 'Función no disponible'})


@main.route('/api/papeleta/<int:id>/cerrar', methods=['POST'])
@login_required
def api_cerrar_papeleta(id):
    """Marca una papeleta como cerrada (proceso completo)"""
    if not current_user.es_admin():
        return jsonify({'success': False, 'error': 'No autorizado'}), 403
    
    papeleta = Papeleta.query.get_or_404(id)
    
    try:
        papeleta.estatus_control = 'cerrada'
        papeleta.cerrada_por_id = current_user.id
        papeleta.fecha_cierre = datetime.utcnow()
        
        db.session.commit()
        
        return jsonify({'success': True})
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'error': str(e)}), 500


# ------------------------------------------------------------
# NOTIFICACIONES DE PAPELETAS PENDIENTES
# ------------------------------------------------------------

@main.route('/api/mis-papeletas-pendientes')
@login_required
def api_mis_papeletas_pendientes():
    """Obtiene papeletas pendientes del usuario actual (para notificaciones)"""
    from datetime import date
    
    pendientes = Papeleta.query.filter(
        Papeleta.usuario_id == current_user.id,
        Papeleta.fecha_venta < fecha_mexico(),
        Papeleta.estatus_control.in_(['activa', None]),
        Papeleta.justificacion_pendiente.is_(None)
    ).all()
    
    return jsonify({
        'count': len(pendientes),
        'papeletas': [{
            'id': p.id,
            'folio': p.folio,
            'fecha': p.fecha_venta.strftime('%d/%m') if p.fecha_venta else '',
            'total': float(p.total or 0),
            'dias': (fecha_mexico() - p.fecha_venta).days if p.fecha_venta else 0
        } for p in pendientes]
    })


# ------------------------------------------------------------
# WIDGET PARA DASHBOARD DEL AGENTE
# ------------------------------------------------------------
# Agregar esto en la ruta dashboard() existente para mostrar alertas

"""
En tu ruta de dashboard, agrega esto para pasar datos de papeletas pendientes:

    # Papeletas pendientes del usuario
    from datetime import date
    papeletas_pendientes = Papeleta.query.filter(
        Papeleta.usuario_id == current_user.id,
        Papeleta.fecha_venta < fecha_mexico(),
        Papeleta.estatus_control.in_(['activa', None]),
        Papeleta.justificacion_pendiente.is_(None)
    ).count()

Y pásalo al template:
    papeletas_pendientes=papeletas_pendientes
"""

# ============================================================
# RECEPCIÓN DE VALES (Vista para Admin)
# ============================================================

@main.route('/entregas/recepcion')
@login_required
@admin_required
def recepcion_vales():
    """Panel para que administración reciba vales pendientes"""
    from datetime import datetime
    
    # Vales pendientes de recibir
    vales_pendientes = EntregaCorte.query.filter_by(estatus='pendiente').order_by(
        EntregaCorte.fecha_hora_creacion.asc()
    ).all()
    
    # Total pendiente
    total_pendiente = sum([float(v.total_fisico or 0) for v in vales_pendientes])
    
    # Vales en custodia del usuario actual
    en_custodia = EntregaCorte.query.filter_by(
        estatus='en_custodia',
        recibido_por_id=current_user.id
    ).count()
    
    return render_template('entregas/recepcion_vales.html',
        vales_pendientes=vales_pendientes,
        total_pendiente=total_pendiente,
        en_custodia=en_custodia,
        now=datetime.now()
    )


@main.route('/entregas/<int:id>/recibir', methods=['POST'])
@login_required
@admin_required
def recibir_vale(id):
    """Admin confirma recepción física del vale"""
    entrega = EntregaCorte.query.get_or_404(id)
    
    if entrega.estatus != 'pendiente':
        flash('Este vale ya fue procesado', 'error')
        return redirect(url_for('main.recepcion_vales'))
    
    notas = request.form.get('notas', '')
    
    try:
        # Actualizar estatus
        entrega.estatus = 'en_custodia'
        entrega.recibido_por_id = current_user.id
        entrega.fecha_recepcion = datetime.now()
        entrega.firma_receptor = 'CONFIRMADO'
        
        # Registrar en historial
        entrega.registrar_historial(
            'en_custodia', 
            current_user.id, 
            'pendiente', 
            f'Recibido por {current_user.nombre}. {notas}'.strip()
        )
        
        db.session.commit()
        flash(f'Vale {entrega.folio} recibido exitosamente', 'success')
        
    except Exception as e:
        db.session.rollback()
        flash(f'Error: {str(e)}', 'error')
    
    return redirect(url_for('main.recepcion_vales'))


# =============================================================================
# CALCULADORA DE DESGLOSE
# =============================================================================

@main.route('/calculadora-desglose')
@login_required
def calculadora_desglose():
    """Calculadora de desglose con fórmulas automáticas por empresa"""
    empresas_list = Empresa.query.filter_by(activa=True).order_by(Empresa.nombre_empresa).all()
    aerolineas_list = Aerolinea.query.order_by(Aerolinea.nombre).all()
    empresas_booking_list = EmpresaBooking.query.order_by(EmpresaBooking.nombre).all()
    return render_template('calculadora_desglose.html', 
                           empresas=empresas_list, 
                           aerolineas=aerolineas_list,
                           empresas_booking=empresas_booking_list)


@main.route('/api/empresa/<int:id>/cargos')
@login_required
def api_empresa_cargos(id):
    """Obtiene los cargos por servicio y bonificación de una empresa"""
    try:
        empresa = Empresa.query.get_or_404(id)
        
        cargo_visible = None
        cargo_oculto = None
        
        for cargo in empresa.cargos_servicio:
            if cargo.activo:
                if cargo.tipo == 'visible':
                    cargo_visible = float(cargo.monto) if cargo.monto else None
                elif cargo.tipo == 'oculto':
                    cargo_oculto = float(cargo.monto) if cargo.monto else None
        
        # Buscar bonificación - primero en empresa, luego en descuentos
        bonificacion = float(empresa.bonificacion_porcentaje) if empresa.bonificacion_porcentaje else 0
        
        # Si no hay bonificación en empresa, buscar en tabla descuentos
        if bonificacion == 0:
            for desc in empresa.descuentos:
                if desc.activo and desc.tipo == 'porcentaje':
                    # El valor viene como porcentaje (ej: 41.75), convertir a decimal
                    bonificacion = float(desc.valor) / 100 if desc.valor else 0
                    break
        
        # Determinar esquema basado en datos
        esquema = empresa.esquema_facturacion or 'cargo_servicio'
        
        # Solo cambiar a 'bonificacion' si NO es un esquema especial (issfam, desglose_especial, etc.)
        esquemas_especiales = ['issfam', 'desglose_especial', 'tarifa_fija']
        if bonificacion > 0 and esquema not in esquemas_especiales:
            esquema = 'bonificacion'
        
        return jsonify({
            'success': True,
            'cargo_visible': cargo_visible,
            'cargo_oculto': cargo_oculto,
            'esquema': esquema,
            'bonificacion': bonificacion
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/empresa/<int:id>/instrucciones')
@login_required
def api_empresa_instrucciones(id):
    """Obtiene las instrucciones de facturación de una empresa"""
    try:
        empresa = Empresa.query.get_or_404(id)
        
        return jsonify({
            'success': True,
            'instrucciones_cotizacion': empresa.instrucciones_cotizacion,
            'notas_emision': empresa.notas_emision,
            'personas_autorizadas': empresa.personas_autorizadas,
            'encargado_cuenta': empresa.encargado_cuenta
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/api/empresa/<int:id>/esquema')
@login_required
def api_empresa_esquema(id):
    """Obtiene el esquema de facturación completo de una empresa"""
    try:
        empresa = Empresa.query.get_or_404(id)
        
        tarifas_fijas = []
        for tarifa in empresa.tarifas_fijas:
            if tarifa.activa:
                tarifas_fijas.append({
                    'ruta_origen': tarifa.ruta_origen,
                    'ruta_destino': tarifa.ruta_destino,
                    'monto': float(tarifa.monto) if tarifa.monto else 0,
                    'tipo_viaje': tarifa.tipo_viaje,
                    'es_default': tarifa.es_ruta_default
                })
        
        descuentos = []
        for desc in empresa.descuentos:
            if desc.activo:
                descuentos.append({
                    'tipo': desc.tipo,
                    'valor': float(desc.valor) if desc.valor else 0,
                    'aplica_sobre': desc.aplica_sobre,
                    'descripcion': desc.descripcion,
                    'incluye_iva': desc.incluye_iva
                })
        
        return jsonify({
            'success': True,
            'empresa': {
                'id': empresa.id,
                'nombre': empresa.nombre_empresa,
                'esquema_facturacion': empresa.esquema_facturacion,
                'bonificacion_porcentaje': float(empresa.bonificacion_porcentaje) if empresa.bonificacion_porcentaje else 0,
                'bonificacion_aplica_sobre': empresa.bonificacion_aplica_sobre,
                'requiere_desglose': empresa.requiere_desglose,
                'tipo_desglose': empresa.tipo_desglose,
                'instrucciones_cotizacion': empresa.instrucciones_cotizacion,
                'notas_emision': empresa.notas_emision
            },
            'tarifas_fijas': tarifas_fijas,
            'descuentos': descuentos
        })
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@main.route('/desgloses/calculadora/guardar', methods=['POST'])
@login_required
def guardar_desglose_calculadora():
    """Guarda un desglose desde la calculadora"""
    from flask import current_app
    from werkzeug.utils import secure_filename
    import sys
    import os
    
    # Debug: imprimir datos recibidos
    current_app.logger.info("=== DATOS DEL FORMULARIO ===")
    print("=== DATOS DEL FORMULARIO ===", file=sys.stderr)
    for key, value in request.form.items():
        current_app.logger.info(f"{key}: {value}")
        print(f"{key}: {value}", file=sys.stderr)
    print("============================", file=sys.stderr)
    
    try:
        ultimo_folio = db.session.query(func.max(Desglose.folio)).scalar() or 0
        nuevo_folio = ultimo_folio + 1
        print(f"Nuevo folio: {nuevo_folio}", file=sys.stderr)
        
        # Validar campos requeridos
        empresa_id = request.form.get('empresa_id')
        aerolinea_id = request.form.get('aerolinea_id')
        clave_reserva = request.form.get('clave_reserva', '').strip()
        
        if not empresa_id:
            return jsonify({'success': False, 'error': 'Empresa es requerida'}), 400
        if not aerolinea_id:
            return jsonify({'success': False, 'error': 'Aerolínea es requerida'}), 400
        if not clave_reserva:
            return jsonify({'success': False, 'error': 'Clave de reservación es requerida'}), 400
        
        # Obtener empresa_booking_id automáticamente (usar el primero disponible)
        empresa_booking = EmpresaBooking.query.first()
        if not empresa_booking:
            return jsonify({'success': False, 'error': 'No hay Empresa Booking configurada'}), 400
        
        # Procesar archivo de boleto BSP si se subió
        archivo_boleto_nombre = None
        print(f"=== DEBUG ARCHIVO ===", file=sys.stderr)
        print(f"request.files keys: {list(request.files.keys())}", file=sys.stderr)
        if 'archivo_boleto' in request.files:
            archivo = request.files['archivo_boleto']
            print(f"archivo.filename: '{archivo.filename}'", file=sys.stderr)
            print(f"archivo.content_length: {archivo.content_length}", file=sys.stderr)
            if archivo and archivo.filename:
                # Validar extensión
                ALLOWED_EXTENSIONS = {'pdf', 'png', 'jpg', 'jpeg'}
                extension = archivo.filename.rsplit('.', 1)[1].lower() if '.' in archivo.filename else ''
                
                if extension in ALLOWED_EXTENSIONS:
                    # Crear nombre único para el archivo
                    archivo_boleto_nombre = f"bsp_{nuevo_folio}_{datetime.now().strftime('%Y%m%d%H%M%S')}.{extension}"
                    archivo_boleto_nombre = secure_filename(archivo_boleto_nombre)
                    
                    # Ruta de guardado
                    upload_folder = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'static', 'uploads', 'boletos')
                    os.makedirs(upload_folder, exist_ok=True)
                    
                    ruta_completa = os.path.join(upload_folder, archivo_boleto_nombre)
                    archivo.save(ruta_completa)
                    print(f"Archivo boleto guardado: {archivo_boleto_nombre}", file=sys.stderr)
                    print(f"Ruta completa: {ruta_completa}", file=sys.stderr)
                    print(f"Existe: {os.path.exists(ruta_completa)}", file=sys.stderr)
                else:
                    print(f"Extension no permitida: '{extension}'", file=sys.stderr)
        else:
            print(f"'archivo_boleto' NO está en request.files", file=sys.stderr)
        
        print(f"archivo_boleto_nombre final: {archivo_boleto_nombre}", file=sys.stderr)
        
        print(f"Creando desglose...", file=sys.stderr)
        
        nuevo = Desglose(
            folio=nuevo_folio,
            empresa_id=int(empresa_id),
            aerolinea_id=int(aerolinea_id),
            empresa_booking_id=empresa_booking.id,
            tarifa_base=float(request.form.get('tarifa_base') or 0),
            iva=float(request.form.get('iva') or 0),
            tua=float(request.form.get('tua') or 0),
            yr=float(request.form.get('yr') or 0),
            otros_cargos=float(request.form.get('otros_cargos') or 0),
            cargo_por_servicio=float(request.form.get('cargo_por_servicio') or 0),
            total=float(request.form.get('total') or 0),
            clave_reserva=clave_reserva,
            clave_sabre=request.form.get('clave_sabre') or None,
            numero_boleto=request.form.get('numero_boleto', '').strip() or None,
            pasajero_nombre=request.form.get('pasajero_nombre') or None,
            ruta=request.form.get('ruta') or None,
            archivo_boleto=archivo_boleto_nombre,
            usuario_id=current_user.id,
            sucursal_id=current_user.sucursal_id,
            estatus='pendiente',
            fecha_emision=fecha_mexico()
        )
        
        print(f"Agregando a sesión...", file=sys.stderr)
        db.session.add(nuevo)
        print(f"Haciendo commit...", file=sys.stderr)
        db.session.commit()
        print(f"¡Éxito! Folio: {nuevo_folio}", file=sys.stderr)
        
        return jsonify({
            'success': True,
            'folio': nuevo_folio,
            'message': f'Desglose {nuevo_folio} guardado correctamente'
        })
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_detail = traceback.format_exc()
        print(f"ERROR: {str(e)}", file=sys.stderr)
        print(f"DETALLE: {error_detail}", file=sys.stderr)
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500# ============================================
# MÓDULO DE BOLETOS - Listado y Conciliación BSP
# Agregar estas rutas al final de routes.py
# ============================================

# ============================================================
# LISTADO DE BOLETOS (DESGLOSES)
# ============================================================

@main.route('/boletos')
@login_required
def listado_boletos():
    """Listado de todos los boletos (desgloses) con filtros"""
    from datetime import date, timedelta
    
    fecha_hoy = fecha_mexico()
    
    # Parámetros de filtro
    aerolinea_id = request.args.get('aerolinea_id', type=int)
    estatus = request.args.get('estatus', '')  # pendiente, emitido, conciliado
    conciliada = request.args.get('conciliada', '')  # si, no
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    buscar = request.args.get('buscar', '').strip()
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 50
    
    # Query base - solo aerolíneas BSP
    query = Desglose.query.join(Aerolinea).filter(Aerolinea.es_bsp == True)
    
    # Filtros
    if aerolinea_id:
        query = query.filter(Desglose.aerolinea_id == aerolinea_id)
    
    if estatus:
        query = query.filter(Desglose.estatus == estatus)
    
    if conciliada == 'si':
        query = query.filter(Desglose.conciliada == True)
    elif conciliada == 'no':
        query = query.filter(db.or_(Desglose.conciliada == False, Desglose.conciliada.is_(None)))
    
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(Desglose.fecha_emision >= fd)
        except:
            pass
    
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(Desglose.fecha_emision <= fh)
        except:
            pass
    
    if buscar:
        query = query.filter(db.or_(
            Desglose.numero_boleto.ilike(f'%{buscar}%'),
            Desglose.pasajero_nombre.ilike(f'%{buscar}%'),
            Desglose.clave_reserva.ilike(f'%{buscar}%'),
            Desglose.ruta.ilike(f'%{buscar}%')
        ))
    
    # Ordenar y paginar
    query = query.order_by(Desglose.fecha_emision.desc(), Desglose.folio.desc())
    
    total = query.count()
    boletos = query.offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    total_paginas = (total + por_pagina - 1) // por_pagina
    
    # Stats (solo BSP)
    total_boletos = Desglose.query.join(Aerolinea).filter(Aerolinea.es_bsp == True).count()
    total_conciliados = Desglose.query.join(Aerolinea).filter(Aerolinea.es_bsp == True, Desglose.conciliada == True).count()
    total_sin_conciliar = total_boletos - total_conciliados
    
    # Aerolíneas para filtro
    aerolineas = Aerolinea.query.filter_by(activa=True, es_bsp=True).order_by(Aerolinea.nombre).all()
    
    return render_template('listado_boletos.html',
        boletos=boletos,
        aerolineas=aerolineas,
        total=total,
        total_boletos=total_boletos,
        total_conciliados=total_conciliados,
        total_sin_conciliar=total_sin_conciliar,
        pagina=pagina,
        total_paginas=total_paginas,
        # Filtros actuales
        filtro_aerolinea=aerolinea_id,
        filtro_estatus=estatus,
        filtro_conciliada=conciliada,
        filtro_fecha_desde=fecha_desde,
        filtro_fecha_hasta=fecha_hasta,
        filtro_buscar=buscar
    )


@main.route('/boletos/conciliar', methods=['POST'])
@login_required
def conciliar_boleto():
    """Conciliar un boleto individual"""
    folio = request.form.get('folio', type=int)
    if not folio:
        flash('Folio no válido', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    desglose = Desglose.query.get(folio)
    if not desglose:
        flash('Desglose no encontrado', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    desglose.conciliada = True
    desglose.fecha_conciliacion = fecha_mexico()
    desglose.conciliada_por_id = current_user.id
    db.session.commit()
    
    flash(f'Boleto {desglose.numero_boleto or desglose.folio} conciliado correctamente', 'success')
    return redirect(request.referrer or url_for('main.listado_boletos'))


@main.route('/boletos/desconciliar', methods=['POST'])
@login_required
def desconciliar_boleto():
    """Quitar conciliación de un boleto"""
    folio = request.form.get('folio', type=int)
    if not folio:
        flash('Folio no válido', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    desglose = Desglose.query.get(folio)
    if not desglose:
        flash('Desglose no encontrado', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    desglose.conciliada = False
    desglose.fecha_conciliacion = None
    desglose.conciliada_por_id = None
    desglose.periodo_bsp = None
    db.session.commit()
    
    flash(f'Conciliación removida para boleto {desglose.numero_boleto or desglose.folio}', 'warning')
    return redirect(request.referrer or url_for('main.listado_boletos'))


# =============================================================================
# CONCILIACIÓN BSP - Parser PDF + Lógica de conciliación
# Reemplaza: conciliar_bsp() + _parsear_bsp_txt()
# Requiere: pip install pdfplumber
# =============================================================================


@main.route('/boletos/conciliar-bsp', methods=['POST'])
@login_required
def conciliar_bsp():
    """Subir archivo BSP (PDF o TXT) y conciliar automáticamente"""
    import tempfile
    import os
    
    archivo = request.files.get('archivo_bsp')
    if not archivo:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    filename = archivo.filename.lower()
    
    # Determinar tipo de archivo y parsear
    if filename.endswith('.pdf'):
        # Guardar temporalmente para pdfplumber
        tmp = tempfile.NamedTemporaryFile(delete=False, suffix='.pdf')
        archivo.save(tmp.name)
        tmp.close()
        try:
            resultado = _parsear_bsp_pdf(tmp.name)
        finally:
            os.unlink(tmp.name)
    elif filename.endswith('.txt') or filename.endswith('.csv'):
        resultado = _parsear_bsp_txt(archivo)
    else:
        flash('Formato no soportado. Sube el archivo PDF o TXT del FCAGBILLDET', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    if not resultado['documentos']:
        flash('No se encontraron documentos en el archivo', 'error')
        return redirect(url_for('main.listado_boletos'))
    
    # Filtrar CANX (cancelaciones) — no se concilian
    docs_a_conciliar = [d for d in resultado['documentos'] if d['trnc'] != 'CANX']
    docs_canx = [d for d in resultado['documentos'] if d['trnc'] == 'CANX']
    
    # Cruzar con desgloses
    encontrados = 0
    no_encontrados = []
    ya_conciliados = 0
    conciliados_ahora = 0
    conciliados_list = []
    ya_conciliados_list = []
    periodo = resultado.get('periodo', '')
    
    for doc in docs_a_conciliar:
        numero_completo = doc['numero_completo']  # Ej: "1395463633094"
        
        # Búsqueda principal: numero_boleto exacto
        desglose = Desglose.query.filter(
            Desglose.numero_boleto == numero_completo
        ).first()
        
        # Fallback: solo document_number sin prefijo
        if not desglose:
            desglose = Desglose.query.filter(
                Desglose.numero_boleto == doc['document_number']
            ).first()
        
        # Fallback: con guión
        if not desglose:
            numero_con_guion = doc['airline_code'] + '-' + doc['document_number']
            desglose = Desglose.query.filter(
                Desglose.numero_boleto == numero_con_guion
            ).first()
        
        if desglose:
            encontrados += 1
            doc['folio_desglose'] = desglose.folio
            doc['pasajero'] = desglose.pasajero_nombre or ''
            if desglose.conciliada:
                ya_conciliados += 1
                ya_conciliados_list.append(doc)
            else:
                desglose.conciliada = True
                desglose.fecha_conciliacion = fecha_mexico()
                desglose.conciliada_por_id = current_user.id
                desglose.periodo_bsp = periodo
                conciliados_ahora += 1
                conciliados_list.append(doc)
        else:
            no_encontrados.append(doc)
    
    db.session.commit()
    
    # Guardar resultados en sesión
    from flask import session
    session['bsp_resultado'] = {
        'total_bsp': len(resultado['documentos']),
        'total_a_conciliar': len(docs_a_conciliar),
        'cancelaciones': len(docs_canx),
        'encontrados': encontrados,
        'ya_conciliados': ya_conciliados,
        'conciliados_ahora': conciliados_ahora,
        'conciliados_list': conciliados_list,
        'ya_conciliados_list': ya_conciliados_list,
        'no_encontrados': no_encontrados,
        'periodo': periodo,
        'por_tipo': {
            'TKTT': sum(1 for d in docs_a_conciliar if d['trnc'] == 'TKTT'),
            'EMDS': sum(1 for d in docs_a_conciliar if d['trnc'] == 'EMDS'),
            'EMDA': sum(1 for d in docs_a_conciliar if d['trnc'] == 'EMDA'),
        },
        'revisados': sum(1 for d in docs_a_conciliar if d.get('es_revisado'))
    }
    
    flash(
        f'Conciliación BSP completada: {conciliados_ahora} conciliados, '
        f'{ya_conciliados} ya estaban conciliados, '
        f'{len(no_encontrados)} no encontrados',
        'success' if not no_encontrados else 'warning'
    )
    
    return redirect(url_for('main.resultado_conciliacion_bsp'))


@main.route('/boletos/resultado-conciliacion')
@login_required
def resultado_conciliacion_bsp():
    """Muestra resultado de la conciliación BSP"""
    from flask import session
    resultado = session.get('bsp_resultado', {})
    if not resultado:
        flash('No hay resultados de conciliación disponibles', 'info')
        return redirect(url_for('main.listado_boletos'))
    return render_template('resultado_conciliacion_bsp.html', resultado=resultado)


def _parsear_bsp_pdf(filepath):
    """Parsear archivo FCAGBILLDET en formato PDF"""
    import pdfplumber
    import re
    
    documentos = []
    periodo = ''
    ultimo_doc = None
    
    # Regex para línea de documento
    doc_pattern = re.compile(
        r'^(\d{3})\s+'                    # CIA (3 dígitos)
        r'(TKTT|EMDS|EMDA|CANX)\s+'       # TRNC
        r'(\d{7,10})\s+'                   # NO. DOCUMENTO
        r'(\d{2}[A-Z]{3}\d{2})\s+'        # FECHA EMISION
        r'(\S+)\s+'                        # CPN
    )
    
    rtdn_pattern = re.compile(r'^\+RTDN:\s*(\d{7,10})')
    periodo_pattern = re.compile(r'Billing Period:(\d{6})')
    
    # Líneas a ignorar
    skip_prefixes = (
        'ESAC:', 'ISSUES TOTAL', 'FCAGBILLDET', 'SCOPE', 'CIA ', 'NO. DE',
        '***', 'GRAND TOTAL', 'DOMESTIC TOTAL', 'INTERNATIONAL TOTAL',
        'ISSUES', 'CA ', 'CC ', '**'
    )
    
    with pdfplumber.open(filepath) as pdf:
        for page in pdf.pages:
            text = page.extract_text()
            if not text:
                continue
            
            for line in text.split('\n'):
                line = line.strip()
                
                # Extraer periodo
                pm = periodo_pattern.search(line)
                if pm:
                    periodo = pm.group(1)
                    continue
                
                # Detectar +RTDN (revisado/cambio)
                rm = rtdn_pattern.match(line)
                if rm and ultimo_doc:
                    ultimo_doc['rtdn'] = rm.group(1)
                    ultimo_doc['es_revisado'] = True
                    continue
                
                # Saltar líneas no relevantes
                if any(line.startswith(p) for p in skip_prefixes) or 'Page :' in line:
                    continue
                
                # Parsear documento
                dm = doc_pattern.match(line)
                if not dm:
                    continue
                
                airline_code = dm.group(1)
                trnc = dm.group(2)
                document_number = dm.group(3)
                fecha_emision = dm.group(4)
                
                # Extraer montos del resto de la línea
                rest = line[dm.end():]
                montos = re.findall(r'-?[\d,]+\.\d{2}', rest)
                
                transaction_amount = float(montos[0].replace(',', '')) if montos else 0.0
                tarifa = float(montos[1].replace(',', '')) if len(montos) >= 2 else 0.0
                balance_payable = float(montos[-1].replace(',', '')) if montos else 0.0
                
                # FOP y Scope
                fop = 'CA'
                fop_match = re.search(r'\s[ID]\s+(CA|CC)\s', line)
                if fop_match:
                    fop = fop_match.group(1)
                
                scope = 'Doméstico'
                scope_match = re.search(r'\s([ID])\s+(?:CA|CC)\s', line)
                if scope_match and scope_match.group(1) == 'I':
                    scope = 'Internacional'
                
                doc = {
                    'airline_code': airline_code,
                    'trnc': trnc,
                    'document_number': document_number,
                    'numero_completo': airline_code + document_number,
                    'fecha_emision': fecha_emision,
                    'fop': fop,
                    'scope': scope,
                    'transaction_amount': transaction_amount,
                    'tarifa': tarifa,
                    'balance_payable': balance_payable,
                    'es_revisado': False,
                    'rtdn': None
                }
                
                documentos.append(doc)
                ultimo_doc = doc
    
    return {
        'periodo': periodo,
        'documentos': documentos
    }



def _parsear_bsp_txt(archivo):
    """Parsear archivo FCAGBILLDETSIMP (.txt/.csv)"""
    import csv
    import io
    
    contenido = archivo.read().decode('utf-8', errors='ignore')
    lineas = contenido.strip().split('\n')
    
    resultado = {
        'periodo': '',
        'documentos': []
    }
    
    for linea in lineas:
        linea = linea.strip()
        
        # Extraer Agent Code (primera línea)
        if linea.startswith('Agent Code'):
            continue
        
        # Saltar header
        if linea.startswith('Airline Code,'):
            continue
        
        # Parsear línea de datos
        if not linea or linea.startswith('#'):
            continue
        
        campos = linea.split(',')
        if len(campos) < 8:
            continue
        
        airline_code = campos[0].strip().zfill(3)  # Asegurar 3 dígitos: 001, 006, etc.
        trnc = campos[1].strip()
        document_number = campos[2].strip()
        fop = campos[3].strip()
        
        try:
            transaction_amount = float(campos[4].strip())
        except:
            transaction_amount = 0.0
        
        try:
            tax_on_commission = float(campos[5].strip())
        except:
            tax_on_commission = 0.0
        
        try:
            balance_payable = float(campos[6].strip())
        except:
            balance_payable = 0.0
        
        currency = campos[7].strip() if len(campos) > 7 else 'MXN'
        
        # Solo incluir TKTT (boletos) y EMD (servicios), ignorar CANX (cancelaciones con monto 0)
        if trnc == 'CANX':
            continue
        
        resultado['documentos'].append({
            'airline_code': airline_code,
            'trnc': trnc,
            'document_number': document_number,
            'fop': fop,
            'transaction_amount': transaction_amount,
            'tax_on_commission': tax_on_commission,
            'balance_payable': balance_payable,
            'currency': currency
        })
    
    # Extraer periodo del nombre del archivo si posible
    # Formato: FCAGBILLDETSIMP_MX_86506696_260104.txt
    try:
        nombre = archivo.filename
        partes = nombre.replace('.txt', '').replace('.csv', '').split('_')
        for p in partes:
            if len(p) == 6 and p.isdigit():
                resultado['periodo'] = p
                break
    except:
        pass
    
    return resultado


# ============================================
# MÓDULO VOLARIS - Listado y Conciliación
# ============================================

@main.route('/papeletas-volaris')
@login_required
def listado_papeletas_volaris():
    """Listado de papeletas de Volaris con filtros"""
    
    fecha_hoy = fecha_mexico()
    
    conciliada = request.args.get('conciliada', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    buscar = request.args.get('buscar', '').strip()
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 50
    
    # Solo Volaris — únicamente papeletas ya revisadas:
    #   - Crédito: factura aprobada (estatus_facturacion = 'aprobada')
    #   - Mostrador: incluida en reporte de ventas (reporte_venta_id IS NOT NULL)
    query = Papeleta.query.options(
        db.joinedload(Papeleta.usuario)
    ).join(Aerolinea).filter(
        Aerolinea.nombre.ilike('%volaris%'),
        db.or_(
            Papeleta.estatus_facturacion == 'aprobada',
            Papeleta.reporte_venta_id.isnot(None)
        )
    )
    
    if conciliada == 'si':
        query = query.filter(Papeleta.conciliada == True)
    elif conciliada == 'no':
        query = query.filter(db.or_(Papeleta.conciliada == False, Papeleta.conciliada.is_(None)))
    
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(Papeleta.fecha_venta >= fd)
        except:
            pass
    
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(Papeleta.fecha_venta <= fh)
        except:
            pass
    
    if buscar:
        query = query.filter(db.or_(
            Papeleta.clave_sabre.ilike(f'%{buscar}%'),
            Papeleta.clave_reserva.ilike(f'%{buscar}%'),
            Papeleta.pasajero_nombre.ilike(f'%{buscar}%'),
            Papeleta.facturar_a.ilike(f'%{buscar}%'),
            Papeleta.solicito.ilike(f'%{buscar}%'),
            Papeleta.folio.ilike(f'%{buscar}%')
        ))
    
    query = query.order_by(Papeleta.fecha_venta.desc(), Papeleta.id.desc())
    
    total = query.count()
    papeletas = query.offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    total_paginas = (total + por_pagina - 1) // por_pagina
    
    base_query = Papeleta.query.join(Aerolinea).filter(
        Aerolinea.nombre.ilike('%volaris%'),
        db.or_(
            Papeleta.estatus_facturacion == 'aprobada',
            Papeleta.reporte_venta_id.isnot(None)
        )
    )
    total_papeletas = base_query.count()
    total_conciliadas = base_query.filter(Papeleta.conciliada == True).count()
    total_sin_conciliar = total_papeletas - total_conciliadas
    
    return render_template('listado_papeletas_volaris.html',
        papeletas=papeletas,
        total=total,
        total_papeletas=total_papeletas,
        total_conciliadas=total_conciliadas,
        total_sin_conciliar=total_sin_conciliar,
        pagina=pagina,
        total_paginas=total_paginas,
        filtro_conciliada=conciliada,
        filtro_fecha_desde=fecha_desde,
        filtro_fecha_hasta=fecha_hasta,
        filtro_buscar=buscar
    )


@main.route('/papeletas-volaris/conciliar', methods=['POST'])
@login_required
def conciliar_papeleta_volaris():
    """Conciliar una papeleta de Volaris manualmente"""
    papeleta_id = request.form.get('papeleta_id', type=int)
    if not papeleta_id:
        flash('ID no válido', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    papeleta = Papeleta.query.get(papeleta_id)
    if not papeleta:
        flash('Papeleta no encontrada', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    papeleta.conciliada = True
    papeleta.fecha_conciliacion = fecha_mexico()
    papeleta.conciliada_por_id = current_user.id
    db.session.commit()
    
    flash(f'Papeleta {papeleta.folio} conciliada correctamente', 'success')
    return redirect(request.referrer or url_for('main.listado_papeletas_volaris'))


@main.route('/papeletas-volaris/desconciliar', methods=['POST'])
@login_required
def desconciliar_papeleta_volaris():
    """Quitar conciliación de una papeleta de Volaris"""
    papeleta_id = request.form.get('papeleta_id', type=int)
    if not papeleta_id:
        flash('ID no válido', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    papeleta = Papeleta.query.get(papeleta_id)
    if not papeleta:
        flash('Papeleta no encontrada', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    papeleta.conciliada = False
    papeleta.fecha_conciliacion = None
    papeleta.conciliada_por_id = None
    papeleta.periodo_conciliacion = None
    db.session.commit()
    
    flash(f'Conciliación removida para papeleta {papeleta.folio}', 'warning')
    return redirect(request.referrer or url_for('main.listado_papeletas_volaris'))


@main.route('/papeletas-volaris/conciliar-archivo', methods=['POST'])
@login_required
def conciliar_volaris():
    """Subir reporte de ventas de Volaris (.xlsx) y conciliar automáticamente"""
    import openpyxl
    
    archivo = request.files.get('archivo_volaris')
    if not archivo:
        flash('No se seleccionó ningún archivo', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    filename = archivo.filename.lower()
    if not (filename.endswith('.xlsx') or filename.endswith('.xls')):
        flash('Formato no soportado. Sube el archivo Excel (.xlsx) de Volaris', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    try:
        wb = openpyxl.load_workbook(archivo, data_only=True)
    except Exception as e:
        flash(f'Error al leer el archivo: {str(e)}', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    documentos = []
    periodo = ''
    
    for ws in wb.worksheets:
        titulo = ws.cell(row=1, column=1).value or ''
        if not periodo and titulo:
            periodo = titulo.replace('REPORTE DE VENTAS DEL ', '').strip()
        
        for row_num in range(2, ws.max_row + 1):
            fecha = ws.cell(row=row_num, column=1).value
            pnr = ws.cell(row=row_num, column=2).value
            agente = ws.cell(row=row_num, column=3).value
            pasajero = ws.cell(row=row_num, column=4).value
            pago = ws.cell(row=row_num, column=6).value
            
            if not pnr or not isinstance(pnr, str) or len(pnr.strip()) < 3:
                continue
            
            pnr = pnr.strip().upper()
            
            if pnr.startswith('=') or pnr == 'PNR':
                continue
            
            try:
                pago_num = float(pago) if pago else 0.0
            except (ValueError, TypeError):
                continue
            
            documentos.append({
                'fecha': fecha.strftime('%d/%m/%Y') if hasattr(fecha, 'strftime') else str(fecha or ''),
                'pnr': pnr,
                'agente': str(agente or '').strip(),
                'pasajero': str(pasajero or '').strip(),
                'pago': pago_num
            })
    
    if not documentos:
        flash('No se encontraron registros en el archivo', 'error')
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    encontrados = 0
    no_encontrados = []
    ya_conciliados = 0
    conciliados_ahora = 0
    con_diferencia_monto = []
    conciliados_list = []
    ya_conciliados_list = []
    
    for doc in documentos:
        # Buscar TODAS las papeletas con esta clave_reserva
        papeletas_found = Papeleta.query.options(
            db.joinedload(Papeleta.usuario)
        ).filter(
            db.func.upper(Papeleta.clave_reserva) == doc['pnr']
        ).order_by(Papeleta.fecha_venta, Papeleta.id).all()
        
        # Fallback: buscar por clave_sabre si no encuentra por clave_reserva
        if not papeletas_found:
            papeletas_found = Papeleta.query.options(
                db.joinedload(Papeleta.usuario)
            ).filter(
                db.func.upper(Papeleta.clave_sabre) == doc['pnr']
            ).order_by(Papeleta.fecha_venta, Papeleta.id).all()
        
        if papeletas_found:
            encontrados += 1
            
            # Sumar totales de todas las papeletas con esta clave
            monto_sistema = sum(float(p.total or 0) for p in papeletas_found)
            monto_volaris = doc['pago']
            diferencia = round(monto_sistema - monto_volaris, 2)
            
            # Datos del primer registro para mostrar
            primera = papeletas_found[0]
            folios = ', '.join(p.folio for p in papeletas_found)
            
            # Pasajero: preferir pasajero_nombre, luego solicito
            pasajero_sis = primera.pasajero_nombre or primera.solicito or primera.facturar_a or ''
            
            doc_info = {
                'pnr': doc['pnr'],
                'fecha': doc['fecha'],
                'agente_volaris': doc['agente'],
                'pasajero_volaris': doc['pasajero'],
                'monto_volaris': monto_volaris,
                'monto_sistema': monto_sistema,
                'diferencia': diferencia,
                'agente_sistema': primera.usuario.nombre if primera.usuario else '',
                'pasajero_sistema': pasajero_sis,
                'folio': folios,
                'clave_reserva': primera.clave_reserva or '',
                'num_papeletas': len(papeletas_found)
            }
            
            # Verificar si TODAS ya están conciliadas
            todas_conciliadas = all(p.conciliada for p in papeletas_found)
            alguna_conciliada = any(p.conciliada for p in papeletas_found)
            
            if todas_conciliadas:
                ya_conciliados += 1
                ya_conciliados_list.append(doc_info)
            else:
                # Conciliar las que falten
                for p in papeletas_found:
                    if not p.conciliada:
                        p.conciliada = True
                        p.fecha_conciliacion = fecha_mexico()
                        p.conciliada_por_id = current_user.id
                        p.periodo_conciliacion = periodo
                conciliados_ahora += 1
                conciliados_list.append(doc_info)
            
            if abs(diferencia) > 0.01:
                con_diferencia_monto.append(doc_info)
        else:
            no_encontrados.append(doc)
    
    db.session.commit()
    
    from flask import session
    session['volaris_resultado'] = {
        'total_reporte': len(documentos),
        'encontrados': encontrados,
        'ya_conciliados': ya_conciliados,
        'conciliados_ahora': conciliados_ahora,
        'conciliados_list': conciliados_list,
        'ya_conciliados_list': ya_conciliados_list,
        'no_encontrados': no_encontrados,
        'con_diferencia_monto': con_diferencia_monto,
        'periodo': periodo
    }
    
    flash(
        f'Conciliación Volaris completada: {conciliados_ahora} conciliados, '
        f'{ya_conciliados} ya estaban conciliados, '
        f'{len(no_encontrados)} no encontrados en el sistema',
        'success' if not no_encontrados else 'warning'
    )
    
    return redirect(url_for('main.resultado_conciliacion_volaris'))


@main.route('/papeletas-volaris/resultado')
@login_required
def resultado_conciliacion_volaris():
    """Mostrar resultado de la conciliación Volaris"""
    from flask import session
    resultado = session.pop('volaris_resultado', None)
    
    if not resultado:
        return redirect(url_for('main.listado_papeletas_volaris'))
    
    return render_template('resultado_conciliacion_volaris.html', resultado=resultado)


# ============================================
# MÓDULO VIVA AEROBUS - Listado y Conciliación Manual
# ============================================

@main.route('/papeletas-viva')
@login_required
def listado_boletos_viva():
    """Listado de papeletas de Viva Aerobus con filtros"""
    
    fecha_hoy = fecha_mexico()
    
    conciliada = request.args.get('conciliada', '')
    fecha_desde = request.args.get('fecha_desde', '')
    fecha_hasta = request.args.get('fecha_hasta', '')
    buscar = request.args.get('buscar', '').strip()
    pagina = request.args.get('pagina', 1, type=int)
    por_pagina = 50
    
    # Solo Viva Aerobus
    query = Papeleta.query.join(Aerolinea).filter(Aerolinea.nombre.ilike('%viva%'))
    
    if conciliada == 'si':
        query = query.filter(Papeleta.conciliada == True)
    elif conciliada == 'no':
        query = query.filter(db.or_(Papeleta.conciliada == False, Papeleta.conciliada.is_(None)))
    
    if fecha_desde:
        try:
            fd = datetime.strptime(fecha_desde, '%Y-%m-%d').date()
            query = query.filter(Papeleta.fecha_venta >= fd)
        except:
            pass
    
    if fecha_hasta:
        try:
            fh = datetime.strptime(fecha_hasta, '%Y-%m-%d').date()
            query = query.filter(Papeleta.fecha_venta <= fh)
        except:
            pass
    
    if buscar:
        query = query.filter(db.or_(
            Papeleta.clave_sabre.ilike(f'%{buscar}%'),
            Papeleta.facturar_a.ilike(f'%{buscar}%'),
            Papeleta.solicito.ilike(f'%{buscar}%'),
            Papeleta.folio.ilike(f'%{buscar}%')
        ))
    
    query = query.order_by(Papeleta.fecha_venta.desc(), Papeleta.id.desc())
    
    total = query.count()
    papeletas = query.offset((pagina - 1) * por_pagina).limit(por_pagina).all()
    total_paginas = (total + por_pagina - 1) // por_pagina
    
    base_query = Papeleta.query.join(Aerolinea).filter(Aerolinea.nombre.ilike('%viva%'))
    total_papeletas = base_query.count()
    total_conciliadas = base_query.filter(Papeleta.conciliada == True).count()
    total_sin_conciliar = total_papeletas - total_conciliadas
    
    return render_template('listado_boletos_viva.html',
        papeletas=papeletas,
        total=total,
        total_papeletas=total_papeletas,
        total_conciliadas=total_conciliadas,
        total_sin_conciliar=total_sin_conciliar,
        pagina=pagina,
        total_paginas=total_paginas,
        filtro_conciliada=conciliada,
        filtro_fecha_desde=fecha_desde,
        filtro_fecha_hasta=fecha_hasta,
        filtro_buscar=buscar
    )


@main.route('/papeletas-viva/conciliar', methods=['POST'])
@login_required
def conciliar_papeleta_viva():
    """Conciliar una papeleta de Viva Aerobus"""
    papeleta_id = request.form.get('papeleta_id', type=int)
    if not papeleta_id:
        flash('ID no válido', 'error')
        return redirect(url_for('main.listado_boletos_viva'))
    
    papeleta = Papeleta.query.get(papeleta_id)
    if not papeleta:
        flash('Papeleta no encontrada', 'error')
        return redirect(url_for('main.listado_boletos_viva'))
    
    papeleta.conciliada = True
    papeleta.fecha_conciliacion = fecha_mexico()
    papeleta.conciliada_por_id = current_user.id
    db.session.commit()
    
    flash(f'Papeleta {papeleta.folio} conciliada correctamente', 'success')
    return redirect(request.referrer or url_for('main.listado_boletos_viva'))


@main.route('/papeletas-viva/desconciliar', methods=['POST'])
@login_required
def desconciliar_papeleta_viva():
    """Quitar conciliación de una papeleta de Viva Aerobus"""
    papeleta_id = request.form.get('papeleta_id', type=int)
    if not papeleta_id:
        flash('ID no válido', 'error')
        return redirect(url_for('main.listado_boletos_viva'))
    
    papeleta = Papeleta.query.get(papeleta_id)
    if not papeleta:
        flash('Papeleta no encontrada', 'error')
        return redirect(url_for('main.listado_boletos_viva'))
    
    papeleta.conciliada = False
    papeleta.fecha_conciliacion = None
    papeleta.conciliada_por_id = None
    papeleta.periodo_conciliacion = None
    db.session.commit()
    
    flash(f'Conciliación removida para papeleta {papeleta.folio}', 'warning')
    return redirect(request.referrer or url_for('main.listado_boletos_viva'))

# ============================================
# MÓDULO RESTRICCIÓN DE CRÉDITO
# ============================================

@main.route('/api/empresa/<int:empresa_id>/estado-credito')
@login_required
def estado_credito_empresa(empresa_id):
    """API: Retorna el estado de crédito de una empresa (para JS en formularios)"""
    empresa = Empresa.query.get(empresa_id)
    if not empresa:
        return jsonify({'error': 'Empresa no encontrada'}), 404
    
    data = {
        'id': empresa.id,
        'nombre': empresa.nombre_empresa,
        'credito_restringido': empresa.credito_restringido or False,
        'motivo_restriccion': empresa.motivo_restriccion or '',
        'fecha_restriccion': empresa.fecha_restriccion.strftime('%d/%m/%Y') if empresa.fecha_restriccion else '',
        'fecha_probable_pago': empresa.fecha_probable_pago.strftime('%d/%m/%Y') if empresa.fecha_probable_pago else '',
        'notas_cobranza': empresa.notas_cobranza or '',
        'credito_activo': empresa.credito_activo or False,
        'limite_credito': float(empresa.limite_credito or 0),
        'credito_disponible': float(empresa.credito_disponible or 0),
        'dias_credito': empresa.dias_credito or 0
    }
    return jsonify(data)


@main.route('/empresas/<int:id>/restringir', methods=['POST'])
@login_required
def restringir_empresa(id):
    """Marcar empresa como restringida por atraso de pago"""
    if not current_user.es_admin():
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('main.empresas'))
    
    empresa = Empresa.query.get_or_404(id)
    
    empresa.credito_restringido = True
    empresa.motivo_restriccion = request.form.get('motivo_restriccion', 'Atraso en pagos mayor a 2 meses')
    empresa.fecha_restriccion = fecha_mexico()
    
    fecha_pp = request.form.get('fecha_probable_pago')
    if fecha_pp:
        try:
            empresa.fecha_probable_pago = datetime.strptime(fecha_pp, '%Y-%m-%d').date()
        except:
            pass
    
    empresa.notas_cobranza = request.form.get('notas_cobranza', '')
    empresa.restringido_por_id = current_user.id
    
    db.session.commit()
    flash(f'Crédito restringido para {empresa.nombre_empresa}', 'warning')
    return redirect(request.referrer or url_for('main.empresas'))


@main.route('/empresas/<int:id>/desrestringir', methods=['POST'])
@login_required
def desrestringir_empresa(id):
    """Quitar restricción de crédito"""
    if not current_user.es_admin():
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('main.empresas'))
    
    empresa = Empresa.query.get_or_404(id)
    
    empresa.credito_restringido = False
    empresa.motivo_restriccion = None
    empresa.fecha_restriccion = None
    empresa.fecha_probable_pago = None
    empresa.notas_cobranza = None
    empresa.restringido_por_id = None
    
    db.session.commit()
    flash(f'Crédito reactivado para {empresa.nombre_empresa}', 'success')
    return redirect(request.referrer or url_for('main.empresas'))


@main.route('/empresas/<int:id>/actualizar-contrato', methods=['POST'])
@login_required
def actualizar_contrato_empresa(id):
    """Actualizar datos del contrato"""
    if not current_user.es_admin():
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('main.empresas'))
    
    empresa = Empresa.query.get_or_404(id)
    
    empresa.numero_contrato = request.form.get('numero_contrato', '').strip() or None
    
    monto = request.form.get('monto_contrato', '').strip()
    empresa.monto_contrato = float(monto) if monto else None
    
    fi = request.form.get('fecha_inicio_contrato')
    if fi:
        try:
            empresa.fecha_inicio_contrato = datetime.strptime(fi, '%Y-%m-%d').date()
        except:
            pass
    
    ff = request.form.get('fecha_fin_contrato')
    if ff:
        try:
            empresa.fecha_fin_contrato = datetime.strptime(ff, '%Y-%m-%d').date()
        except:
            pass
    
    db.session.commit()
    flash(f'Contrato actualizado para {empresa.nombre_empresa}', 'success')
    return redirect(request.referrer or url_for('main.empresas'))


@main.route('/credito-empresas')
@login_required
def panel_credito_empresas():
    """Panel de seguimiento de crédito y contratos de empresas"""
    if not current_user.es_admin():
        flash('Acceso no autorizado', 'danger')
        return redirect(url_for('main.dashboard'))
    
    fecha_hoy = fecha_mexico()
    
    # Todas las empresas activas
    empresas = Empresa.query.filter(
        Empresa.activa == True
    ).order_by(
        Empresa.credito_restringido.desc(),  # Restringidas primero
        Empresa.nombre_empresa
    ).all()
    
    # Calcular datos para cada empresa
    empresas_data = []
    for emp in empresas:
        # Consumo total (papeletas a crédito en los últimos 6 meses)
        seis_meses = fecha_hoy - timedelta(days=180)
        consumo = db.session.query(func.sum(Papeleta.total)).filter(
            Papeleta.empresa_id == emp.id,
            Papeleta.fecha_venta >= seis_meses
        ).scalar()
        consumo = float(consumo or 0)
        
        # Facturas pendientes (sin factura o pendientes)
        facturas_pend = Papeleta.query.filter(
            Papeleta.empresa_id == emp.id,
            Papeleta.estatus_facturacion.in_(['pendiente', None]),
            Papeleta.numero_factura.is_(None)
        ).count()
        
        # Monto pendiente de facturar
        monto_pend = db.session.query(func.sum(Papeleta.total)).filter(
            Papeleta.empresa_id == emp.id,
            Papeleta.estatus_facturacion.in_(['pendiente', None]),
            Papeleta.numero_factura.is_(None)
        ).scalar()
        monto_pend = float(monto_pend or 0)
        
        # Progreso del contrato
        progreso_contrato = 0
        if emp.monto_contrato and float(emp.monto_contrato) > 0:
            progreso_contrato = min(round((consumo / float(emp.monto_contrato)) * 100, 1), 100)
        
        # Días de contrato restantes
        dias_contrato_rest = None
        if emp.fecha_fin_contrato:
            dias_contrato_rest = (emp.fecha_fin_contrato - fecha_hoy).days
        
        empresas_data.append({
            'empresa': emp,
            'consumo_6m': consumo,
            'facturas_pendientes': facturas_pend,
            'monto_pendiente': monto_pend,
            'progreso_contrato': progreso_contrato,
            'dias_contrato_restantes': dias_contrato_rest
        })
    
    # Stats
    total_empresas = len(empresas)
    restringidas = len([e for e in empresas if e.credito_restringido])
    
    return render_template('panel_credito_empresas.html',
        empresas_data=empresas_data,
        total_empresas=total_empresas,
        restringidas=restringidas,
        fecha_hoy=fecha_hoy
    )